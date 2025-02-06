import traceback

try:
    import os
    import re
    import openpyxl
    import numpy as np
    import pandas as pd
    from datetime import datetime, timedelta
    from openpyxl.styles import Alignment, PatternFill, Font

    # import selenium_tambau
    
    # Deletar o arquivo de saída após o uso
    if os.path.exists("Erro na automação.txt"):
        os.remove("Erro na automação.txt")

    caminho_saida = "outputs"

    # Função para criar uma pasta apenas se ela não existir
    def criar_pasta_se_nao_existir(caminho):
        if not os.path.exists(caminho):
            os.makedirs(caminho)
            print(f"Pasta '{caminho}' criada com sucesso.")
        else:
            print(f"Pasta '{caminho}' já existe.")

    criar_pasta_se_nao_existir(caminho_saida)


    # -------------------------- Importando bases

    # (INVOLVES) ------------------------------------------------
    justificativas = pd.read_excel("bases_involves/relatorio_de_operacao.xlsx")

    checkinweek = justificativas  # Separando o checkinweek para outros cálculos

    relatorio_colaboradores = pd.read_excel("bases_involves/Relatório_Colaboradores.xlsx")

    relatorio_de_checkin = pd.read_excel("bases_involves/relatorio_de_check_in.xlsx")

    gerencial = pd.read_excel("bases_involves/relatorio-gerencial-visitas.xlsx")

    pesquisasporcentagem_x = pd.read_excel("bases_involves/Relatório.xlsx")
    tarefas = pesquisasporcentagem_x.loc[pesquisasporcentagem_x["Status"] == "Respondida"]

    # Bases Fixas ------------------------------------------------
    pesos = pd.read_excel("bases_fixas/VALORES BASE DO CALCULO.xlsx")

    almoco = gerencial.loc[gerencial["Ponto de Venda"].str.contains("INTERVALO")]

    # Utilizando nome do cliente para nomear o output final do relatório
    nome_do_cliente = "TAMBAÚ"

    # Filtrando e renomeando colunas
    relatorio_colaboradores = relatorio_colaboradores[
        [
            "UF",
            "Nome do Colaborador",
            "Colaborador Superior",
            "Perfil de Acesso",
            "Jornada de Trabalho",
            "Equipe de Campo",
            "Nome do Pai",
        ]
    ]
    relatorio_colaboradores["almoco - SEMANAL"] = pd.to_timedelta("0 days 00:00:00")
    relatorio_colaboradores = relatorio_colaboradores.loc[relatorio_colaboradores["Nome do Colaborador"]!="TESTE (NÃO INATIVAR)"]
    
    relatorio_colaboradores = relatorio_colaboradores.loc[(relatorio_colaboradores["Perfil de Acesso"].str.contains("PROMOTOR")) & 
                                (relatorio_colaboradores["Perfil de Acesso"].str.contains("CX"))]
    # print(relatorio_colaboradores)
    relatorio_colaboradores = relatorio_colaboradores.rename(
        {"Nome do Colaborador": "Colaborador", "Nome do Pai": "jornada_semanal"}, axis=1
    )


    # Função para limpar caracteres não numéricos e truncar inteiros com mais de dois dígitos
    def clean_and_truncate(value):
        if pd.isna(value):  # Manter valores NaN
            return value
        value = str(value)  # Converter para string
        cleaned_value = re.sub(r"[^\d.]", "", value)  # Remover tudo que não é número ou ponto decimal
        try:
            # Verificar se o valor é float ou inteiro
            num_value = float(cleaned_value)
            if num_value.is_integer() and len(cleaned_value.split('.')[0]) > 2:  # Truncar inteiros com mais de 2 dígitos
                return int(str(int(num_value))[:2])  # Truncar para os dois primeiros dígitos
            return num_value  # Retornar floats sem alterações
        except ValueError:
            return None  # Valores não numéricos retornam como None/NaN

    # Aplicar a função na coluna
    relatorio_colaboradores["jornada_semanal"] = relatorio_colaboradores["jornada_semanal"].apply(clean_and_truncate)
    relatorio_colaboradores["jornada_semanal"] = relatorio_colaboradores["jornada_semanal"].astype(float)

    # Dicionário para mapeamento das jornadas e almocos
    jornada_dict = {
        44.0: {
            "dias": "0 days 08:00:00",
            "sabado": "0 days 04:00:00",
            "semanal": "0 days 44:00:00",
        },
        34.0: {
            "dias": "0 days 06:00:00",
            "sabado": "0 days 04:00:00",
            "semanal": "0 days 34:00:00",
        },
        30.0: {
            "dias": "0 days 05:00:00",
            "sabado": "0 days 05:00:00",
            "semanal": "0 days 30:00:00",
        },
        24.0: {
            "dias": "0 days 04:00:00",
            "sabado": "0 days 04:00:00",
            "semanal": "0 days 24:00:00",
        },
        12.0: {
            "dias": "0 days 02:00:00",
            "sabado": "0 days 02:00:00",
            "semanal": "0 days 12:00:00",
        },
    }

    # Função para aplicar valores de jornada e almoco
    def aplicar_jornada_almoco(df, jornada_dict):
        for jornada, valores in jornada_dict.items():
            dias_semana = [
                "SEGUNDA-FEIRA",
                "TERÇA-FEIRA",
                "QUARTA-FEIRA",
                "QUINTA-FEIRA",
                "SEXTA-FEIRA",
            ]
            # almoco_semana = ["almoco - SEG", "almoco - TER", "almoco - QUA", "almoco - QUI", "almoco - SEX"]

            df.loc[df["jornada_semanal"] == jornada, dias_semana] = pd.to_timedelta(
                valores["dias"]
            )
            df.loc[df["jornada_semanal"] == jornada, "SÁBADO"] = pd.to_timedelta(
                valores["sabado"]
            )
            df.loc[
                df["jornada_semanal"] == jornada, "Jornada de Trabalho - SEMANAL"
            ] = pd.to_timedelta(valores["semanal"])

            # df.loc[df["jornada_semanal"] == jornada, almoco_semana] = pd.to_timedelta(valores["almoco"])
            # df.loc[df["jornada_semanal"] == jornada, "almoco - SAB"] = pd.to_timedelta(valores["almoco_sab"])
            # df.loc[df["jornada_semanal"] == jornada, "almoco - SEMANAL"] = pd.to_timedelta(valores["almoco_semanal"])

    # Aplicando a função
    aplicar_jornada_almoco(relatorio_colaboradores, jornada_dict)

    # Removendo colunas desnecessárias
    jornada_diaria = relatorio_colaboradores.drop(
        columns=[
            "Perfil de Acesso",
            "Jornada de Trabalho",
            "Equipe de Campo",
            "jornada_semanal",
        ]
    )

    # ----------- #1

    # ----------- #2
    import pandas as pd

    # Função para separar data e hora
    def separa_hora_data(df, coluna):
        df["data"] = pd.to_datetime(
            df[coluna].dt.date
        )  # Extraindo e convertendo a data
        df["Hora"] = pd.to_datetime(
            df[coluna].dt.time, format="%H:%M:%S"
        )  # Extraindo e convertendo a hora
        return df

    # Função para converter colunas de data e filtrar com base nas datas únicas
    def converte_e_filtra_data(df, coluna_data, datas_unicas):
        df[coluna_data] = pd.to_datetime(df[coluna_data], format="%d/%m/%Y")
        df_filtrado = df[df[coluna_data].isin(datas_unicas)]
        return df_filtrado

    # Processamento das datas únicas
    datas_unicas = pd.to_datetime(
        justificativas["Dia Referência"].dropna(), format="%d/%m/%Y"
    )

    # Obtendo a primeira e última datas e formatando
    primeira_data_formatada = datas_unicas.min().strftime("%d.%m.%y")
    ultima_data_formatada = datas_unicas.max().strftime("%d.%m.%y")

    print(primeira_data_formatada)
    print(ultima_data_formatada)

    # Separando hora e data nas colunas de interesse
    tarefas = separa_hora_data(tarefas, "Data de Conclusão")
    pesquisasporcentagem_x = pesquisasporcentagem_x.fillna(
        pd.to_datetime("1999-01-01 23:59:00")
    )
    pesquisasporcentagem_x = separa_hora_data(
        pesquisasporcentagem_x, "Data de Conclusão"
    )

    # Filtrando DataFrames com base nas datas únicas
    gerencial = converte_e_filtra_data(gerencial, "Data da Visita", datas_unicas)
    relatorio_de_checkin = converte_e_filtra_data(
        relatorio_de_checkin, "Data do Roteiro", datas_unicas
    )
    tarefas = converte_e_filtra_data(tarefas, "data", datas_unicas)

    # Critérios de exclusão para pesquisasporcentagem_x
    pesquisasporcentagem_x = pesquisasporcentagem_x[
        (pesquisasporcentagem_x["data"].isin(datas_unicas))
        | (pesquisasporcentagem_x["Data de Expiração"].isin(datas_unicas))
    ]

    # Validação de justificativas abonáveis
    justificativas["Justificativas válidas"] = "NÃO"
    # justificativas_validas_lista = justificativas_abonaveis[
    #     "Justificativas válidas"
    # ].tolist()


    justificativas_validas_lista = ["ATESTADO MEDICO - INSS", "ATESTADO MÉDICO 1 DIA", "ATESTADO MÉDICO 2 DIAS", "ATESTADO MÉDICO 3 DIAS", 
                                    "ATESTADO MÉDICO 4 DIAS", "ATESTADO MÉDICO 5 DIAS", "ATESTADO MÉDICO 6 A 14 DIAS", "AUDIÊNCIA", 
                                    "CASAMENTO", "CONFRATERNIZAÇÕES", "DAY OFF (FOLGA DE ANIVERSARIO)", "DECLARAÇÃO DE HORAS", "FALECIMENTO FAMILIAR",
                                    "FERIADO", "FÉRIAS", "FOLGA - BANCO DE HORAS", "LICENÇA MATERNIDADE", "LICENÇA PATERNIDADE", "REUNIÃO",
                                    "SUSPENSO", "TREINAMENTO (NOVO PROMOTOR)"]


    for justificativa in justificativas_validas_lista:
        justificativas.loc[
            justificativas["Afastado"] == justificativa, "Justificativas válidas"
        ] = "SIM"

    # Tratamento do DataFrame relatorio_de_checkin
    relatorio_de_checkin = relatorio_de_checkin.loc[
        relatorio_de_checkin["Tipo de Checkin"] == "Checkin Manual"
    ]
    relatorio_de_checkin["Hora Saída"].fillna(
        pd.to_datetime("1900-01-01 23:59:00"), inplace=True
    )
    relatorio_de_checkin["ID"] = ""
    relatorio_de_checkin["CONTRATO"] = "RT"
    relatorio_de_checkin.rename(columns={"Tipo de Checkin": "RÓTULO"}, inplace=True)
    relatorio_de_checkin = relatorio_de_checkin[
        [
            "ID",
            "Colaborador",
            "RÓTULO",
            "CONTRATO",
            "Colaborador Superior",
            "Bandeira",
            "Ponto de Venda",
            "Estado",
            "Data do Roteiro",
            "Hora Entrada",
            "Hora Saída",
        ]
    ]
    relatorio_de_checkin.reset_index(drop=True, inplace=True)

    def separa_hora_data(df, coluna):
        variavel = df
        variavel["data"] = [d.date() for d in variavel[coluna]]
        variavel["Hora"] = [d.time() for d in variavel[coluna]]

        variavel["data"] = pd.to_datetime(variavel["data"])
        variavel["Hora"] = pd.to_datetime(variavel["Hora"], format="%H:%M:%S")

        return variavel

    # -------------------------- Criando novas colunas de Data e Hora

    # -------------------------- Transformando valores da coluna em formado data & hora
    deslocamento = separa_hora_data(relatorio_de_checkin, "Hora Entrada")

    # -------------------------- Criando coluna com informação "Deslocamento"
    deslocamento["ID"] = "Deslocamento"

    # -------------------------- Selecionando as colunas que desejo trabalhar & renomeando "deslocamento" para "checkin"
    checkin = deslocamento[
        [
            "ID",
            "Colaborador",
            "RÓTULO",
            "CONTRATO",
            "Colaborador Superior",
            "Bandeira",
            "Ponto de Venda",
            "Estado",
            "Data do Roteiro",
            "Hora",
        ]
    ]
    checkin = checkin.rename({"Data do Roteiro": "data"}, axis=1)

    # ------------------------------------------------------ Separando DATA da hora e formatando-a
    checkin["Hora"] = checkin["Hora"].astype(str)

    checkin[["xxx", "Hora"]] = checkin["Hora"].str.split(" ", expand=True)

    checkin["Hora"] = pd.to_timedelta(checkin["Hora"])

    checkin = checkin.drop(["xxx"], axis=1)

    # -------------------------- Criando novas colunas de Data e Hora

    # -------------------------- Transformando valores da coluna em formado data & hora
    ocio = separa_hora_data(relatorio_de_checkin, "Hora Saída")

    # -------------------------- Criando coluna com informação "ocio"
    ocio["ID"] = "Ocio"
    ocio["RÓTULO"] = "Checkout Manual"

    # -------------------------- Selecionando as colunas que desejo trabalhar & renomeando "ocio" para "checkout"
    checkout = ocio[
        [
            "ID",
            "Colaborador",
            "RÓTULO",
            "CONTRATO",
            "Colaborador Superior",
            "Bandeira",
            "Ponto de Venda",
            "Estado",
            "Data do Roteiro",
            "Hora",
        ]
    ]
    checkout = checkout.rename({"Data do Roteiro": "data"}, axis=1)

    # ------------------------------------------------------ Separando DATA da hora e formatando-a
    checkout["Hora"] = checkout["Hora"].astype(str)

    checkout[["xxx", "Hora"]] = checkout["Hora"].str.split(" ", expand=True)

    checkout["Hora"] = pd.to_timedelta(checkout["Hora"])

    checkout = checkout.drop(["xxx"], axis=1)

    # -------------------------- Tratando planilha de tarefas
    tarefas["ID"] = "Produtividade"

    tarefas = tarefas.loc[tarefas["Status"] == "Respondida"]

    produtividade = tarefas[
        [
            "Id",
            "ID",
            "Tipo de Coleta",
            "Rótulo",
            "Ponto de Venda",
            "Bandeira",
            "Estado",
            "Responsável",
            "Colaborador Superior",
            "data",
            "Hora",
        ]
    ]
    produtividade = produtividade.rename({"Rótulo": "RÓTULO"}, axis=1)
    produtividade = produtividade.rename({"Responsável": "Colaborador"}, axis=1)

    # ------------------------------------------------------ Separando DATA da hora e formatando-a
    produtividade["Hora"] = produtividade["Hora"].astype(str)

    produtividade[["xxx", "Hora"]] = produtividade["Hora"].str.split(" ", expand=True)

    produtividade["Hora"] = pd.to_timedelta(produtividade["Hora"])

    produtividade = produtividade.drop(["xxx"], axis=1)

    produtividade = produtividade[
        [
            "ID",
            "Colaborador",
            "RÓTULO",
            "Colaborador Superior",
            "Bandeira",
            "Ponto de Venda",
            "Estado",
            "data",
            "Hora",
        ]
    ]

    # Agrupar por colaborador e data e encontrar a última pesquisa realizada
    ultima_pesquisa = (
        produtividade.groupby(["Colaborador", "data"])["Hora"].max().reset_index()
    )

    ultima_pesquisa.rename({"Hora": "Hora Correta"}, axis=1, inplace=True)

    # Concatenando
    PDOH = pd.concat([checkin, checkout, produtividade])

    PDOH = pd.merge(PDOH, ultima_pesquisa, on=["Colaborador", "data"], how="left")

    PDOH["Hora Correta"] = PDOH["Hora Correta"].fillna(pd.to_timedelta(PDOH["Hora"]))

    PDOH.loc[PDOH["Hora"] >= pd.to_timedelta("0 days 23:30:00"), "Hora"] = PDOH[
        "Hora Correta"
    ]

    # PDOH.to_excel("PDOH.xlsx")

    # Selecionando Filtros
    PDOH.sort_values(by=["Colaborador", "data", "Hora"], inplace=True)

    # Reorganizando índices
    PDOH.index = range(PDOH.shape[0])

    # Criando Variável para receber os valores
    timein = []
    linhas = PDOH.shape[0] - 1

    # Criando uma das condições e armazenando em uma variável
    t0 = pd.to_timedelta("00:00:00")

    # Criando um laço de repetição por linha

    for i in range(linhas):
        if PDOH["data"][i + 1] != PDOH["data"][i]:
            timein.append(t0)
        else:
            timein.append(PDOH["Hora"][i + 1] - PDOH["Hora"][i])

    # Criando Data Frame com a variável que armazenou as informações
    timein = pd.DataFrame(timein)

    # Adicionando uma coluna Nula na primeira linha da coluna
    timein.loc[-1] = t0
    timein.index = timein.index + 1
    timein = timein.sort_index()

    # Conferindo se a coluna de cálculos criada é do mesmo tamanho da do PDOH
    timein.shape[0] - PDOH.shape[0]

    # Adicionando coluna de cálculos ao PDOH
    PDOH["Tempo Investido"] = timein

    PDOH.loc[
        PDOH["Tempo Investido"] < pd.to_timedelta("0 days 00:00:00"), "Tempo Investido"
    ] = pd.to_timedelta("0 days 00:00:00")

    dias_ptbr = {
        "Sunday": "Domingo",
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
    }

    PDOH["nome_do_dia"] = PDOH["data"].dt.day_name().replace(dias_ptbr)


    # Converter a coluna "data" para datetime para facilitar a manipulação
    PDOH["data"] = pd.to_datetime(PDOH["data"])

    # Obter uma lista de pares únicos de "nome_do_dia" e "data" com a semana correspondente
    dias_datas_unicas = PDOH[["nome_do_dia", "data"]].drop_duplicates().reset_index(drop=True)
    dias_datas_unicas["semana_do_ano"] = dias_datas_unicas["data"].dt.isocalendar().week

    # Lista completa dos dias da semana
    dias_da_semana_geral = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado"]

    # Identificar a semana e o ano presentes em dias_datas_unicas
    semana_alvo = dias_datas_unicas["semana_do_ano"].iloc[0]
    ano_alvo = dias_datas_unicas["data"].dt.year.iloc[0]

    # Obter o primeiro dia com registros na semana e calcular a segunda-feira dessa semana
    primeiro_dia_com_registro = dias_datas_unicas[dias_datas_unicas["semana_do_ano"] == semana_alvo]["data"].min()
    primeira_segunda_feira = primeiro_dia_com_registro - timedelta(days=primeiro_dia_com_registro.weekday())

    # Criar um DataFrame com todos os dias da semana e suas datas correspondentes para essa semana específica
    todos_os_dias = pd.DataFrame({
        "nome_do_dia": dias_da_semana_geral,
        "data": [primeira_segunda_feira + timedelta(days=i) for i in range(6)],  # Gera datas de segunda a sábado
        "semana_do_ano": semana_alvo  # Preencher todos com a semana alvo
    })

    # Filtrar para manter apenas os dias ausentes
    dias_ausentes = todos_os_dias[~todos_os_dias["nome_do_dia"].isin(dias_datas_unicas["nome_do_dia"])]

    # Adicionar os dias ausentes ao DataFrame dias_datas_unicas
    dias_datas_unicas = pd.concat([dias_datas_unicas, dias_ausentes], ignore_index=True).sort_values(by="data").reset_index(drop=True)

    # Lista dos dias da semana
    dias_da_semana_geral = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado"]

    # Obter valores únicos das colunas
    colaboradores_unicos = PDOH["Colaborador"].unique()
    superiores_unicos = PDOH["Colaborador Superior"].unique()

    # Iterar sobre cada colaborador e verificar dias ausentes
    for colaborador in colaboradores_unicos:
        for dia in dias_da_semana_geral:
            # Filtrar registros para verificar se o colaborador tem o dia registrado
            registros_dia = PDOH[(PDOH["Colaborador"] == colaborador) & (PDOH["nome_do_dia"] == dia)]
            
            # Se o dia não existir para o colaborador, adiciona as linhas com "Produtividade", "Deslocamento" e "Ocio"
            if registros_dia.empty:
                novas_linhas = pd.DataFrame({
                    "ID": ["Produtividade", "Deslocamento", "Ocio"],
                    "nome_do_dia": [dia, dia, dia],
                    "Colaborador": [colaborador, colaborador, colaborador],
                    "Colaborador Superior": [PDOH.loc[PDOH["Colaborador"] == colaborador, "Colaborador Superior"].iloc[0]] * 3,  # Associa o mesmo superior
                    "Tempo Investido": [pd.to_timedelta("0 days 00:00:00")] * 3
                })
                
                # Adicionar as novas linhas ao DataFrame PDOH
                PDOH = pd.concat([PDOH, novas_linhas], ignore_index=True)

    # Resetar o índice final
    PDOH = PDOH.reset_index(drop=True)

    # Realizar um merge entre PDOH e dias_datas_unicas baseado em "nome_do_dia" para trazer as datas correspondentes
    PDOH = PDOH.merge(dias_datas_unicas[["nome_do_dia", "data"]], on="nome_do_dia", how="left", suffixes=("", "_unicas"))

    # Sobrescrever as datas em branco na coluna "data" de PDOH com os valores de "data_unicas" quando houver valores ausentes
    PDOH["data"] = PDOH["data"].fillna(PDOH["data_unicas"])

    # Remover a coluna auxiliar "data_unicas" após o preenchimento
    PDOH = PDOH.drop(columns=["data_unicas"])

    PDOH.to_excel(
        f"outputs/BASE DE DADOS PDOH {nome_do_cliente} {primeira_data_formatada} à {ultima_data_formatada}.xlsx",
        sheet_name="Base de Dados PDOH",
        index=False,
    )

    dias_da_semana = PDOH[["nome_do_dia"]]

    dias_da_semana = dias_da_semana.drop_duplicates()

    dias_da_semana = dias_da_semana.reset_index(drop=True)

    print(dias_da_semana)

    comprimento_dias = dias_da_semana["nome_do_dia"].shape[0]

    PDOHSEG = PDOH.loc[PDOH["nome_do_dia"] == "Segunda-feira"]
    PDOHTER = PDOH.loc[PDOH["nome_do_dia"] == "Terça-feira"]
    PDOHQUA = PDOH.loc[PDOH["nome_do_dia"] == "Quarta-feira"]
    PDOHQUI = PDOH.loc[PDOH["nome_do_dia"] == "Quinta-feira"]
    PDOHSEX = PDOH.loc[PDOH["nome_do_dia"] == "Sexta-feira"]
    PDOHSAB = PDOH.loc[PDOH["nome_do_dia"] == "Sábado"]

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # ----------------- Formatando e organizando planilha de almoco
    almoco = almoco[["Colaborador", "Tempo em Loja Executado", "Data da Visita"]]

    almoco["Tempo em Loja Executado"] = pd.to_timedelta(
        almoco["Tempo em Loja Executado"]
    )

    almoco["Data da Visita"] = pd.to_datetime(
        almoco["Data da Visita"], format="%d/%m/%Y"
    )

    dias_ptbr = {
        "Sunday": "Domingo",
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
    }

    almoco["Data da Visita"].dt.day_name().replace(dias_ptbr)

    almoco["nome_do_dia"] = almoco["Data da Visita"].dt.day_name().replace(dias_ptbr)

    almoco = almoco.rename({"Tempo em Loja Executado": "almoco"}, axis=1)

    almoco.loc[(almoco["nome_do_dia"] == "Segunda-feira"), "almoco - SEG"] = almoco[
        "almoco"
    ]
    almoco.loc[(almoco["nome_do_dia"] == "Terça-feira"), "almoco - TER"] = almoco[
        "almoco"
    ]
    almoco.loc[(almoco["nome_do_dia"] == "Quarta-feira"), "almoco - QUA"] = almoco[
        "almoco"
    ]
    almoco.loc[(almoco["nome_do_dia"] == "Quinta-feira"), "almoco - QUI"] = almoco[
        "almoco"
    ]
    almoco.loc[(almoco["nome_do_dia"] == "Sexta-feira"), "almoco - SEX"] = almoco[
        "almoco"
    ]
    almoco.loc[(almoco["nome_do_dia"] == "Sábado"), "almoco - SAB"] = almoco["almoco"]

    almoco_SEG = almoco.loc[~almoco["almoco - SEG"].isnull()]

    almoco_SEG = almoco_SEG[["Colaborador", "almoco - SEG"]]

    almoco_SEG = almoco_SEG.groupby("Colaborador").sum()

    almoco_SEG = almoco_SEG.reset_index(drop=False)

    almoco_TER = almoco.loc[~almoco["almoco - TER"].isnull()]

    almoco_TER = almoco_TER[["Colaborador", "almoco - TER"]]

    almoco_TER = almoco_TER.groupby("Colaborador").sum()

    almoco_TER = almoco_TER.reset_index(drop=False)

    almoco_QUA = almoco.loc[~almoco["almoco - QUA"].isnull()]

    almoco_QUA = almoco_QUA[["Colaborador", "almoco - QUA"]]

    almoco_QUA = almoco_QUA.groupby("Colaborador").sum()

    almoco_QUA = almoco_QUA.reset_index(drop=False)

    almoco_QUI = almoco.loc[~almoco["almoco - QUI"].isnull()]

    almoco_QUI = almoco_QUI[["Colaborador", "almoco - QUI"]]

    almoco_QUI = almoco_QUI.groupby("Colaborador").sum()

    almoco_QUI = almoco_QUI.reset_index(drop=False)

    almoco_SEX = almoco.loc[~almoco["almoco - SEX"].isnull()]

    almoco_SEX = almoco_SEX[["Colaborador", "almoco - SEX"]]

    almoco_SEX = almoco_SEX.groupby("Colaborador").sum()

    almoco_SEX = almoco_SEX.reset_index(drop=False)

    almoco_SAB = almoco.loc[~almoco["almoco - SAB"].isnull()]

    almoco_SAB = almoco_SAB[["Colaborador", "almoco - SAB"]]

    almoco_SAB = almoco_SAB.groupby("Colaborador").sum()

    almoco_SAB = almoco_SAB.reset_index(drop=False)

    jornada_diaria = pd.merge(jornada_diaria, almoco_SEG, how="left", on="Colaborador")
    jornada_diaria = pd.merge(jornada_diaria, almoco_TER, how="left", on="Colaborador")
    jornada_diaria = pd.merge(jornada_diaria, almoco_QUA, how="left", on="Colaborador")
    jornada_diaria = pd.merge(jornada_diaria, almoco_QUI, how="left", on="Colaborador")
    jornada_diaria = pd.merge(jornada_diaria, almoco_SEX, how="left", on="Colaborador")
    jornada_diaria = pd.merge(jornada_diaria, almoco_SAB, how="left", on="Colaborador")

    jornada_diaria["almoco - SEG"] = jornada_diaria["almoco - SEG"].fillna(
        pd.to_timedelta("0 days 00:00:00")
    )
    jornada_diaria["almoco - TER"] = jornada_diaria["almoco - TER"].fillna(
        pd.to_timedelta("0 days 00:00:00")
    )
    jornada_diaria["almoco - QUA"] = jornada_diaria["almoco - QUA"].fillna(
        pd.to_timedelta("0 days 00:00:00")
    )
    jornada_diaria["almoco - QUI"] = jornada_diaria["almoco - QUI"].fillna(
        pd.to_timedelta("0 days 00:00:00")
    )
    jornada_diaria["almoco - SEX"] = jornada_diaria["almoco - SEX"].fillna(
        pd.to_timedelta("0 days 00:00:00")
    )
    jornada_diaria["almoco - SAB"] = jornada_diaria["almoco - SAB"].fillna(
        pd.to_timedelta("0 days 00:00:00")
    )

    if dias_da_semana["nome_do_dia"].str.contains("Segunda-feira").any():
        jornada_diaria["SEGUNDA-FEIRA"] = pd.to_timedelta(
            jornada_diaria["SEGUNDA-FEIRA"]
        )
        jornada_diaria["almoco - SEG"] = pd.to_timedelta(jornada_diaria["almoco - SEG"])

    else:
        jornada_diaria["SEGUNDA-FEIRA"] = pd.to_timedelta("0 days 00:00:00")
        jornada_diaria["almoco - SEG"] = pd.to_timedelta("0 days 00:00:00")

    if dias_da_semana["nome_do_dia"].str.contains("Terça-feira").any():
        jornada_diaria["TERÇA-FEIRA"] = pd.to_timedelta(jornada_diaria["TERÇA-FEIRA"])
        jornada_diaria["almoco - TER"] = pd.to_timedelta(jornada_diaria["almoco - TER"])
    else:
        jornada_diaria["TERÇA-FEIRA"] = pd.to_timedelta("0 days 00:00:00")
        jornada_diaria["almoco - TER"] = pd.to_timedelta("0 days 00:00:00")

    if dias_da_semana["nome_do_dia"].str.contains("Quarta-feira").any():
        jornada_diaria["QUARTA-FEIRA"] = pd.to_timedelta(jornada_diaria["QUARTA-FEIRA"])
        jornada_diaria["almoco - QUA"] = pd.to_timedelta(jornada_diaria["almoco - QUA"])
    else:
        jornada_diaria["QUARTA-FEIRA"] = pd.to_timedelta("0 days 00:00:00")
        jornada_diaria["almoco - QUA"] = pd.to_timedelta("0 days 00:00:00")

    if dias_da_semana["nome_do_dia"].str.contains("Quinta-feira").any():
        jornada_diaria["QUINTA-FEIRA"] = pd.to_timedelta(jornada_diaria["QUINTA-FEIRA"])
        jornada_diaria["almoco - QUI"] = pd.to_timedelta(jornada_diaria["almoco - QUI"])
    else:
        jornada_diaria["QUINTA-FEIRA"] = pd.to_timedelta("0 days 00:00:00")
        jornada_diaria["almoco - QUI"] = pd.to_timedelta("0 days 00:00:00")

    if dias_da_semana["nome_do_dia"].str.contains("Sexta-feira").any():
        jornada_diaria["SEXTA-FEIRA"] = pd.to_timedelta(jornada_diaria["SEXTA-FEIRA"])
        jornada_diaria["almoco - SEX"] = pd.to_timedelta(jornada_diaria["almoco - SEX"])
    else:
        jornada_diaria["SEXTA-FEIRA"] = pd.to_timedelta("0 days 00:00:00")
        jornada_diaria["almoco - SEX"] = pd.to_timedelta("0 days 00:00:00")

    if dias_da_semana["nome_do_dia"].str.contains("Sábado").any():
        jornada_diaria["SÁBADO"] = pd.to_timedelta(jornada_diaria["SÁBADO"])
        jornada_diaria["almoco - SAB"] = pd.to_timedelta(jornada_diaria["almoco - SAB"])
    else:
        jornada_diaria["SÁBADO"] = pd.to_timedelta("0 days 00:00:00")
        jornada_diaria["almoco - SAB"] = pd.to_timedelta("0 days 00:00:00")

    jornada_diaria["Jornada de Trabalho - SEMANAL"] = pd.to_timedelta(
        jornada_diaria["Jornada de Trabalho - SEMANAL"]
    )
    jornada_diaria["almoco - SEMANAL"] = pd.to_timedelta(
        jornada_diaria["almoco - SEMANAL"]
    )

    # ---------------------------------------------- JUSTIFICATIVAS ------------------------------------------------------

    justificativas = justificativas[
        ["Colaborador", "Dia Referência", "Afastado", "Justificativas válidas"]
    ]

    justificativas["nome_do_dia"] = (
        justificativas["Dia Referência"].dt.day_name().replace(dias_ptbr)
    )

    justificativas_seg = justificativas.loc[
        justificativas["nome_do_dia"] == "Segunda-feira"
    ]

    justificativas_seg = justificativas_seg.sort_values(
        ["Justificativas válidas"], ascending=False
    )
    justificativas_seg = justificativas_seg.sort_values(["Colaborador"], ascending=True)

    justificativas_seg.drop_duplicates(subset=["Colaborador"], inplace=True)

    justificativas_ter = justificativas.loc[
        justificativas["nome_do_dia"] == "Terça-feira"
    ]

    justificativas_ter = justificativas_ter.sort_values(
        ["Justificativas válidas"], ascending=False
    )
    justificativas_ter = justificativas_ter.sort_values(["Colaborador"], ascending=True)

    justificativas_ter.drop_duplicates(subset=["Colaborador"], inplace=True)

    justificativas_qua = justificativas.loc[
        justificativas["nome_do_dia"] == "Quarta-feira"
    ]

    justificativas_qua = justificativas_qua.sort_values(
        ["Justificativas válidas"], ascending=False
    )
    justificativas_qua = justificativas_qua.sort_values(["Colaborador"], ascending=True)

    justificativas_qua.drop_duplicates(subset=["Colaborador"], inplace=True)

    justificativas_qui = justificativas.loc[
        justificativas["nome_do_dia"] == "Quinta-feira"
    ]

    justificativas_qui = justificativas_qui.sort_values(
        ["Justificativas válidas"], ascending=False
    )
    justificativas_qui = justificativas_qui.sort_values(["Colaborador"], ascending=True)

    justificativas_qui.drop_duplicates(subset=["Colaborador"], inplace=True)

    justificativas_sex = justificativas.loc[
        justificativas["nome_do_dia"] == "Sexta-feira"
    ]

    justificativas_sex = justificativas_sex.sort_values(
        ["Justificativas válidas"], ascending=False
    )
    justificativas_sex = justificativas_sex.sort_values(["Colaborador"], ascending=True)

    justificativas_sex.drop_duplicates(subset=["Colaborador"], inplace=True)

    justificativas_sab = justificativas.loc[justificativas["nome_do_dia"] == "Sábado"]

    justificativas_sab = justificativas_sab.sort_values(
        ["Justificativas válidas"], ascending=False
    )
    justificativas_sab = justificativas_sab.sort_values(["Colaborador"], ascending=True)

    justificativas_sab.drop_duplicates(subset=["Colaborador"], inplace=True)

    colaboradores_SEG = jornada_diaria[
        ["UF", "Colaborador", "Colaborador Superior", "SEGUNDA-FEIRA", "almoco - SEG"]
    ]

    colaboradores_SEG = pd.merge(
        colaboradores_SEG, justificativas_seg, how="left", on="Colaborador"
    )

    colaboradores_SEG = colaboradores_SEG.rename(
        {
            "SEGUNDA-FEIRA": "Jornada de Trabalho",
            "almoco - SEG": "almoco",
            "Afastado": "justificativas_seg",
        },
        axis=1,
    )
    print(colaboradores_SEG)

    jus_SEG = colaboradores_SEG[["Colaborador", "justificativas_seg"]]

    colaboradores_TER = jornada_diaria[
        ["UF", "Colaborador", "Colaborador Superior", "TERÇA-FEIRA", "almoco - TER"]
    ]

    colaboradores_TER = pd.merge(
        colaboradores_TER, justificativas_ter, how="left", on="Colaborador"
    )

    colaboradores_TER = colaboradores_TER.rename(
        {
            "TERÇA-FEIRA": "Jornada de Trabalho",
            "almoco - TER": "almoco",
            "Afastado": "justificativas_ter",
        },
        axis=1,
    )

    jus_TER = colaboradores_TER[["Colaborador", "justificativas_ter"]]

    colaboradores_QUA = jornada_diaria[
        ["UF", "Colaborador", "Colaborador Superior", "QUARTA-FEIRA", "almoco - QUA"]
    ]

    colaboradores_QUA = pd.merge(
        colaboradores_QUA, justificativas_qua, how="left", on="Colaborador"
    )

    colaboradores_QUA = colaboradores_QUA.rename(
        {
            "QUARTA-FEIRA": "Jornada de Trabalho",
            "almoco - QUA": "almoco",
            "Afastado": "justificativas_qua",
        },
        axis=1,
    )

    jus_QUA = colaboradores_QUA[["Colaborador", "justificativas_qua"]]

    colaboradores_QUI = jornada_diaria[
        ["UF", "Colaborador", "Colaborador Superior", "QUINTA-FEIRA", "almoco - QUI"]
    ]

    colaboradores_QUI = pd.merge(
        colaboradores_QUI, justificativas_qui, how="left", on="Colaborador"
    )

    colaboradores_QUI = colaboradores_QUI.rename(
        {
            "QUINTA-FEIRA": "Jornada de Trabalho",
            "almoco - QUI": "almoco",
            "Afastado": "justificativas_qui",
        },
        axis=1,
    )

    jus_QUI = colaboradores_QUI[["Colaborador", "justificativas_qui"]]

    colaboradores_SEX = jornada_diaria[
        ["UF", "Colaborador", "Colaborador Superior", "SEXTA-FEIRA", "almoco - SEX"]
    ]

    colaboradores_SEX = pd.merge(
        colaboradores_SEX, justificativas_sex, how="left", on="Colaborador"
    )

    colaboradores_SEX = colaboradores_SEX.rename(
        {
            "SEXTA-FEIRA": "Jornada de Trabalho",
            "almoco - SEX": "almoco",
            "Afastado": "justificativas_sex",
        },
        axis=1,
    )

    jus_SEX = colaboradores_SEX[["Colaborador", "justificativas_sex"]]

    colaboradores_SAB = jornada_diaria[
        ["UF", "Colaborador", "Colaborador Superior", "SÁBADO", "almoco - SAB"]
    ]

    colaboradores_SAB = pd.merge(
        colaboradores_SAB, justificativas_sab, how="left", on="Colaborador"
    )

    colaboradores_SAB = colaboradores_SAB.rename(
        {
            "SÁBADO": "Jornada de Trabalho",
            "almoco - SAB": "almoco",
            "Afastado": "justificativas_sab",
        },
        axis=1,
    )

    jus_SAB = colaboradores_SAB[["Colaborador", "justificativas_sab"]]

    colaboradores_SEG.loc[
        colaboradores_SEG["Justificativas válidas"] == "SIM", "Jornada de Trabalho"
    ] = pd.to_timedelta("0 days 00:00:00")
    colaboradores_SEG.loc[
        colaboradores_SEG["Justificativas válidas"] == "SIM", "almoco"
    ] = pd.to_timedelta("0 days 00:00:00")

    colaboradores_TER.loc[
        colaboradores_TER["Justificativas válidas"] == "SIM", "Jornada de Trabalho"
    ] = pd.to_timedelta("0 days 00:00:00")
    colaboradores_TER.loc[
        colaboradores_TER["Justificativas válidas"] == "SIM", "almoco"
    ] = pd.to_timedelta("0 days 00:00:00")

    colaboradores_QUA.loc[
        colaboradores_QUA["Justificativas válidas"] == "SIM", "Jornada de Trabalho"
    ] = pd.to_timedelta("0 days 00:00:00")
    colaboradores_QUA.loc[
        colaboradores_QUA["Justificativas válidas"] == "SIM", "almoco"
    ] = pd.to_timedelta("0 days 00:00:00")

    colaboradores_QUI.loc[
        colaboradores_QUI["Justificativas válidas"] == "SIM", "Jornada de Trabalho"
    ] = pd.to_timedelta("0 days 00:00:00")
    colaboradores_QUI.loc[
        colaboradores_QUI["Justificativas válidas"] == "SIM", "almoco"
    ] = pd.to_timedelta("0 days 00:00:00")

    colaboradores_SEX.loc[
        colaboradores_SEX["Justificativas válidas"] == "SIM", "Jornada de Trabalho"
    ] = pd.to_timedelta("0 days 00:00:00")
    colaboradores_SEX.loc[
        colaboradores_SEX["Justificativas válidas"] == "SIM", "almoco"
    ] = pd.to_timedelta("0 days 00:00:00")

    colaboradores_SAB.loc[
        colaboradores_SAB["Justificativas válidas"] == "SIM", "Jornada de Trabalho"
    ] = pd.to_timedelta("0 days 00:00:00")
    colaboradores_SAB.loc[
        colaboradores_SAB["Justificativas válidas"] == "SIM", "almoco"
    ] = pd.to_timedelta("0 days 00:00:00")

    colaboradores = jornada_diaria[
        [
            "UF",
            "Colaborador",
            "Colaborador Superior",
            "Jornada de Trabalho - SEMANAL",
            "almoco - SEMANAL",
        ]
    ]

    colaboradores = colaboradores.rename(
        {
            "Jornada de Trabalho - SEMANAL": "Jornada de Trabalho",
            "almoco - SEMANAL": "almoco",
        },
        axis=1,
    )


    colaboradores = pd.merge(colaboradores, colaboradores_SEG, on=["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco"], how="outer")
    colaboradores = pd.merge(colaboradores, colaboradores_TER, on=["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco", "Dia Referência", "Justificativas válidas", "nome_do_dia"], how="outer")
    colaboradores = pd.merge(colaboradores, colaboradores_QUA, on=["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco", "Dia Referência", "Justificativas válidas", "nome_do_dia"], how="outer")
    colaboradores = pd.merge(colaboradores, colaboradores_QUI, on=["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco", "Dia Referência", "Justificativas válidas", "nome_do_dia"], how="outer")
    colaboradores = pd.merge(colaboradores, colaboradores_SEX, on=["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco", "Dia Referência", "Justificativas válidas", "nome_do_dia"], how="outer")
    colaboradores = pd.merge(colaboradores, colaboradores_SAB, on=["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco", "Dia Referência", "Justificativas válidas", "nome_do_dia"], how="outer")


    colaboradores = colaboradores.dropna(subset="Dia Referência")

    colaboradores = colaboradores[["UF", "Colaborador", "Colaborador Superior", "Jornada de Trabalho", "almoco"]]

    # Substituir NaNs por uma string temporária (por exemplo, "Indefinido") antes de agrupar
    colaboradores["UF"] = colaboradores["UF"].fillna("SEM UF")
    colaboradores["Colaborador Superior"] = colaboradores["Colaborador Superior"].fillna("SEM UF")

    # Agrupar e somar as colunas especificadas
    colaboradores = colaboradores.groupby(["UF", "Colaborador", "Colaborador Superior"]).agg({"Jornada de Trabalho": "sum", "almoco": "sum"}).reset_index()

    print(colaboradores)

    # Concatenando as justificativas

    colaboradores = pd.merge(colaboradores, jus_SEG, how="left", on="Colaborador")
    colaboradores = pd.merge(colaboradores, jus_TER, how="left", on="Colaborador")
    colaboradores = pd.merge(colaboradores, jus_QUA, how="left", on="Colaborador")
    colaboradores = pd.merge(colaboradores, jus_QUI, how="left", on="Colaborador")
    colaboradores = pd.merge(colaboradores, jus_SEX, how="left", on="Colaborador")
    colaboradores = pd.merge(colaboradores, jus_SAB, how="left", on="Colaborador")

    # definir uma função para concatenar as strings das colunas
    def concatenar_strings(row):
        return "; ".join(row.values.astype(str))

    # aplicar a função em cada linha do dataframe para criar a nova coluna
    colaboradores["justificativas"] = colaboradores[
        [
            "justificativas_seg",
            "justificativas_ter",
            "justificativas_qua",
            "justificativas_qui",
            "justificativas_sex",
            "justificativas_sab",
        ]
    ].apply(concatenar_strings, axis=1)

    colaboradores["justificativas"] = colaboradores["justificativas"].str.replace(
        "nan;", ""
    )
    colaboradores["justificativas"] = colaboradores["justificativas"].str.replace(
        "; nan", ""
    )
    colaboradores["justificativas"] = colaboradores["justificativas"].str.replace(
        "nan", ""
    )

    colaboradores = colaboradores[
        [
            "UF",
            "Colaborador",
            "Colaborador Superior",
            "Jornada de Trabalho",
            "almoco",
            "justificativas",
        ]
    ]

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------PDOH SEMANAL COMPLETO--------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    def AnalisePor(df, filtro):
        selecao = df["ID"] == filtro
        filtrado = df[selecao]
        agrupado = filtrado.groupby(["Colaborador"])
        somado = (agrupado["Tempo Investido"].sum()).to_frame()
        IndexReset = somado.reset_index()
        saida = IndexReset.rename(columns={"Tempo Investido": filtro})
        return saida

    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    Ocio = AnalisePor(PDOH, "Ocio")

    #Criando variável para alocar o corte do almoco dentro do Ocio
    ocio_sem_almoco = Ocio

    ocio_sem_almoco = pd.merge(ocio_sem_almoco, colaboradores, on="Colaborador", how="right")

    ocio_sem_almoco = ocio_sem_almoco[["Colaborador", "Ocio", "almoco"]]

    ocio_sem_almoco = ocio_sem_almoco.rename({"almoco":"almoco"}, axis=1)

    ocio_sem_almoco["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)


    ocio_sem_almoco["ocio_sem_almoco"] = ocio_sem_almoco["Ocio"] - ocio_sem_almoco["almoco"]

    ocio_sem_almoco.loc[ocio_sem_almoco["ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco["almoco"] - ocio_sem_almoco["Ocio"]

    ocio_sem_almoco.loc[ocio_sem_almoco["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")

    ocio_sem_almoco.loc[ocio_sem_almoco["ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco["ocio_sem_almoco"]
    ocio_sem_almoco.loc[ocio_sem_almoco["ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")

    ocio_sem_almoco["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)


    Ocio = ocio_sem_almoco[["Colaborador", "Ocio_real"]]

    Ocio.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    Deslocamento = AnalisePor(PDOH,"Deslocamento")

    desloc_com_cortes = pd.merge(Deslocamento, ocio_sem_almoco, on="Colaborador", how="right")

    desloc_com_cortes["Deslocamento_pos_corte"] = desloc_com_cortes["Deslocamento"] - desloc_com_cortes["Corte_pendente"]

    desloc_com_cortes["corte_pendente_prod"] = desloc_com_cortes["Corte_pendente"] - desloc_com_cortes["Deslocamento"]

    desloc_com_cortes.loc[desloc_com_cortes["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")

    desloc_com_cortes.loc[desloc_com_cortes["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes["Deslocamento_pos_corte"]

    desloc_com_cortes.loc[desloc_com_cortes["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

    Deslocamento = desloc_com_cortes[["Colaborador", "Deslocamento"]]
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    Produtividadex = AnalisePor(PDOH,"Produtividade")

    Produtividadex = pd.merge(Produtividadex, desloc_com_cortes, how = "right", on = "Colaborador")

    Produtividadex.loc[Produtividadex["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex["Produtividade"] - Produtividadex["corte_pendente_prod"]

    Produtividadex.loc[Produtividadex["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")

    Produtividadex.loc[Produtividadex["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")

    Produtividadex["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

    Produtividade = Produtividadex[["Colaborador", "Produtividade"]]
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


    htr = pd.merge(Produtividade, Deslocamento, how="outer", on="Colaborador")
    HorasTotaisRegistradas = pd.merge(htr, Ocio, how="outer", on="Colaborador")
    HorasTotaisRegistradas["Tempo Investido"] = (
        HorasTotaisRegistradas["Produtividade"]
        + HorasTotaisRegistradas["Deslocamento"]
        + HorasTotaisRegistradas["Ocio"]
    )

    Horastotaisxxxx = pd.merge(
        HorasTotaisRegistradas, colaboradores, how="outer", on="Colaborador"
    )

    Horastotaisxxxx.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

    Horastotaisxxxx["HT sem almoco"] = Horastotaisxxxx["Tempo Investido"]

    JornadaDeTrabalho = Horastotaisxxxx[
        [
            "Colaborador",
            "Colaborador Superior",
            "UF",
            "HT sem almoco",
            "Jornada de Trabalho",
            "almoco",
            "justificativas",
        ]
    ]

    JornadaDeTrabalho["Hora Excedente"] = (
        JornadaDeTrabalho["HT sem almoco"] - JornadaDeTrabalho["Jornada de Trabalho"]
    )

    JornadaDeTrabalho["Hora Pendente"] = (
        JornadaDeTrabalho["Jornada de Trabalho"] - JornadaDeTrabalho["HT sem almoco"]
    )

    JornadaDeTrabalho.loc[
        JornadaDeTrabalho["Hora Excedente"] < pd.to_timedelta("0 days 00:00:00"),
        "Hora Excedente",
    ] = pd.to_timedelta("0 days 00:00:00")

    JornadaDeTrabalho.loc[
        JornadaDeTrabalho["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
        "Hora Pendente",
    ] = pd.to_timedelta("0 days 00:00:00")

    Horasnaoregistradas = JornadaDeTrabalho[
        [
            "Colaborador",
            "Colaborador Superior",
            "UF",
            "Hora Pendente",
            "almoco",
            "justificativas",
        ]
    ]

    corteszx = pd.merge(JornadaDeTrabalho, Deslocamento, how="outer", on="Colaborador")

    corteszx["Deslocamento_cortado"] = (
        corteszx["Deslocamento"] - corteszx["Hora Excedente"]
    )

    corteszx["Deslocamento_cortado_excedente"] = (
        corteszx["Hora Excedente"] - corteszx["Deslocamento"]
    )

    corteszx.loc[
        corteszx["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
        "Deslocamento_cortado",
    ] = pd.to_timedelta("0 days 00:00:00")
    corteszx.loc[
        corteszx["Deslocamento_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
        "Deslocamento_cortado_excedente",
    ] = pd.to_timedelta("0 days 00:00:00")

    cortesyy = pd.merge(corteszx, Ocio, how="outer", on="Colaborador")

    cortesyy["Ocio_cortado"] = (
        cortesyy["Ocio"] - cortesyy["Deslocamento_cortado_excedente"]
    )

    cortesyy["Ocio_cortado_excedente"] = (
        cortesyy["Deslocamento_cortado_excedente"] - cortesyy["Ocio"]
    )

    cortesyy.loc[
        cortesyy["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"), "Ocio_cortado"
    ] = pd.to_timedelta("0 days 00:00:00")
    cortesyy.loc[
        cortesyy["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
        "Ocio_cortado_excedente",
    ] = pd.to_timedelta("0 days 00:00:00")

    corteszz = pd.merge(cortesyy, Produtividade, how="right", on="Colaborador")

    corteszz["Produtividade_cortada"] = (
        corteszz["Produtividade"] - corteszz["Ocio_cortado_excedente"]
    )

    corteszz.loc[
        corteszz["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
        "Produtividade_cortada",
    ] = pd.to_timedelta("0 days 00:00:00")

    corteszz["TOTAL DE HR"] = (
        corteszz["Produtividade_cortada"]
        + corteszz["Deslocamento_cortado"]
        + corteszz["Ocio_cortado"]
    )

    cortes = corteszz[
        [
            "Colaborador",
            "Colaborador Superior",
            "UF",
            "Produtividade_cortada",
            "Deslocamento_cortado",
            "Ocio_cortado",
            "TOTAL DE HR",
            "Jornada de Trabalho",
            "Hora Pendente",
            "Hora Excedente",
            "almoco",
            "justificativas",
        ]
    ]

    cortes = cortes.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
    cortes = cortes.rename({"almoco": "almoco"}, axis=1)
    cortes = cortes.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

    PDOH_P = cortes.rename(
        {
            "Produtividade_cortada": "PRODUTIVIDADE",
            "Deslocamento_cortado": "DESLOCAMENTO",
            "Ocio_cortado": "ÓCIO",
            "TOTAL DE HR": "TOTAL",
            "Hora Pendente": "H N REGISTRADAS",
            "Jornada de Trabalho": "horas_programadas",
        },
        axis=1,
    )
    ####################################################################################################################

    # Indo para as demais colunas pós calculo do PDOH

    checkinweek = checkinweek.rename(
        {"Primeiro CheckIn": "primeiro_checkin", "Último CheckOut": "ultimo_checkout"},
        axis=1,
    )
    checkinweek = checkinweek[
        ["Colaborador", "Dia Referência", "primeiro_checkin", "ultimo_checkout"]
    ]

    checkinweek.dropna(subset=["Colaborador"], axis=0, inplace=True)
    checkinweek.dropna(subset=["primeiro_checkin"], axis=0, inplace=True)
    checkinweek.dropna(subset=["ultimo_checkout"], axis=0, inplace=True)
    checkinweek.index = range(checkinweek.shape[0])

    def separa_hora_data(df, coluna):
        variavel = df
        variavel["data"] = [d.date() for d in variavel[coluna]]
        variavel["Hora"] = [d.time() for d in variavel[coluna]]

        variavel["data"] = pd.to_datetime(variavel["data"])
        variavel["Hora"] = pd.to_datetime(variavel["Hora"], format="%H:%M:%S")

        return variavel

    primeirocheckin = checkinweek[["Colaborador", "Dia Referência", "primeiro_checkin"]]

    ultimocheckout = checkinweek[["Colaborador", "Dia Referência", "ultimo_checkout"]]

    primeirocheckin = separa_hora_data(primeirocheckin, "primeiro_checkin")

    ultimocheckout = separa_hora_data(ultimocheckout, "ultimo_checkout")

    primeirocheckin = primeirocheckin.rename(columns={"Hora": "Primeiro Checkin OK"})

    primeirocheckin = primeirocheckin[
        ["Colaborador", "Dia Referência", "Primeiro Checkin OK"]
    ]

    ultimocheckout = ultimocheckout.rename(columns={"Hora": "Último Checkout OK"})

    ultimocheckout = ultimocheckout[
        ["Colaborador", "Dia Referência", "Último Checkout OK"]
    ]

    primeirocheckin["Último Checkout OK"] = ultimocheckout["Último Checkout OK"]

    dias_ptbr = {
        "Sunday": "Domingo",
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
    }

    primeirocheckin["Dia Referência"].dt.day_name().replace(dias_ptbr)

    primeirocheckin["nome_do_dia"] = (
        primeirocheckin["Dia Referência"].dt.day_name().replace(dias_ptbr)
    )

    checkin_seg_a_sex = primeirocheckin[primeirocheckin["nome_do_dia"] != "Sábado"]
    checkin_seg_a_sex = primeirocheckin[primeirocheckin["nome_do_dia"] != "Domingo"]

    checkin_SEG = primeirocheckin[primeirocheckin["nome_do_dia"] == "Segunda-feira"]

    checkin_SEG = checkin_SEG.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin",
            "Último Checkout OK": "ultimo_checkout",
        }
    )

    checkin_SEG = checkin_SEG[["Colaborador", "primeiro_checkin", "ultimo_checkout"]]

    checkin_TER = primeirocheckin[primeirocheckin["nome_do_dia"] == "Terça-feira"]

    checkin_TER = checkin_TER.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin",
            "Último Checkout OK": "ultimo_checkout",
        }
    )

    checkin_TER = checkin_TER[["Colaborador", "primeiro_checkin", "ultimo_checkout"]]

    checkin_QUA = primeirocheckin[primeirocheckin["nome_do_dia"] == "Quarta-feira"]

    checkin_QUA = checkin_QUA.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin",
            "Último Checkout OK": "ultimo_checkout",
        }
    )

    checkin_QUA = checkin_QUA[["Colaborador", "primeiro_checkin", "ultimo_checkout"]]

    checkin_QUI = primeirocheckin[primeirocheckin["nome_do_dia"] == "Quinta-feira"]

    checkin_QUI = checkin_QUI.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin",
            "Último Checkout OK": "ultimo_checkout",
        }
    )

    checkin_QUI = checkin_QUI[["Colaborador", "primeiro_checkin", "ultimo_checkout"]]

    checkin_SEX = primeirocheckin[primeirocheckin["nome_do_dia"] == "Sexta-feira"]

    checkin_SEX = checkin_SEX.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin",
            "Último Checkout OK": "ultimo_checkout",
        }
    )

    checkin_SEX = checkin_SEX[["Colaborador", "primeiro_checkin", "ultimo_checkout"]]

    checkin_SABD = primeirocheckin[primeirocheckin["nome_do_dia"] == "Sábado"]

    checkin_SABD = checkin_SABD.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin",
            "Último Checkout OK": "ultimo_checkout",
        }
    )

    checkin_SABD = checkin_SABD[["Colaborador", "primeiro_checkin", "ultimo_checkout"]]

    checkin_sab = primeirocheckin[primeirocheckin["nome_do_dia"] == "Sábado"]

    checkin_sab = checkin_sab.rename(
        columns={
            "Primeiro Checkin OK": "primeiro_checkin_sab",
            "Último Checkout OK": "ultimo_checkout_sab",
        }
    )

    checkin_sab = checkin_sab[
        ["Colaborador", "primeiro_checkin_sab", "ultimo_checkout_sab"]
    ]

    checkin_seg_a_sex = checkin_seg_a_sex.groupby(["Colaborador"], as_index=False).agg(
        {"Primeiro Checkin OK": "mean", "Último Checkout OK": "mean"}
    )

    PDOH_P = pd.merge(PDOH_P, checkin_seg_a_sex, on=["Colaborador"], how="left")

    PDOH_P = pd.merge(PDOH_P, checkin_sab, on=["Colaborador"], how="left")

    # PDOH_P = PDOH_P[["Colaborador", "deslocamento", "ocio", "produtividade", "Horas N Registradas", "Total Geral", "Primeiro Checkin OK", "Último Checkout OK", "primeiro_checkin_sab", "ultimo_checkout_sab"]]
    print(PDOH_P)
    #######################################################################################################################################################

    visitasgerencial = gerencial[
        ["Colaborador", "Tipo de Check-in", "Tempo em Loja Executado"]
    ]

    visitasgerencial["visitas_semanais"] = "1"
    visitasgerencial["percentual_visitas"] = "1"
    visitasgerencial["MANUAL E GPS"] = "0"
    visitasgerencial["MANUAL"] = "0"
    visitasgerencial["SEM CHECKIN"] = "0"

    visitasgerencial.loc[
        visitasgerencial["Tipo de Check-in"] == "Sem Checkin", "percentual_visitas"
    ] = "0"
    visitasgerencial.loc[
        visitasgerencial["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
    ] = "1"
    visitasgerencial.loc[visitasgerencial["Tipo de Check-in"] == "Manual", "MANUAL"] = (
        "1"
    )
    visitasgerencial.loc[
        visitasgerencial["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
    ] = "1"

    # transformando em int() os valores das colunas que atribuimos.
    visitasgerencial["visitas_semanais"] = visitasgerencial["visitas_semanais"].astype(
        int
    )
    visitasgerencial["percentual_visitas"] = visitasgerencial[
        "percentual_visitas"
    ].astype(int)
    visitasgerencial["MANUAL E GPS"] = visitasgerencial["MANUAL E GPS"].astype(int)
    visitasgerencial["MANUAL"] = visitasgerencial["MANUAL"].astype(int)
    visitasgerencial["SEM CHECKIN"] = visitasgerencial["SEM CHECKIN"].astype(int)

    porcentagem_visitas = visitasgerencial[
        ["Colaborador", "visitas_semanais", "percentual_visitas"]
    ]

    porcentagem_visitas = porcentagem_visitas.groupby(["Colaborador"]).agg(
        {"visitas_semanais": "sum", "percentual_visitas": "sum"}
    )

    porcentagem_visitas["% de Visitas R"] = (
        porcentagem_visitas["percentual_visitas"]
        / porcentagem_visitas["visitas_semanais"]
    )

    porcentagem_visitas = porcentagem_visitas.rename(
        {
            "visitas_semanais": "visitas_semanais",
            "percentual_visitas": "visitas_semanais_realizadas",
            "% de Visitas R": "percentual_visitas",
        },
        axis=1,
    )

    TESTEKX = pd.merge(PDOH_P, porcentagem_visitas, on=["Colaborador"], how="left")

    md_tempoemloja = visitasgerencial[["Colaborador", "Tempo em Loja Executado"]]
    md_tempoemloja["Tempo em Loja Executado"] = pd.to_timedelta(
        md_tempoemloja["Tempo em Loja Executado"]
    )

    md_tempoemloja = md_tempoemloja.groupby(["Colaborador"], as_index=False).agg(
        {"Tempo em Loja Executado": "mean"}
    )

    md_tempoemloja = md_tempoemloja.rename(
        {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
    )

    TESTEKX = pd.merge(TESTEKX, md_tempoemloja, on=["Colaborador"], how="left")

    TESTEKX["PRODUTIVIDADE"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

    TESTEKX["DESLOCAMENTO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

    TESTEKX["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

    # TESTEKX["PROD_SOPRANO"] = TESTEKX["PRODUTIVIDADE"] + TESTEKX["ÓCIO"]

    # TESTEKX["PRODUTIVIDADE"] = TESTEKX["PROD_SOPRANO"]

    TESTEKX["PRODUTIVIDADE"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

    TESTEKX.loc[
        TESTEKX["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
        "% produtividade",
    ] = (
        TESTEKX["PRODUTIVIDADE"] / TESTEKX["horas_programadas"]
    )

    TESTEKX.loc[
        (TESTEKX["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
        & (TESTEKX["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))),
        "% produtividade",
    ] = int(0)

    pesquisasporcentagem = pesquisasporcentagem_x[["Responsável", "Status"]]

    pesquisasporcentagem["pesquisas_semanais"] = "1"
    pesquisasporcentagem["% de Pesquisasx"] = "1"

    pesquisasporcentagem.loc[
        pesquisasporcentagem["Status"] != "Respondida", "% de Pesquisasx"
    ] = "0"

    # transformando em int() os valores das colunas que atribuimos.

    pesquisasporcentagem["pesquisas_semanais"] = pesquisasporcentagem[
        "pesquisas_semanais"
    ].astype(int)
    pesquisasporcentagem["% de Pesquisasx"] = pesquisasporcentagem[
        "% de Pesquisasx"
    ].astype(int)

    pesquisasporcentagem = pesquisasporcentagem.rename(
        {"Responsável": "Colaborador"}, axis=1
    )

    pesquisasporcentagem = pesquisasporcentagem[
        ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
    ]

    pesquisasporcentagem = pesquisasporcentagem.groupby(["Colaborador"]).agg(
        {"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"}
    )

    pesquisasporcentagem["percentual_pesquisas"] = (
        pesquisasporcentagem["% de Pesquisasx"]
        / pesquisasporcentagem["pesquisas_semanais"]
    )

    pesquisasporcentagem = pesquisasporcentagem[
        ["pesquisas_semanais", "percentual_pesquisas"]
    ]

    TESTEKX = pd.merge(TESTEKX, pesquisasporcentagem, on=["Colaborador"], how="left")

    TESTEKX["pesquisas_semanais"].fillna(int(0), inplace=True)
    TESTEKX["percentual_pesquisas"].fillna(int(0), inplace=True)
    # TESTEKX["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

    TESTEKX["visitas_semanais"].fillna(int(0), inplace=True)
    TESTEKX["percentual_visitas"].fillna(int(0), inplace=True)

    PDTT = TESTEKX["PRODUTIVIDADE"].sum()

    DSLCT = TESTEKX["DESLOCAMENTO"].sum()

    OCT = TESTEKX["ÓCIO"].sum()

    TT = TESTEKX["TOTAL"].sum()

    HPTT = TESTEKX["horas_programadas"].sum()

    HNRT = TESTEKX["H N REGISTRADAS"].sum()

    # Condição para evitar divisão por zero
    if HPTT != pd.to_timedelta("0 days 00:00:00"):
        PDESLOC = DSLCT / HPTT
    else:
        PDESLOC = pd.to_timedelta("0 days 00:00:00")

    if HPTT != pd.to_timedelta("0 days 00:00:00"):
        POCIO = OCT / HPTT
    else:
        POCIO = pd.to_timedelta("0 days 00:00:00")

    if HPTT != pd.to_timedelta("0 days 00:00:00"):
        PPROD = PDTT / HPTT
    else:
        PPROD = pd.to_timedelta("0 days 00:00:00")

    if HPTT != pd.to_timedelta("0 days 00:00:00"):
        PHNR = HNRT / HPTT
    else:
        PHNR = pd.to_timedelta("0 days 00:00:00")

    if HPTT != pd.to_timedelta("0 days 00:00:00"):
        PHT = HPTT / HPTT
    else:
        PHT = pd.to_timedelta("0 days 00:00:00")

    TESTEKX["pesquisas_semanais_realizadas"] = (
        TESTEKX["pesquisas_semanais"] * TESTEKX["percentual_pesquisas"]
    )
    TESTEKX.replace([float("inf"), -float("inf")], 0, inplace=True)
    TESTEKX["pesquisas_semanais_realizadas"] = TESTEKX[
        "pesquisas_semanais_realizadas"
    ].fillna(int(0))
    TESTEKX["pesquisas_semanais_realizadas"] = TESTEKX[
        "pesquisas_semanais_realizadas"
    ].astype(int)

    SVS = TESTEKX["visitas_semanais"].sum()
    SVSR = TESTEKX["visitas_semanais_realizadas"].sum()
    SPS = TESTEKX["pesquisas_semanais"].sum()
    SPSR = TESTEKX["pesquisas_semanais_realizadas"].sum()

    TESTEKX.loc[
        TESTEKX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "visitas_semanais",
    ] = int(0)
    TESTEKX.loc[
        TESTEKX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "visitas_semanais_realizadas",
    ] = int(0)
    TESTEKX.loc[
        TESTEKX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "pesquisas_semanais_realizadas",
    ] = int(0)
    TESTEKX.loc[
        TESTEKX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "pesquisas_semanais",
    ] = int(0)

    # Condição para evitar divisão por zero
    if HPTT != pd.to_timedelta("0 days 00:00:00"):
        MdPROD = PDTT / HPTT
    else:
        MdPROD = int(0)

    if SVS != pd.to_timedelta("0 days 00:00:00"):
        MdVST = SVSR / SVS
    else:
        MdVST = int(0)

    if SPS != pd.to_timedelta("0 days 00:00:00"):
        MdPSQ = SPSR / SPS
    else:
        MdPSQ = int(0)

    TESTEKX = TESTEKX.rename(
        {
            "Primeiro Checkin OK": "media_primeiro_checkin",
            "Último Checkout OK": "media_ultimo_checkout",
        },
        axis=1,
    )

    # Criar um novo DataFrame com a linha que deseja adicionar
    nova_linha = pd.DataFrame(
        {
            "Colaborador": ["TOTAL"],
            "SUPERIOR": ["-"],
            "UF": ["-"],
            "PRODUTIVIDADE": [PDTT],
            "DESLOCAMENTO": [DSLCT],
            "ÓCIO": [OCT],
            "TOTAL": [TT],
            "horas_programadas": [HPTT],
            "H N REGISTRADAS": [HNRT],
            "almoco": "-",
            "Média de tempo em loja": "-",
            "media_primeiro_checkin": "-",
            "media_ultimo_checkout": "-",
            "primeiro_checkin_sab": "-",
            "ultimo_checkout_sab": "-",
            "visitas_semanais": [SVS],
            "visitas_semanais_realizadas": [SVSR],
            "pesquisas_semanais": [SPS],
            "pesquisas_semanais_realizadas": [SPSR],
            "% produtividade": [MdPROD],
            "percentual_visitas": [MdVST],
            "percentual_pesquisas": [MdPSQ],
            "justificativas": ["-"],
        }
    )

    # Usar a função concat para adicionar a nova linha ao DataFrame
    TESTEKX = pd.concat([TESTEKX, nova_linha], ignore_index=True)

    TESTEKX.loc[TESTEKX["visitas_semanais"] == 0, "percentual_visitas"] = "-"
    TESTEKX.loc[TESTEKX["pesquisas_semanais"] == 0, "percentual_pesquisas"] = "-"

    # Step 2: Create a dictionary to hold the pesos for each indicator
    pesos = dict(zip(pesos["Indicadores"], pesos["Pesos"]))

    # Convertendo as colunas para tipos numéricos
    TESTEKX["% produtividade"] = pd.to_numeric(
        TESTEKX["% produtividade"], errors="coerce"
    )
    TESTEKX["percentual_visitas"] = pd.to_numeric(
        TESTEKX["percentual_visitas"], errors="coerce"
    )
    TESTEKX["percentual_pesquisas"] = pd.to_numeric(
        TESTEKX["percentual_pesquisas"], errors="coerce"
    )

    # Verificando se há valores não numéricos que foram convertidos para NaN
    if (
        TESTEKX[["% produtividade", "percentual_visitas", "percentual_pesquisas"]]
        .isnull()
        .values.any()
    ):
        print(
            "Existem valores não numéricos nas colunas que foram convertidos para NaN."
        )
        print(TESTEKX)
    # Step 3: Calculate the weighted average using the pesos from the spreadsheet
    # Assuming TESTEKX is a pandas DataFrame that has the columns "% produtividade", "percentual_visitas", and "percentual_pesquisas"

    # Commenting out the actual calculation since the DataFrame is not defined here.
    TESTEKX["percentual_efetividade"] = (
        TESTEKX["% produtividade"] * pesos["Produtividade"]
        + TESTEKX["percentual_visitas"] * pesos["Visitas"]
        + TESTEKX["percentual_pesquisas"] * pesos["Pesquisas"]
    ) / sum(pesos.values())

    # Criar um novo DataFrame com a linha que deseja adicionar
    nova_linha2 = pd.DataFrame(
        {
            "Colaborador": ["-"],
            "SUPERIOR": ["-"],
            "UF": ["-"],
            "PRODUTIVIDADE": [PPROD],
            "DESLOCAMENTO": [PDESLOC],
            "ÓCIO": [POCIO],
            "H N REGISTRADAS": [PHNR],
            "horas_programadas": [PHT],
            "almoco": "-",
            "Média de tempo em loja": "-",
            "media_primeiro_checkin": "-",
            "media_ultimo_checkout": "-",
            "primeiro_checkin_sab": "-",
            "ultimo_checkout_sab": "-",
            "visitas_semanais": ["-"],
            "visitas_semanais_realizadas": ["-"],
            "pesquisas_semanais": ["-"],
            "pesquisas_semanais_realizadas": ["-"],
            "% produtividade": ["-"],
            "percentual_visitas": ["-"],
            "percentual_pesquisas": ["-"],
            "percentual_efetividade": ["-"],
            "justificativas": ["-"],
        }
    )

    # Usar a função concat para adicionar a nova linha ao DataFrame
    PDOHCOMPLETO = pd.concat([TESTEKX, nova_linha2], ignore_index=True)

    PDOHCOMPLETO = PDOHCOMPLETO.rename(
        {
            "Colaborador": "colaborador",
            "SUPERIOR": "superior",
            "UF": "estado",
            "DESLOCAMENTO": "deslocamento",
            "ÓCIO": "ocio",
            "PRODUTIVIDADE": "produtividade",
            "H N REGISTRADAS": "horas_nao_registradas",
            "horas_programadas": "horas_programadas",
            "Média de tempo em loja": "media_tempo_em_loja",
            "media_primeiro_checkin": "media_primeiro_checkin",
            "% produtividade": "percentual_produtividade",
            "percentual_visitas": "percentual_visitas",
            "percentual_pesquisas": "percentual_pesquisas",
            "percentual_efetividade": "percentual_efetividade",
            "justificativas": "justificativas",
        },
        axis=1,
    )

    PDOHCOMPLETO.loc[
        PDOHCOMPLETO["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_produtividade",
    ] = "-"
    PDOHCOMPLETO.loc[
        PDOHCOMPLETO["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_visitas",
    ] = "-"
    PDOHCOMPLETO.loc[
        PDOHCOMPLETO["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_pesquisas",
    ] = "-"
    PDOHCOMPLETO.loc[
        PDOHCOMPLETO["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_efetividade",
    ] = "-"

    PDOHCOMPLETO = PDOHCOMPLETO[
        [
            "colaborador",
            "superior",
            "estado",
            "deslocamento",
            "ocio",
            "produtividade",
            "horas_nao_registradas",
            "horas_programadas",
            "almoco",
            "media_tempo_em_loja",
            "media_primeiro_checkin",
            "media_ultimo_checkout",
            "primeiro_checkin_sab",
            "ultimo_checkout_sab",
            "visitas_semanais",
            "visitas_semanais_realizadas",
            "pesquisas_semanais",
            "pesquisas_semanais_realizadas",
            "percentual_produtividade",
            "percentual_visitas",
            "percentual_pesquisas",
            "percentual_efetividade",
            "justificativas",
        ]
    ]

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------PDOH SEMANAL COMPLETO--------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    # ----------------------- Criando variável vazia para alocar todos os PDOHs diários
    concat_analise = []

    # --------------------------------------------------------------------------------------------------------------------
    # -------------------------------------------------SEGUNDA-FEIRA------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------
    if True in dias_da_semana["nome_do_dia"].isin(["Segunda-feira"]).tolist():
        PDOH_SEG = PDOH.loc[PDOH["nome_do_dia"] == "Segunda-feira"]

        def AnalisePDOH(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador", "data"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida  #####################################################

        def AnalisePor(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida

        Ocio_SEG = AnalisePDOH(PDOH_SEG, "Ocio")
        print(Ocio_SEG)

        teste_x_SEG = Ocio_SEG.iloc[0, 1]


        #Criando variável para alocar o corte do almoco dentro do Ocio
        ocio_sem_almoco_SEG = Ocio_SEG
        ocio_sem_almoco_SEG = pd.merge(ocio_sem_almoco_SEG, colaboradores_SEG, on="Colaborador", how="right")
        ocio_sem_almoco_SEG = ocio_sem_almoco_SEG[["Colaborador", "Ocio", "almoco"]]
        ocio_sem_almoco_SEG = ocio_sem_almoco_SEG.rename({"almoco":"almoco"}, axis=1)
        ocio_sem_almoco_SEG["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)
        ocio_sem_almoco_SEG["Ocio_sem_almoco"] = ocio_sem_almoco_SEG["Ocio"] - ocio_sem_almoco_SEG["almoco"]
        ocio_sem_almoco_SEG.loc[ocio_sem_almoco_SEG["Ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco_SEG["almoco"] - ocio_sem_almoco_SEG["Ocio"]
        ocio_sem_almoco_SEG.loc[ocio_sem_almoco_SEG["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")
        ocio_sem_almoco_SEG.loc[ocio_sem_almoco_SEG["Ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco_SEG["Ocio_sem_almoco"]
        ocio_sem_almoco_SEG.loc[ocio_sem_almoco_SEG["Ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")
        ocio_sem_almoco_SEG["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Ocio_SEG = ocio_sem_almoco_SEG[["Colaborador", "Ocio_real"]]
        Ocio_SEG.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)



        Deslocamento_SEG = AnalisePor(PDOH_SEG,"Deslocamento")

        desloc_com_cortes_SEG = pd.merge(Deslocamento_SEG, ocio_sem_almoco_SEG, on="Colaborador", how="right")
        desloc_com_cortes_SEG["Deslocamento_pos_corte"] = desloc_com_cortes_SEG["Deslocamento"] - desloc_com_cortes_SEG["Corte_pendente"]
        desloc_com_cortes_SEG["corte_pendente_prod"] = desloc_com_cortes_SEG["Corte_pendente"] - desloc_com_cortes_SEG["Deslocamento"]
        desloc_com_cortes_SEG.loc[desloc_com_cortes_SEG["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")
        desloc_com_cortes_SEG.loc[desloc_com_cortes_SEG["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes_SEG["Deslocamento_pos_corte"]
        desloc_com_cortes_SEG.loc[desloc_com_cortes_SEG["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

        Deslocamento_SEG = desloc_com_cortes_SEG[["Colaborador", "Deslocamento"]]


        Produtividadex_SEG = AnalisePor(PDOH_SEG,"Produtividade")
        Produtividadex_SEG = pd.merge(Produtividadex_SEG, desloc_com_cortes_SEG, how = "right", on = "Colaborador")
        Produtividadex_SEG.loc[Produtividadex_SEG["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex_SEG["Produtividade"] - Produtividadex_SEG["corte_pendente_prod"]
        Produtividadex_SEG.loc[Produtividadex_SEG["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")
        Produtividadex_SEG.loc[Produtividadex_SEG["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")
        Produtividadex_SEG["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)
        Produtividade_SEG = Produtividadex_SEG[["Colaborador", "Produtividade"]]


        htr_SEG = pd.merge(
            Produtividade_SEG, Deslocamento_SEG, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_SEG = pd.merge(
            htr_SEG, Ocio_SEG, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_SEG["Tempo Investido"] = (
            HorasTotaisRegistradas_SEG["Produtividade"]
            + HorasTotaisRegistradas_SEG["Deslocamento"]
            + HorasTotaisRegistradas_SEG["Ocio"]
        )

        Horastotaisxxxx_SEG = pd.merge(
            HorasTotaisRegistradas_SEG, colaboradores_SEG, how="right", on="Colaborador"
        )

        Horastotaisxxxx_SEG["justificativas_seg"].fillna(" ", inplace=True)

        Horastotaisxxxx_SEG.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Horastotaisxxxx_SEG["HT sem almoco"] = Horastotaisxxxx_SEG["Tempo Investido"]

        JornadaDeTrabalho_SEG = Horastotaisxxxx_SEG[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "HT sem almoco",
                "Jornada de Trabalho",
                "almoco",
                "justificativas_seg",
            ]
        ]

        JornadaDeTrabalho_SEG["Hora Excedente"] = (
            JornadaDeTrabalho_SEG["HT sem almoco"]
            - JornadaDeTrabalho_SEG["Jornada de Trabalho"]
        )

        JornadaDeTrabalho_SEG["Hora Pendente"] = (
            JornadaDeTrabalho_SEG["Jornada de Trabalho"]
            - JornadaDeTrabalho_SEG["HT sem almoco"]
        )

        JornadaDeTrabalho_SEG.loc[
            JornadaDeTrabalho_SEG["Hora Excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Hora Excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        JornadaDeTrabalho_SEG.loc[
            JornadaDeTrabalho_SEG["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
            "Hora Pendente",
        ] = pd.to_timedelta("0 days 00:00:00")

        Horasnaoregistradas_SEG = JornadaDeTrabalho_SEG[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Hora Pendente",
                "almoco",
                "justificativas_seg",
            ]
        ]

        corteszx_SEG = pd.merge(
            JornadaDeTrabalho_SEG, Deslocamento_SEG, how="left", on="Colaborador"
        )

        corteszx_SEG["Deslocamento_cortado"] = (
            corteszx_SEG["Deslocamento"] - corteszx_SEG["Hora Excedente"]
        )

        corteszx_SEG["Deslocamento_cortado_excedente"] = (
            corteszx_SEG["Hora Excedente"] - corteszx_SEG["Deslocamento"]
        )

        corteszx_SEG.loc[
            corteszx_SEG["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        corteszx_SEG.loc[
            corteszx_SEG["Deslocamento_cortado_excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        cortesyy_SEG = pd.merge(corteszx_SEG, Ocio_SEG, how="outer", on="Colaborador")

        cortesyy_SEG["Ocio_cortado"] = (
            cortesyy_SEG["Ocio"] - cortesyy_SEG["Deslocamento_cortado_excedente"]
        )

        cortesyy_SEG["Ocio_cortado_excedente"] = (
            cortesyy_SEG["Deslocamento_cortado_excedente"] - cortesyy_SEG["Ocio"]
        )

        cortesyy_SEG.loc[
            cortesyy_SEG["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        cortesyy_SEG.loc[
            cortesyy_SEG["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_SEG = pd.merge(
            cortesyy_SEG, Produtividade_SEG, how="right", on="Colaborador"
        )

        corteszz_SEG["Produtividade_cortada"] = (
            corteszz_SEG["Produtividade"] - corteszz_SEG["Ocio_cortado_excedente"]
        )

        corteszz_SEG.loc[
            corteszz_SEG["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
            "Produtividade_cortada",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_SEG["TOTAL DE HR"] = (
            corteszz_SEG["Produtividade_cortada"]
            + corteszz_SEG["Deslocamento_cortado"]
            + corteszz_SEG["Ocio_cortado"]
        )

        cortes_SEG = corteszz_SEG[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Produtividade_cortada",
                "Deslocamento_cortado",
                "Ocio_cortado",
                "TOTAL DE HR",
                "Jornada de Trabalho",
                "Hora Pendente",
                "Hora Excedente",
                "almoco",
                "justificativas_seg",
            ]
        ]

        cortes_SEG = cortes_SEG.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
        cortes_SEG = cortes_SEG.rename({"almoco": "almoco"}, axis=1)
        cortes_SEG = cortes_SEG.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

        PDOH_P_SEG = cortes_SEG.rename(
            {
                "Produtividade_cortada": "PRODUTIVIDADE",
                "Deslocamento_cortado": "DESLOCAMENTO",
                "Ocio_cortado": "ÓCIO",
                "TOTAL DE HR": "TOTAL",
                "Hora Pendente": "H N REGISTRADAS",
                "Jornada de Trabalho": "horas_programadas",
            },
            axis=1,
        )

        PDOH_P_SEG = pd.merge(PDOH_P_SEG, checkin_SEG, on=["Colaborador"], how="left")

        ########
        visitasgerencial_SEG = gerencial[
            [
                "Colaborador",
                "Data da Visita",
                "Tipo de Check-in",
                "Tempo em Loja Executado",
            ]
        ]

        visitasgerencial_SEG["Data da Visita"] = pd.to_datetime(
            visitasgerencial_SEG["Data da Visita"], format="%d/%m/%Y"
        )

        dias_ptbr = {
            "Sunday": "Domingo",
            "Monday": "Segunda-feira",
            "Tuesday": "Terça-feira",
            "Wednesday": "Quarta-feira",
            "Thursday": "Quinta-feira",
            "Friday": "Sexta-feira",
            "Saturday": "Sábado",
        }

        visitasgerencial_SEG["nome_do_dia"] = (
            visitasgerencial_SEG["Data da Visita"].dt.day_name().replace(dias_ptbr)
        )

        visitasgerencial_SEG = visitasgerencial_SEG.loc[
            visitasgerencial_SEG["nome_do_dia"] == "Segunda-feira"
        ]

        ########

        visitasgerencial_SEG["visitas_diarias"] = "1"
        visitasgerencial_SEG["percentual_visitas"] = "1"
        visitasgerencial_SEG["MANUAL E GPS"] = "0"
        visitasgerencial_SEG["MANUAL"] = "0"
        visitasgerencial_SEG["SEM CHECKIN"] = "0"

        visitasgerencial_SEG.loc[
            visitasgerencial_SEG["Tipo de Check-in"] == "Sem Checkin",
            "percentual_visitas",
        ] = "0"
        visitasgerencial_SEG.loc[
            visitasgerencial_SEG["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
        ] = "1"
        visitasgerencial_SEG.loc[
            visitasgerencial_SEG["Tipo de Check-in"] == "Manual", "MANUAL"
        ] = "1"
        visitasgerencial_SEG.loc[
            visitasgerencial_SEG["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
        ] = "1"

        # transformando em int() os valores das colunas que atribuimos.

        visitasgerencial_SEG["visitas_diarias"] = visitasgerencial_SEG[
            "visitas_diarias"
        ].astype(int)
        visitasgerencial_SEG["percentual_visitas"] = visitasgerencial_SEG[
            "percentual_visitas"
        ].astype(int)
        visitasgerencial_SEG["MANUAL E GPS"] = visitasgerencial_SEG[
            "MANUAL E GPS"
        ].astype(int)
        visitasgerencial_SEG["MANUAL"] = visitasgerencial_SEG["MANUAL"].astype(int)
        visitasgerencial_SEG["SEM CHECKIN"] = visitasgerencial_SEG[
            "SEM CHECKIN"
        ].astype(int)

        porcentagem_visitas_SEG = visitasgerencial_SEG[
            ["Colaborador", "visitas_diarias", "percentual_visitas"]
        ]

        porcentagem_visitas_SEG = porcentagem_visitas_SEG.groupby(["Colaborador"]).agg(
            {"visitas_diarias": "sum", "percentual_visitas": "sum"}
        )

        porcentagem_visitas_SEG["% de Visitas R"] = (
            porcentagem_visitas_SEG["percentual_visitas"]
            / porcentagem_visitas_SEG["visitas_diarias"]
        )

        porcentagem_visitas_SEG = porcentagem_visitas_SEG.rename(
            {
                "visitas_diarias": "visitas_diarias",
                "percentual_visitas": "visitas_diarias_realizadas",
                "% de Visitas R": "percentual_visitas",
            },
            axis=1,
        )

        TESTEKX_SEG = pd.merge(
            PDOH_P_SEG, porcentagem_visitas_SEG, on=["Colaborador"], how="left"
        )

        md_tempoemloja_SEG = visitasgerencial_SEG[
            ["Colaborador", "Tempo em Loja Executado"]
        ]
        md_tempoemloja_SEG["Tempo em Loja Executado"] = pd.to_timedelta(
            md_tempoemloja_SEG["Tempo em Loja Executado"]
        )

        md_tempoemloja_SEG = md_tempoemloja_SEG.groupby(
            ["Colaborador"], as_index=False
        ).agg({"Tempo em Loja Executado": "mean"})

        md_tempoemloja_SEG = md_tempoemloja_SEG.rename(
            {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
        )

        TESTEKX_SEG = pd.merge(
            TESTEKX_SEG, md_tempoemloja_SEG, on=["Colaborador"], how="left"
        )

        TESTEKX_SEG["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SEG["DESLOCAMENTO"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SEG["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        # TESTEKX_SEG["PROD_SOPRANO"] = TESTEKX_SEG["PRODUTIVIDADE"] + TESTEKX_SEG["ÓCIO"]

        # TESTEKX_SEG["PRODUTIVIDADE"] = TESTEKX_SEG["PROD_SOPRANO"]

        TESTEKX_SEG["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SEG.loc[
            TESTEKX_SEG["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
            "% produtividade",
        ] = (
            TESTEKX_SEG["PRODUTIVIDADE"] / TESTEKX_SEG["horas_programadas"]
        )

        TESTEKX_SEG.loc[
            (TESTEKX_SEG["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
            & (
                TESTEKX_SEG["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))
            ),
            "% produtividade",
        ] = int(0)

        ########
        pesquisasporcentagem_SEG = pesquisasporcentagem_x[
            ["Responsável", "Status", "Data de Expiração"]
        ]

        pesquisasporcentagem_SEG["Data de Expiração"] = pd.to_datetime(
            pesquisasporcentagem_SEG["Data de Expiração"]
        )

        pesquisasporcentagem_SEG["nome_do_dia"] = (
            pesquisasporcentagem_SEG["Data de Expiração"]
            .dt.day_name()
            .replace(dias_ptbr)
        )

        pesquisasporcentagem_SEG = pesquisasporcentagem_SEG.loc[
            pesquisasporcentagem_SEG["nome_do_dia"] == "Segunda-feira"
        ]

        ########

        pesquisasporcentagem_SEG["pesquisas_semanais"] = "1"
        pesquisasporcentagem_SEG["% de Pesquisasx"] = "1"

        pesquisasporcentagem_SEG.loc[
            pesquisasporcentagem_SEG["Status"] != "Respondida", "% de Pesquisasx"
        ] = "0"

        # transformando em int() os valores das colunas que atribuimos.

        pesquisasporcentagem_SEG["pesquisas_semanais"] = pesquisasporcentagem_SEG[
            "pesquisas_semanais"
        ].astype(int)
        pesquisasporcentagem_SEG["% de Pesquisasx"] = pesquisasporcentagem_SEG[
            "% de Pesquisasx"
        ].astype(int)

        pesquisasporcentagem_SEG = pesquisasporcentagem_SEG.rename(
            {"Responsável": "Colaborador"}, axis=1
        )

        pesquisasporcentagem_SEG = pesquisasporcentagem_SEG[
            ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
        ]

        pesquisasporcentagem_SEG = pesquisasporcentagem_SEG.groupby(
            ["Colaborador"]
        ).agg({"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"})

        pesquisasporcentagem_SEG["percentual_pesquisas"] = (
            pesquisasporcentagem_SEG["% de Pesquisasx"]
            / pesquisasporcentagem_SEG["pesquisas_semanais"]
        )

        pesquisasporcentagem_SEG = pesquisasporcentagem_SEG[
            ["pesquisas_semanais", "percentual_pesquisas"]
        ]

        TESTEKX_SEG = pd.merge(
            TESTEKX_SEG, pesquisasporcentagem_SEG, on=["Colaborador"], how="left"
        )

        TESTEKX_SEG["pesquisas_diarias_realizadas"] = TESTEKX_SEG[
            "pesquisas_semanais"
        ].astype(float) * TESTEKX_SEG["percentual_pesquisas"].astype(float)

        TESTEKX_SEG["pesquisas_semanais"].fillna(int(0), inplace=True)
        TESTEKX_SEG["percentual_pesquisas"].fillna(int(0), inplace=True)
        TESTEKX_SEG["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

        TESTEKX_SEG["visitas_diarias"].fillna(int(0), inplace=True)
        TESTEKX_SEG["percentual_visitas"].fillna(int(0), inplace=True)

        PDTT_SEG = TESTEKX_SEG["PRODUTIVIDADE"].sum()

        DSLCT_SEG = TESTEKX_SEG["DESLOCAMENTO"].sum()

        OCT_SEG = TESTEKX_SEG["ÓCIO"].sum()

        TT_SEG = TESTEKX_SEG["TOTAL"].sum()

        HPTT_SEG = TESTEKX_SEG["horas_programadas"].sum()

        HNRT_SEG = TESTEKX_SEG["H N REGISTRADAS"].sum()

        # Condição para evitar divisão por zero
        if HPTT_SEG != pd.to_timedelta("0 days 00:00:00"):
            PDESLOC_SEG = DSLCT_SEG / HPTT_SEG
        else:
            PDESLOC_SEG = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEG != pd.to_timedelta("0 days 00:00:00"):
            POCIO_SEG = OCT_SEG / HPTT_SEG
        else:
            POCIO_SEG = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEG != pd.to_timedelta("0 days 00:00:00"):
            PPROD_SEG = PDTT_SEG / HPTT_SEG
        else:
            PPROD_SEG = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEG != pd.to_timedelta("0 days 00:00:00"):
            PHNR_SEG = HNRT_SEG / HPTT_SEG
        else:
            PHNR_SEG = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEG != pd.to_timedelta("0 days 00:00:00"):
            PHT_SEG = HPTT_SEG / HPTT_SEG
        else:
            PHT_SEG = pd.to_timedelta("0 days 00:00:00")

        TESTEKX_SEG["pesquisas_diarias_realizadas"] = (
            TESTEKX_SEG["pesquisas_semanais"] * TESTEKX_SEG["percentual_pesquisas"]
        )
        TESTEKX_SEG.replace([float("inf"), -float("inf")], 0, inplace=True)
        TESTEKX_SEG["pesquisas_diarias_realizadas"] = TESTEKX_SEG[
            "pesquisas_diarias_realizadas"
        ].fillna(int(0))
        TESTEKX_SEG["pesquisas_diarias_realizadas"] = TESTEKX_SEG[
            "pesquisas_diarias_realizadas"
        ].astype(int)

        SVS_SEG = TESTEKX_SEG["visitas_diarias"].sum()
        SVSR_SEG = TESTEKX_SEG["visitas_diarias_realizadas"].sum()
        SPS_SEG = TESTEKX_SEG["pesquisas_semanais"].sum()
        SPSR_SEG = TESTEKX_SEG["pesquisas_diarias_realizadas"].sum()

        TESTEKX_SEG.loc[
            TESTEKX_SEG["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias",
        ] = int(0)
        TESTEKX_SEG.loc[
            TESTEKX_SEG["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias_realizadas",
        ] = int(0)
        TESTEKX_SEG.loc[
            TESTEKX_SEG["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_diarias_realizadas",
        ] = int(0)
        TESTEKX_SEG.loc[
            TESTEKX_SEG["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_semanais",
        ] = int(0)

        # Condição para evitar divisão por zero
        if HPTT_SEG != pd.to_timedelta("0 days 00:00:00"):
            MdPROD_SEG = PDTT_SEG / HPTT_SEG
        else:
            MdPROD_SEG = int(0)

        if SVS_SEG != pd.to_timedelta("0 days 00:00:00"):
            MdVST_SEG = SVSR_SEG / SVS_SEG
        else:
            MdVST_SEG = int(0)

        if SPS_SEG != pd.to_timedelta("0 days 00:00:00"):
            MdPSQ_SEG = SPSR_SEG / SPS_SEG
        else:
            MdPSQ_SEG = int(0)

        TESTEKX_SEG = TESTEKX_SEG.rename(
            {"pesquisas_semanais": "pesquisas_diarias"}, axis=1
        )

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha_SEG = pd.DataFrame(
            {
                "Colaborador": ["TOTAL"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PDTT_SEG],
                "DESLOCAMENTO": [DSLCT_SEG],
                "ÓCIO": [OCT_SEG],
                "TOTAL": [TT_SEG],
                "horas_programadas": [HPTT_SEG],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "H N REGISTRADAS": [HNRT_SEG],
                "visitas_diarias": [SVS_SEG],
                "visitas_diarias_realizadas": [SVSR_SEG],
                "pesquisas_diarias": [SPS_SEG],
                "pesquisas_diarias_realizadas": [SPSR_SEG],
                "% produtividade": [MdPROD_SEG],
                "percentual_visitas": [MdVST_SEG],
                "percentual_pesquisas": [MdPSQ_SEG],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        TESTEKX_SEG = pd.concat([TESTEKX_SEG, nova_linha_SEG], ignore_index=True)

        TESTEKX_SEG.loc[TESTEKX_SEG["visitas_diarias"] == 0, "percentual_visitas"] = "-"
        TESTEKX_SEG.loc[
            TESTEKX_SEG["pesquisas_diarias"] == 0, "percentual_pesquisas"
        ] = "-"

        # Step 2: Create a dictionary to hold the pesos for each indicator
        # pesos = dict(zip(pesos["Indicadores"], pesos["Pesos"]))
        # Convertendo as colunas para tipos numéricos
        TESTEKX_SEG["% produtividade"] = pd.to_numeric(
            TESTEKX_SEG["% produtividade"], errors="coerce"
        )
        TESTEKX_SEG["percentual_visitas"] = pd.to_numeric(
            TESTEKX_SEG["percentual_visitas"], errors="coerce"
        )
        TESTEKX_SEG["percentual_pesquisas"] = pd.to_numeric(
            TESTEKX_SEG["percentual_pesquisas"], errors="coerce"
        )

        # Verificando se há valores não numéricos que foram convertidos para NaN
        if (
            TESTEKX_SEG[
                ["% produtividade", "percentual_visitas", "percentual_pesquisas"]
            ]
            .isnull()
            .values.any()
        ):
            print(
                "Existem valores não numéricos nas colunas que foram convertidos para NaN."
            )
            print(TESTEKX_SEG)
        # Step 3: Calculate the weighted average using the pesos from the spreadsheet
        # Assuming TESTEKX_SEG is a pandas DataFrame that has the columns "% produtividade", "percentual_visitas", and "percentual_pesquisas"

        # Commenting out the actual calculation since the DataFrame is not defined here.
        TESTEKX_SEG["percentual_efetividade"] = (
            TESTEKX_SEG["% produtividade"] * pesos["Produtividade"]
            + TESTEKX_SEG["percentual_visitas"] * pesos["Visitas"]
            + TESTEKX_SEG["percentual_pesquisas"] * pesos["Pesquisas"]
        ) / sum(pesos.values())

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha2_SEG = pd.DataFrame(
            {
                "Colaborador": ["-"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PPROD_SEG],
                "DESLOCAMENTO": [PDESLOC_SEG],
                "ÓCIO": [POCIO_SEG],
                "H N REGISTRADAS": [PHNR_SEG],
                "horas_programadas": [PHT_SEG],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "visitas_diarias": ["-"],
                "visitas_diarias_realizadas": ["-"],
                "pesquisas_diarias": ["-"],
                "pesquisas_diarias_realizadas": ["-"],
                "% produtividade": ["-"],
                "percentual_visitas": ["-"],
                "percentual_pesquisas": ["-"],
                "percentual_efetividade": ["-"],
                "justificativas_seg": ["-"],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        PDOHCOMPLETO_SEG = pd.concat([TESTEKX_SEG, nova_linha2_SEG], ignore_index=True)

        PDOHCOMPLETO_SEG["nome_do_dia"] = "Segunda-feira"
        PDOHCOMPLETO_SEG["data"] = teste_x_SEG

        PDOHCOMPLETO_SEG = PDOHCOMPLETO_SEG.rename(
            {
                "Colaborador": "colaborador",
                "SUPERIOR": "superior",
                "UF": "estado",
                "DESLOCAMENTO": "deslocamento",
                "ÓCIO": "ocio",
                "PRODUTIVIDADE": "produtividade",
                "H N REGISTRADAS": "horas_nao_registradas",
                "horas_programadas": "horas_programadas",
                "Média de tempo em loja": "media_tempo_em_loja",
                "media_ultimo_checkout": "media_ultimo_checkout",
                "% produtividade": "percentual_produtividade",
                "percentual_efetividade": "percentual_efetividade",
            },
            axis=1,
        )

        PDOHCOMPLETO_SEG.loc[
            PDOHCOMPLETO_SEG["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_produtividade",
        ] = "-"
        PDOHCOMPLETO_SEG.loc[
            PDOHCOMPLETO_SEG["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_visitas",
        ] = "-"
        PDOHCOMPLETO_SEG.loc[
            PDOHCOMPLETO_SEG["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_pesquisas",
        ] = "-"
        PDOHCOMPLETO_SEG.loc[
            PDOHCOMPLETO_SEG["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_efetividade",
        ] = "-"

        PDOHCOMPLETO_SEG_XXX = PDOHCOMPLETO_SEG[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_seg",
            ]
        ]

        PDOHCOMPLETO_SEG = PDOHCOMPLETO_SEG[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_seg",
            ]
        ]

        concat_analise.append(PDOHCOMPLETO_SEG_XXX)

    # --------------------------------------------------------------------------------------------------------------------
    # -------------------------------------------------SEGUNDA-FEIRA------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------TERÇA-FEIRA------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    if True in dias_da_semana["nome_do_dia"].isin(["Terça-feira"]).tolist():
        PDOH_TER = PDOH.loc[PDOH["nome_do_dia"] == "Terça-feira"]

        def AnalisePDOH(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador", "data"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida  #####################################################

        def AnalisePor(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida

        Ocio_TER = AnalisePDOH(PDOH_TER, "Ocio")
        print(Ocio_TER)

        teste_x_TER = Ocio_TER.iloc[0, 1]


        #Criando variável para alocar o corte do almoco dentro do Ocio
        ocio_sem_almoco_TER = Ocio_TER

        ocio_sem_almoco_TER = pd.merge(ocio_sem_almoco_TER, colaboradores_TER, on="Colaborador", how="right")

        ocio_sem_almoco_TER = ocio_sem_almoco_TER[["Colaborador", "Ocio", "almoco"]]

        ocio_sem_almoco_TER = ocio_sem_almoco_TER.rename({"almoco":"almoco"}, axis=1)

        ocio_sem_almoco_TER["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        ocio_sem_almoco_TER["Ocio_sem_almoco"] = ocio_sem_almoco_TER["Ocio"] - ocio_sem_almoco_TER["almoco"]

        ocio_sem_almoco_TER.loc[ocio_sem_almoco_TER["Ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco_TER["almoco"] - ocio_sem_almoco_TER["Ocio"]

        ocio_sem_almoco_TER.loc[ocio_sem_almoco_TER["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_TER.loc[ocio_sem_almoco_TER["Ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco_TER["Ocio_sem_almoco"]
        ocio_sem_almoco_TER.loc[ocio_sem_almoco_TER["Ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_TER["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Ocio_TER = ocio_sem_almoco_TER[["Colaborador", "Ocio_real"]]

        Ocio_TER.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)


        # --------------------------------------- DESLOCAMENTO
        Deslocamento_TER = AnalisePor(PDOH_TER,"Deslocamento")

        desloc_com_cortes_TER = pd.merge(Deslocamento_TER, ocio_sem_almoco_TER, on="Colaborador", how="right")

        desloc_com_cortes_TER["Deslocamento_pos_corte"] = desloc_com_cortes_TER["Deslocamento"] - desloc_com_cortes_TER["Corte_pendente"]

        desloc_com_cortes_TER["corte_pendente_prod"] = desloc_com_cortes_TER["Corte_pendente"] - desloc_com_cortes_TER["Deslocamento"]

        desloc_com_cortes_TER.loc[desloc_com_cortes_TER["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")

        desloc_com_cortes_TER.loc[desloc_com_cortes_TER["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes_TER["Deslocamento_pos_corte"]

        desloc_com_cortes_TER.loc[desloc_com_cortes_TER["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

        Deslocamento_TER = desloc_com_cortes_TER[["Colaborador", "Deslocamento"]]


        Produtividadex_TER = AnalisePor(PDOH_TER,"Produtividade")

        Produtividadex_TER = pd.merge(Produtividadex_TER, desloc_com_cortes_TER, how = "right", on = "Colaborador")

        Produtividadex_TER.loc[Produtividadex_TER["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex_TER["Produtividade"] - Produtividadex_TER["corte_pendente_prod"]

        Produtividadex_TER.loc[Produtividadex_TER["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")

        Produtividadex_TER.loc[Produtividadex_TER["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")

        Produtividadex_TER["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Produtividade_TER = Produtividadex_TER[["Colaborador", "Produtividade"]]

        htr_TER = pd.merge(
            Produtividade_TER, Deslocamento_TER, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_TER = pd.merge(
            htr_TER, Ocio_TER, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_TER["Tempo Investido"] = (
            HorasTotaisRegistradas_TER["Produtividade"]
            + HorasTotaisRegistradas_TER["Deslocamento"]
            + HorasTotaisRegistradas_TER["Ocio"]
        )

        Horastotaisxxxx_TER = pd.merge(
            HorasTotaisRegistradas_TER, colaboradores_TER, how="right", on="Colaborador"
        )

        Horastotaisxxxx_TER["justificativas_ter"].fillna(" ", inplace=True)

        Horastotaisxxxx_TER.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Horastotaisxxxx_TER["HT sem almoco"] = Horastotaisxxxx_TER["Tempo Investido"]

        JornadaDeTrabalho_TER = Horastotaisxxxx_TER[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "HT sem almoco",
                "Jornada de Trabalho",
                "almoco",
                "justificativas_ter",
            ]
        ]

        JornadaDeTrabalho_TER["Hora Excedente"] = (
            JornadaDeTrabalho_TER["HT sem almoco"]
            - JornadaDeTrabalho_TER["Jornada de Trabalho"]
        )

        JornadaDeTrabalho_TER["Hora Pendente"] = (
            JornadaDeTrabalho_TER["Jornada de Trabalho"]
            - JornadaDeTrabalho_TER["HT sem almoco"]
        )

        JornadaDeTrabalho_TER.loc[
            JornadaDeTrabalho_TER["Hora Excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Hora Excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        JornadaDeTrabalho_TER.loc[
            JornadaDeTrabalho_TER["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
            "Hora Pendente",
        ] = pd.to_timedelta("0 days 00:00:00")

        Horasnaoregistradas_TER = JornadaDeTrabalho_TER[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Hora Pendente",
                "almoco",
                "justificativas_ter",
            ]
        ]

        corteszx_TER = pd.merge(
            JornadaDeTrabalho_TER, Deslocamento_TER, how="left", on="Colaborador"
        )

        corteszx_TER["Deslocamento_cortado"] = (
            corteszx_TER["Deslocamento"] - corteszx_TER["Hora Excedente"]
        )

        corteszx_TER["Deslocamento_cortado_excedente"] = (
            corteszx_TER["Hora Excedente"] - corteszx_TER["Deslocamento"]
        )

        corteszx_TER.loc[
            corteszx_TER["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        corteszx_TER.loc[
            corteszx_TER["Deslocamento_cortado_excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        cortesyy_TER = pd.merge(corteszx_TER, Ocio_TER, how="outer", on="Colaborador")

        cortesyy_TER["Ocio_cortado"] = (
            cortesyy_TER["Ocio"] - cortesyy_TER["Deslocamento_cortado_excedente"]
        )

        cortesyy_TER["Ocio_cortado_excedente"] = (
            cortesyy_TER["Deslocamento_cortado_excedente"] - cortesyy_TER["Ocio"]
        )

        cortesyy_TER.loc[
            cortesyy_TER["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        cortesyy_TER.loc[
            cortesyy_TER["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_TER = pd.merge(
            cortesyy_TER, Produtividade_TER, how="right", on="Colaborador"
        )

        corteszz_TER["Produtividade_cortada"] = (
            corteszz_TER["Produtividade"] - corteszz_TER["Ocio_cortado_excedente"]
        )

        corteszz_TER.loc[
            corteszz_TER["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
            "Produtividade_cortada",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_TER["TOTAL DE HR"] = (
            corteszz_TER["Produtividade_cortada"]
            + corteszz_TER["Deslocamento_cortado"]
            + corteszz_TER["Ocio_cortado"]
        )

        cortes_TER = corteszz_TER[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Produtividade_cortada",
                "Deslocamento_cortado",
                "Ocio_cortado",
                "TOTAL DE HR",
                "Jornada de Trabalho",
                "Hora Pendente",
                "Hora Excedente",
                "almoco",
                "justificativas_ter",
            ]
        ]

        cortes_TER = cortes_TER.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
        cortes_TER = cortes_TER.rename({"almoco": "almoco"}, axis=1)
        cortes_TER = cortes_TER.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

        PDOH_P_TER = cortes_TER.rename(
            {
                "Produtividade_cortada": "PRODUTIVIDADE",
                "Deslocamento_cortado": "DESLOCAMENTO",
                "Ocio_cortado": "ÓCIO",
                "TOTAL DE HR": "TOTAL",
                "Hora Pendente": "H N REGISTRADAS",
                "Jornada de Trabalho": "horas_programadas",
            },
            axis=1,
        )

        PDOH_P_TER = pd.merge(PDOH_P_TER, checkin_TER, on=["Colaborador"], how="left")

        ########
        visitasgerencial_TER = gerencial[
            [
                "Colaborador",
                "Data da Visita",
                "Tipo de Check-in",
                "Tempo em Loja Executado",
            ]
        ]

        visitasgerencial_TER["Data da Visita"] = pd.to_datetime(
            visitasgerencial_TER["Data da Visita"], format="%d/%m/%Y"
        )

        dias_ptbr = {
            "Sunday": "Domingo",
            "Monday": "Segunda-feira",
            "Tuesday": "Terça-feira",
            "Wednesday": "Quarta-feira",
            "Thursday": "Quinta-feira",
            "Friday": "Sexta-feira",
            "Saturday": "Sábado",
        }

        visitasgerencial_TER["nome_do_dia"] = (
            visitasgerencial_TER["Data da Visita"].dt.day_name().replace(dias_ptbr)
        )

        visitasgerencial_TER = visitasgerencial_TER.loc[
            visitasgerencial_TER["nome_do_dia"] == "Terça-feira"
        ]

        ########

        visitasgerencial_TER["visitas_diarias"] = "1"
        visitasgerencial_TER["percentual_visitas"] = "1"
        visitasgerencial_TER["MANUAL E GPS"] = "0"
        visitasgerencial_TER["MANUAL"] = "0"
        visitasgerencial_TER["SEM CHECKIN"] = "0"

        visitasgerencial_TER.loc[
            visitasgerencial_TER["Tipo de Check-in"] == "Sem Checkin",
            "percentual_visitas",
        ] = "0"
        visitasgerencial_TER.loc[
            visitasgerencial_TER["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
        ] = "1"
        visitasgerencial_TER.loc[
            visitasgerencial_TER["Tipo de Check-in"] == "Manual", "MANUAL"
        ] = "1"
        visitasgerencial_TER.loc[
            visitasgerencial_TER["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
        ] = "1"

        # transformando em int() os valores das colunas que atribuimos.

        visitasgerencial_TER["visitas_diarias"] = visitasgerencial_TER[
            "visitas_diarias"
        ].astype(int)
        visitasgerencial_TER["percentual_visitas"] = visitasgerencial_TER[
            "percentual_visitas"
        ].astype(int)
        visitasgerencial_TER["MANUAL E GPS"] = visitasgerencial_TER[
            "MANUAL E GPS"
        ].astype(int)
        visitasgerencial_TER["MANUAL"] = visitasgerencial_TER["MANUAL"].astype(int)
        visitasgerencial_TER["SEM CHECKIN"] = visitasgerencial_TER[
            "SEM CHECKIN"
        ].astype(int)

        porcentagem_visitas_TER = visitasgerencial_TER[
            ["Colaborador", "visitas_diarias", "percentual_visitas"]
        ]

        porcentagem_visitas_TER = porcentagem_visitas_TER.groupby(["Colaborador"]).agg(
            {"visitas_diarias": "sum", "percentual_visitas": "sum"}
        )

        porcentagem_visitas_TER["% de Visitas R"] = (
            porcentagem_visitas_TER["percentual_visitas"]
            / porcentagem_visitas_TER["visitas_diarias"]
        )

        porcentagem_visitas_TER = porcentagem_visitas_TER.rename(
            {
                "visitas_diarias": "visitas_diarias",
                "percentual_visitas": "visitas_diarias_realizadas",
                "% de Visitas R": "percentual_visitas",
            },
            axis=1,
        )

        TESTEKX_TER = pd.merge(
            PDOH_P_TER, porcentagem_visitas_TER, on=["Colaborador"], how="left"
        )

        md_tempoemloja_TER = visitasgerencial_TER[
            ["Colaborador", "Tempo em Loja Executado"]
        ]
        md_tempoemloja_TER["Tempo em Loja Executado"] = pd.to_timedelta(
            md_tempoemloja_TER["Tempo em Loja Executado"]
        )

        md_tempoemloja_TER = md_tempoemloja_TER.groupby(
            ["Colaborador"], as_index=False
        ).agg({"Tempo em Loja Executado": "mean"})

        md_tempoemloja_TER = md_tempoemloja_TER.rename(
            {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
        )

        TESTEKX_TER = pd.merge(
            TESTEKX_TER, md_tempoemloja_TER, on=["Colaborador"], how="left"
        )

        TESTEKX_TER["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_TER["DESLOCAMENTO"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_TER["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        # TESTEKX_TER["PROD_SOPRANO"] = TESTEKX_TER["PRODUTIVIDADE"] + TESTEKX_TER["ÓCIO"]

        # TESTEKX_TER["PRODUTIVIDADE"] = TESTEKX_TER["PROD_SOPRANO"]

        TESTEKX_TER["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_TER.loc[
            TESTEKX_TER["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
            "% produtividade",
        ] = (
            TESTEKX_TER["PRODUTIVIDADE"] / TESTEKX_TER["horas_programadas"]
        )

        TESTEKX_TER.loc[
            (TESTEKX_TER["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
            & (
                TESTEKX_TER["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))
            ),
            "% produtividade",
        ] = int(0)

        ########
        pesquisasporcentagem_TER = pesquisasporcentagem_x[
            ["Responsável", "Status", "Data de Expiração"]
        ]

        pesquisasporcentagem_TER["Data de Expiração"] = pd.to_datetime(
            pesquisasporcentagem_TER["Data de Expiração"]
        )

        pesquisasporcentagem_TER["nome_do_dia"] = (
            pesquisasporcentagem_TER["Data de Expiração"]
            .dt.day_name()
            .replace(dias_ptbr)
        )

        pesquisasporcentagem_TER = pesquisasporcentagem_TER.loc[
            pesquisasporcentagem_TER["nome_do_dia"] == "Terça-feira"
        ]

        ########

        pesquisasporcentagem_TER["pesquisas_semanais"] = "1"
        pesquisasporcentagem_TER["% de Pesquisasx"] = "1"

        pesquisasporcentagem_TER.loc[
            pesquisasporcentagem_TER["Status"] != "Respondida", "% de Pesquisasx"
        ] = "0"

        # transformando em int() os valores das colunas que atribuimos.

        pesquisasporcentagem_TER["pesquisas_semanais"] = pesquisasporcentagem_TER[
            "pesquisas_semanais"
        ].astype(int)
        pesquisasporcentagem_TER["% de Pesquisasx"] = pesquisasporcentagem_TER[
            "% de Pesquisasx"
        ].astype(int)

        pesquisasporcentagem_TER = pesquisasporcentagem_TER.rename(
            {"Responsável": "Colaborador"}, axis=1
        )

        pesquisasporcentagem_TER = pesquisasporcentagem_TER[
            ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
        ]

        pesquisasporcentagem_TER = pesquisasporcentagem_TER.groupby(
            ["Colaborador"]
        ).agg({"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"})

        pesquisasporcentagem_TER["percentual_pesquisas"] = (
            pesquisasporcentagem_TER["% de Pesquisasx"]
            / pesquisasporcentagem_TER["pesquisas_semanais"]
        )

        pesquisasporcentagem_TER = pesquisasporcentagem_TER[
            ["pesquisas_semanais", "percentual_pesquisas"]
        ]

        TESTEKX_TER = pd.merge(
            TESTEKX_TER, pesquisasporcentagem_TER, on=["Colaborador"], how="left"
        )

        TESTEKX_TER["pesquisas_diarias_realizadas"] = TESTEKX_TER[
            "pesquisas_semanais"
        ].astype(float) * TESTEKX_TER["percentual_pesquisas"].astype(float)

        TESTEKX_TER["pesquisas_semanais"].fillna(int(0), inplace=True)
        TESTEKX_TER["percentual_pesquisas"].fillna(int(0), inplace=True)
        TESTEKX_TER["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

        TESTEKX_TER["visitas_diarias"].fillna(int(0), inplace=True)
        TESTEKX_TER["percentual_visitas"].fillna(int(0), inplace=True)

        PDTT_TER = TESTEKX_TER["PRODUTIVIDADE"].sum()

        DSLCT_TER = TESTEKX_TER["DESLOCAMENTO"].sum()

        OCT_TER = TESTEKX_TER["ÓCIO"].sum()

        TT_TER = TESTEKX_TER["TOTAL"].sum()

        HPTT_TER = TESTEKX_TER["horas_programadas"].sum()

        HNRT_TER = TESTEKX_TER["H N REGISTRADAS"].sum()

        # Condição para evitar divisão por zero
        if HPTT_TER != pd.to_timedelta("0 days 00:00:00"):
            PDESLOC_TER = DSLCT_TER / HPTT_TER
        else:
            PDESLOC_TER = pd.to_timedelta("0 days 00:00:00")

        if HPTT_TER != pd.to_timedelta("0 days 00:00:00"):
            POCIO_TER = OCT_TER / HPTT_TER
        else:
            POCIO_TER = pd.to_timedelta("0 days 00:00:00")

        if HPTT_TER != pd.to_timedelta("0 days 00:00:00"):
            PPROD_TER = PDTT_TER / HPTT_TER
        else:
            PPROD_TER = pd.to_timedelta("0 days 00:00:00")

        if HPTT_TER != pd.to_timedelta("0 days 00:00:00"):
            PHNR_TER = HNRT_TER / HPTT_TER
        else:
            PHNR_TER = pd.to_timedelta("0 days 00:00:00")

        if HPTT_TER != pd.to_timedelta("0 days 00:00:00"):
            PHT_TER = HPTT_TER / HPTT_TER
        else:
            PHT_TER = pd.to_timedelta("0 days 00:00:00")

        TESTEKX_TER["pesquisas_diarias_realizadas"] = (
            TESTEKX_TER["pesquisas_semanais"] * TESTEKX_TER["percentual_pesquisas"]
        )
        TESTEKX_TER.replace([float("inf"), -float("inf")], 0, inplace=True)
        TESTEKX_TER["pesquisas_diarias_realizadas"] = TESTEKX_TER[
            "pesquisas_diarias_realizadas"
        ].fillna(int(0))
        TESTEKX_TER["pesquisas_diarias_realizadas"] = TESTEKX_TER[
            "pesquisas_diarias_realizadas"
        ].astype(int)

        SVS_TER = TESTEKX_TER["visitas_diarias"].sum()
        SVSR_TER = TESTEKX_TER["visitas_diarias_realizadas"].sum()
        SPS_TER = TESTEKX_TER["pesquisas_semanais"].sum()
        SPSR_TER = TESTEKX_TER["pesquisas_diarias_realizadas"].sum()

        TESTEKX_TER.loc[
            TESTEKX_TER["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias",
        ] = int(0)
        TESTEKX_TER.loc[
            TESTEKX_TER["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias_realizadas",
        ] = int(0)
        TESTEKX_TER.loc[
            TESTEKX_TER["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_diarias_realizadas",
        ] = int(0)
        TESTEKX_TER.loc[
            TESTEKX_TER["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_semanais",
        ] = int(0)

        # Condição para evitar divisão por zero
        if HPTT_TER != pd.to_timedelta("0 days 00:00:00"):
            MdPROD_TER = PDTT_TER / HPTT_TER
        else:
            MdPROD_TER = int(0)

        if SVS_TER != pd.to_timedelta("0 days 00:00:00"):
            MdVST_TER = SVSR_TER / SVS_TER
        else:
            MdVST_TER = int(0)

        if SPS_TER != pd.to_timedelta("0 days 00:00:00"):
            MdPSQ_TER = SPSR_TER / SPS_TER
        else:
            MdPSQ_TER = int(0)

        TESTEKX_TER = TESTEKX_TER.rename(
            {"pesquisas_semanais": "pesquisas_diarias"}, axis=1
        )

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha_TER = pd.DataFrame(
            {
                "Colaborador": ["TOTAL"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PDTT_TER],
                "DESLOCAMENTO": [DSLCT_TER],
                "ÓCIO": [OCT_TER],
                "TOTAL": [TT_TER],
                "horas_programadas": [HPTT_TER],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "H N REGISTRADAS": [HNRT_TER],
                "visitas_diarias": [SVS_TER],
                "visitas_diarias_realizadas": [SVSR_TER],
                "pesquisas_diarias": [SPS_TER],
                "pesquisas_diarias_realizadas": [SPSR_TER],
                "% produtividade": [MdPROD_TER],
                "percentual_visitas": [MdVST_TER],
                "percentual_pesquisas": [MdPSQ_TER],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        TESTEKX_TER = pd.concat([TESTEKX_TER, nova_linha_TER], ignore_index=True)

        TESTEKX_TER.loc[TESTEKX_TER["visitas_diarias"] == 0, "percentual_visitas"] = "-"
        TESTEKX_TER.loc[
            TESTEKX_TER["pesquisas_diarias"] == 0, "percentual_pesquisas"
        ] = "-"

        # Convertendo as colunas para tipos numéricos
        TESTEKX_TER["% produtividade"] = pd.to_numeric(
            TESTEKX_TER["% produtividade"], errors="coerce"
        )
        TESTEKX_TER["percentual_visitas"] = pd.to_numeric(
            TESTEKX_TER["percentual_visitas"], errors="coerce"
        )
        TESTEKX_TER["percentual_pesquisas"] = pd.to_numeric(
            TESTEKX_TER["percentual_pesquisas"], errors="coerce"
        )

        # Verificando se há valores não numéricos que foram convertidos para NaN
        if (
            TESTEKX_TER[
                ["% produtividade", "percentual_visitas", "percentual_pesquisas"]
            ]
            .isnull()
            .values.any()
        ):
            print(
                "Existem valores não numéricos nas colunas que foram convertidos para NaN."
            )
            print(TESTEKX_TER)

        # Commenting out the actual calculation since the DataFrame is not defined here.
        TESTEKX_TER["percentual_efetividade"] = (
            TESTEKX_TER["% produtividade"] * pesos["Produtividade"]
            + TESTEKX_TER["percentual_visitas"] * pesos["Visitas"]
            + TESTEKX_TER["percentual_pesquisas"] * pesos["Pesquisas"]
        ) / sum(pesos.values())

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha2_TER = pd.DataFrame(
            {
                "Colaborador": ["-"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PPROD_TER],
                "DESLOCAMENTO": [PDESLOC_TER],
                "ÓCIO": [POCIO_TER],
                "H N REGISTRADAS": [PHNR_TER],
                "horas_programadas": [PHT_TER],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "visitas_diarias": ["-"],
                "visitas_diarias_realizadas": ["-"],
                "pesquisas_diarias": ["-"],
                "pesquisas_diarias_realizadas": ["-"],
                "% produtividade": ["-"],
                "percentual_visitas": ["-"],
                "percentual_pesquisas": ["-"],
                "percentual_efetividade": ["-"],
                "justificativas_ter": ["-"],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        PDOHCOMPLETO_TER = pd.concat([TESTEKX_TER, nova_linha2_TER], ignore_index=True)

        PDOHCOMPLETO_TER["nome_do_dia"] = "Terça-feira"
        PDOHCOMPLETO_TER["data"] = teste_x_TER

        PDOHCOMPLETO_TER = PDOHCOMPLETO_TER.rename(
            {
                "Colaborador": "colaborador",
                "SUPERIOR": "superior",
                "UF": "estado",
                "DESLOCAMENTO": "deslocamento",
                "ÓCIO": "ocio",
                "PRODUTIVIDADE": "produtividade",
                "H N REGISTRADAS": "horas_nao_registradas",
                "horas_programadas": "horas_programadas",
                "Média de tempo em loja": "media_tempo_em_loja",
                "% produtividade": "percentual_produtividade",
                "percentual_visitas": "percentual_visitas",
                "percentual_pesquisas": "percentual_pesquisas",
                "percentual_efetividade": "percentual_efetividade",
            },
            axis=1,
        )

        PDOHCOMPLETO_TER.loc[
            PDOHCOMPLETO_TER["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_produtividade",
        ] = "-"
        PDOHCOMPLETO_TER.loc[
            PDOHCOMPLETO_TER["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_visitas",
        ] = "-"
        PDOHCOMPLETO_TER.loc[
            PDOHCOMPLETO_TER["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_pesquisas",
        ] = "-"
        PDOHCOMPLETO_TER.loc[
            PDOHCOMPLETO_TER["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_efetividade",
        ] = "-"

        PDOHCOMPLETO_TER_XXX = PDOHCOMPLETO_TER[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_ter",
            ]
        ]

        PDOHCOMPLETO_TER = PDOHCOMPLETO_TER[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_ter",
            ]
        ]

        concat_analise.append(PDOHCOMPLETO_TER_XXX)

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------TERÇA-FEIRA------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------QUARTA-FEIRA-----------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    if True in dias_da_semana["nome_do_dia"].isin(["Quarta-feira"]).tolist():
        PDOH_QUA = PDOH.loc[PDOH["nome_do_dia"] == "Quarta-feira"]

        def AnalisePDOH(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador", "data"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida  #####################################################

        def AnalisePor(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida

        Ocio_QUA = AnalisePDOH(PDOH_QUA, "Ocio")
        print(Ocio_QUA)

        teste_x_QUA = Ocio_QUA.iloc[0, 1]


        #Criando variável para alocar o corte do almoco dentro do Ocio
        ocio_sem_almoco_QUA = Ocio_QUA

        ocio_sem_almoco_QUA = pd.merge(ocio_sem_almoco_QUA, colaboradores_QUA, on="Colaborador", how="right")

        ocio_sem_almoco_QUA = ocio_sem_almoco_QUA[["Colaborador", "Ocio", "almoco"]]

        ocio_sem_almoco_QUA = ocio_sem_almoco_QUA.rename({"almoco":"almoco"}, axis=1)

        ocio_sem_almoco_QUA["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        ocio_sem_almoco_QUA["Ocio_sem_almoco"] = ocio_sem_almoco_QUA["Ocio"] - ocio_sem_almoco_QUA["almoco"]

        ocio_sem_almoco_QUA.loc[ocio_sem_almoco_QUA["Ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco_QUA["almoco"] - ocio_sem_almoco_QUA["Ocio"]

        ocio_sem_almoco_QUA.loc[ocio_sem_almoco_QUA["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_QUA.loc[ocio_sem_almoco_QUA["Ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco_QUA["Ocio_sem_almoco"]
        ocio_sem_almoco_QUA.loc[ocio_sem_almoco_QUA["Ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_QUA["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Ocio_QUA = ocio_sem_almoco_QUA[["Colaborador", "Ocio_real"]]

        Ocio_QUA.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)


        # --------------------------------------- DESLOCAMENTO
        Deslocamento_QUA = AnalisePor(PDOH_QUA,"Deslocamento")

        desloc_com_cortes_QUA = pd.merge(Deslocamento_QUA, ocio_sem_almoco_QUA, on="Colaborador", how="right")

        desloc_com_cortes_QUA["Deslocamento_pos_corte"] = desloc_com_cortes_QUA["Deslocamento"] - desloc_com_cortes_QUA["Corte_pendente"]

        desloc_com_cortes_QUA["corte_pendente_prod"] = desloc_com_cortes_QUA["Corte_pendente"] - desloc_com_cortes_QUA["Deslocamento"]

        desloc_com_cortes_QUA.loc[desloc_com_cortes_QUA["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")

        desloc_com_cortes_QUA.loc[desloc_com_cortes_QUA["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes_QUA["Deslocamento_pos_corte"]

        desloc_com_cortes_QUA.loc[desloc_com_cortes_QUA["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

        Deslocamento_QUA = desloc_com_cortes_QUA[["Colaborador", "Deslocamento"]]


        Produtividadex_QUA = AnalisePor(PDOH_QUA,"Produtividade")

        Produtividadex_QUA = pd.merge(Produtividadex_QUA, desloc_com_cortes_QUA, how = "right", on = "Colaborador")

        Produtividadex_QUA.loc[Produtividadex_QUA["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex_QUA["Produtividade"] - Produtividadex_QUA["corte_pendente_prod"]

        Produtividadex_QUA.loc[Produtividadex_QUA["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")

        Produtividadex_QUA.loc[Produtividadex_QUA["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")

        Produtividadex_QUA["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Produtividade_QUA = Produtividadex_QUA[["Colaborador", "Produtividade"]]


        htr_QUA = pd.merge(
            Produtividade_QUA, Deslocamento_QUA, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_QUA = pd.merge(
            htr_QUA, Ocio_QUA, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_QUA["Tempo Investido"] = (
            HorasTotaisRegistradas_QUA["Produtividade"]
            + HorasTotaisRegistradas_QUA["Deslocamento"]
            + HorasTotaisRegistradas_QUA["Ocio"]
        )

        Horastotaisxxxx_QUA = pd.merge(
            HorasTotaisRegistradas_QUA, colaboradores_QUA, how="right", on="Colaborador"
        )

        Horastotaisxxxx_QUA["justificativas_qua"].fillna(" ", inplace=True)

        Horastotaisxxxx_QUA.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Horastotaisxxxx_QUA["HT sem almoco"] = Horastotaisxxxx_QUA["Tempo Investido"]

        JornadaDeTrabalho_QUA = Horastotaisxxxx_QUA[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "HT sem almoco",
                "Jornada de Trabalho",
                "almoco",
                "justificativas_qua",
            ]
        ]

        JornadaDeTrabalho_QUA["Hora Excedente"] = (
            JornadaDeTrabalho_QUA["HT sem almoco"]
            - JornadaDeTrabalho_QUA["Jornada de Trabalho"]
        )

        JornadaDeTrabalho_QUA["Hora Pendente"] = (
            JornadaDeTrabalho_QUA["Jornada de Trabalho"]
            - JornadaDeTrabalho_QUA["HT sem almoco"]
        )

        JornadaDeTrabalho_QUA.loc[
            JornadaDeTrabalho_QUA["Hora Excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Hora Excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        JornadaDeTrabalho_QUA.loc[
            JornadaDeTrabalho_QUA["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
            "Hora Pendente",
        ] = pd.to_timedelta("0 days 00:00:00")

        Horasnaoregistradas_QUA = JornadaDeTrabalho_QUA[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Hora Pendente",
                "almoco",
                "justificativas_qua",
            ]
        ]

        corteszx_QUA = pd.merge(
            JornadaDeTrabalho_QUA, Deslocamento_QUA, how="left", on="Colaborador"
        )

        corteszx_QUA["Deslocamento_cortado"] = (
            corteszx_QUA["Deslocamento"] - corteszx_QUA["Hora Excedente"]
        )

        corteszx_QUA["Deslocamento_cortado_excedente"] = (
            corteszx_QUA["Hora Excedente"] - corteszx_QUA["Deslocamento"]
        )

        corteszx_QUA.loc[
            corteszx_QUA["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        corteszx_QUA.loc[
            corteszx_QUA["Deslocamento_cortado_excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        cortesyy_QUA = pd.merge(corteszx_QUA, Ocio_QUA, how="outer", on="Colaborador")

        cortesyy_QUA["Ocio_cortado"] = (
            cortesyy_QUA["Ocio"] - cortesyy_QUA["Deslocamento_cortado_excedente"]
        )

        cortesyy_QUA["Ocio_cortado_excedente"] = (
            cortesyy_QUA["Deslocamento_cortado_excedente"] - cortesyy_QUA["Ocio"]
        )

        cortesyy_QUA.loc[
            cortesyy_QUA["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        cortesyy_QUA.loc[
            cortesyy_QUA["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_QUA = pd.merge(
            cortesyy_QUA, Produtividade_QUA, how="right", on="Colaborador"
        )

        corteszz_QUA["Produtividade_cortada"] = (
            corteszz_QUA["Produtividade"] - corteszz_QUA["Ocio_cortado_excedente"]
        )

        corteszz_QUA.loc[
            corteszz_QUA["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
            "Produtividade_cortada",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_QUA["TOTAL DE HR"] = (
            corteszz_QUA["Produtividade_cortada"]
            + corteszz_QUA["Deslocamento_cortado"]
            + corteszz_QUA["Ocio_cortado"]
        )

        cortes_QUA = corteszz_QUA[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Produtividade_cortada",
                "Deslocamento_cortado",
                "Ocio_cortado",
                "TOTAL DE HR",
                "Jornada de Trabalho",
                "Hora Pendente",
                "Hora Excedente",
                "almoco",
                "justificativas_qua",
            ]
        ]

        cortes_QUA = cortes_QUA.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
        cortes_QUA = cortes_QUA.rename({"almoco": "almoco"}, axis=1)
        cortes_QUA = cortes_QUA.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

        PDOH_P_QUA = cortes_QUA.rename(
            {
                "Produtividade_cortada": "PRODUTIVIDADE",
                "Deslocamento_cortado": "DESLOCAMENTO",
                "Ocio_cortado": "ÓCIO",
                "TOTAL DE HR": "TOTAL",
                "Hora Pendente": "H N REGISTRADAS",
                "Jornada de Trabalho": "horas_programadas",
            },
            axis=1,
        )

        PDOH_P_QUA = pd.merge(PDOH_P_QUA, checkin_QUA, on=["Colaborador"], how="left")

        ########
        visitasgerencial_QUA = gerencial[
            [
                "Colaborador",
                "Data da Visita",
                "Tipo de Check-in",
                "Tempo em Loja Executado",
            ]
        ]

        visitasgerencial_QUA["Data da Visita"] = pd.to_datetime(
            visitasgerencial_QUA["Data da Visita"], format="%d/%m/%Y"
        )

        dias_ptbr = {
            "Sunday": "Domingo",
            "Monday": "Segunda-feira",
            "Tuesday": "Terça-feira",
            "Wednesday": "Quarta-feira",
            "Thursday": "Quinta-feira",
            "Friday": "Sexta-feira",
            "Saturday": "Sábado",
        }

        visitasgerencial_QUA["nome_do_dia"] = (
            visitasgerencial_QUA["Data da Visita"].dt.day_name().replace(dias_ptbr)
        )

        visitasgerencial_QUA = visitasgerencial_QUA.loc[
            visitasgerencial_QUA["nome_do_dia"] == "Quarta-feira"
        ]

        ########

        visitasgerencial_QUA["visitas_diarias"] = "1"
        visitasgerencial_QUA["percentual_visitas"] = "1"
        visitasgerencial_QUA["MANUAL E GPS"] = "0"
        visitasgerencial_QUA["MANUAL"] = "0"
        visitasgerencial_QUA["SEM CHECKIN"] = "0"

        visitasgerencial_QUA.loc[
            visitasgerencial_QUA["Tipo de Check-in"] == "Sem Checkin",
            "percentual_visitas",
        ] = "0"
        visitasgerencial_QUA.loc[
            visitasgerencial_QUA["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
        ] = "1"
        visitasgerencial_QUA.loc[
            visitasgerencial_QUA["Tipo de Check-in"] == "Manual", "MANUAL"
        ] = "1"
        visitasgerencial_QUA.loc[
            visitasgerencial_QUA["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
        ] = "1"

        # transformando em int() os valores das colunas que atribuimos.

        visitasgerencial_QUA["visitas_diarias"] = visitasgerencial_QUA[
            "visitas_diarias"
        ].astype(int)
        visitasgerencial_QUA["percentual_visitas"] = visitasgerencial_QUA[
            "percentual_visitas"
        ].astype(int)
        visitasgerencial_QUA["MANUAL E GPS"] = visitasgerencial_QUA[
            "MANUAL E GPS"
        ].astype(int)
        visitasgerencial_QUA["MANUAL"] = visitasgerencial_QUA["MANUAL"].astype(int)
        visitasgerencial_QUA["SEM CHECKIN"] = visitasgerencial_QUA[
            "SEM CHECKIN"
        ].astype(int)

        porcentagem_visitas_QUA = visitasgerencial_QUA[
            ["Colaborador", "visitas_diarias", "percentual_visitas"]
        ]

        porcentagem_visitas_QUA = porcentagem_visitas_QUA.groupby(["Colaborador"]).agg(
            {"visitas_diarias": "sum", "percentual_visitas": "sum"}
        )

        porcentagem_visitas_QUA["% de Visitas R"] = (
            porcentagem_visitas_QUA["percentual_visitas"]
            / porcentagem_visitas_QUA["visitas_diarias"]
        )

        porcentagem_visitas_QUA = porcentagem_visitas_QUA.rename(
            {
                "visitas_diarias": "visitas_diarias",
                "percentual_visitas": "visitas_diarias_realizadas",
                "% de Visitas R": "percentual_visitas",
            },
            axis=1,
        )

        TESTEKX_QUA = pd.merge(
            PDOH_P_QUA, porcentagem_visitas_QUA, on=["Colaborador"], how="left"
        )

        md_tempoemloja_QUA = visitasgerencial_QUA[
            ["Colaborador", "Tempo em Loja Executado"]
        ]
        md_tempoemloja_QUA["Tempo em Loja Executado"] = pd.to_timedelta(
            md_tempoemloja_QUA["Tempo em Loja Executado"]
        )

        md_tempoemloja_QUA = md_tempoemloja_QUA.groupby(
            ["Colaborador"], as_index=False
        ).agg({"Tempo em Loja Executado": "mean"})

        md_tempoemloja_QUA = md_tempoemloja_QUA.rename(
            {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
        )

        TESTEKX_QUA = pd.merge(
            TESTEKX_QUA, md_tempoemloja_QUA, on=["Colaborador"], how="left"
        )

        TESTEKX_QUA["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_QUA["DESLOCAMENTO"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_QUA["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        # TESTEKX_QUA["PROD_SOPRANO"] = TESTEKX_QUA["PRODUTIVIDADE"] + TESTEKX_QUA["ÓCIO"]

        # TESTEKX_QUA["PRODUTIVIDADE"] = TESTEKX_QUA["PROD_SOPRANO"]

        TESTEKX_QUA["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_QUA.loc[
            TESTEKX_QUA["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
            "% produtividade",
        ] = (
            TESTEKX_QUA["PRODUTIVIDADE"] / TESTEKX_QUA["horas_programadas"]
        )

        TESTEKX_QUA.loc[
            (TESTEKX_QUA["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
            & (
                TESTEKX_QUA["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))
            ),
            "% produtividade",
        ] = int(0)

        ########
        pesquisasporcentagem_QUA = pesquisasporcentagem_x[
            ["Responsável", "Status", "Data de Expiração"]
        ]

        pesquisasporcentagem_QUA["Data de Expiração"] = pd.to_datetime(
            pesquisasporcentagem_QUA["Data de Expiração"]
        )

        pesquisasporcentagem_QUA["nome_do_dia"] = (
            pesquisasporcentagem_QUA["Data de Expiração"]
            .dt.day_name()
            .replace(dias_ptbr)
        )

        pesquisasporcentagem_QUA = pesquisasporcentagem_QUA.loc[
            pesquisasporcentagem_QUA["nome_do_dia"] == "Quarta-feira"
        ]

        ########

        pesquisasporcentagem_QUA["pesquisas_semanais"] = "1"
        pesquisasporcentagem_QUA["% de Pesquisasx"] = "1"

        pesquisasporcentagem_QUA.loc[
            pesquisasporcentagem_QUA["Status"] != "Respondida", "% de Pesquisasx"
        ] = "0"

        # transformando em int() os valores das colunas que atribuimos.

        pesquisasporcentagem_QUA["pesquisas_semanais"] = pesquisasporcentagem_QUA[
            "pesquisas_semanais"
        ].astype(int)
        pesquisasporcentagem_QUA["% de Pesquisasx"] = pesquisasporcentagem_QUA[
            "% de Pesquisasx"
        ].astype(int)

        pesquisasporcentagem_QUA = pesquisasporcentagem_QUA.rename(
            {"Responsável": "Colaborador"}, axis=1
        )

        pesquisasporcentagem_QUA = pesquisasporcentagem_QUA[
            ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
        ]

        pesquisasporcentagem_QUA = pesquisasporcentagem_QUA.groupby(
            ["Colaborador"]
        ).agg({"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"})

        pesquisasporcentagem_QUA["percentual_pesquisas"] = (
            pesquisasporcentagem_QUA["% de Pesquisasx"]
            / pesquisasporcentagem_QUA["pesquisas_semanais"]
        )

        pesquisasporcentagem_QUA = pesquisasporcentagem_QUA[
            ["pesquisas_semanais", "percentual_pesquisas"]
        ]

        TESTEKX_QUA = pd.merge(
            TESTEKX_QUA, pesquisasporcentagem_QUA, on=["Colaborador"], how="left"
        )

        TESTEKX_QUA["pesquisas_diarias_realizadas"] = TESTEKX_QUA[
            "pesquisas_semanais"
        ].astype(float) * TESTEKX_QUA["percentual_pesquisas"].astype(float)

        TESTEKX_QUA["pesquisas_semanais"].fillna(int(0), inplace=True)
        TESTEKX_QUA["percentual_pesquisas"].fillna(int(0), inplace=True)
        TESTEKX_QUA["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

        TESTEKX_QUA["visitas_diarias"].fillna(int(0), inplace=True)
        TESTEKX_QUA["percentual_visitas"].fillna(int(0), inplace=True)

        PDTT_QUA = TESTEKX_QUA["PRODUTIVIDADE"].sum()

        DSLCT_QUA = TESTEKX_QUA["DESLOCAMENTO"].sum()

        OCT_QUA = TESTEKX_QUA["ÓCIO"].sum()

        TT_QUA = TESTEKX_QUA["TOTAL"].sum()

        HPTT_QUA = TESTEKX_QUA["horas_programadas"].sum()

        HNRT_QUA = TESTEKX_QUA["H N REGISTRADAS"].sum()

        # Condição para evitar divisão por zero
        if HPTT_QUA != pd.to_timedelta("0 days 00:00:00"):
            PDESLOC_QUA = DSLCT_QUA / HPTT_QUA
        else:
            PDESLOC_QUA = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUA != pd.to_timedelta("0 days 00:00:00"):
            POCIO_QUA = OCT_QUA / HPTT_QUA
        else:
            POCIO_QUA = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUA != pd.to_timedelta("0 days 00:00:00"):
            PPROD_QUA = PDTT_QUA / HPTT_QUA
        else:
            PPROD_QUA = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUA != pd.to_timedelta("0 days 00:00:00"):
            PHNR_QUA = HNRT_QUA / HPTT_QUA
        else:
            PHNR_QUA = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUA != pd.to_timedelta("0 days 00:00:00"):
            PHT_QUA = HPTT_QUA / HPTT_QUA
        else:
            PHT_QUA = pd.to_timedelta("0 days 00:00:00")

        TESTEKX_QUA["pesquisas_diarias_realizadas"] = (
            TESTEKX_QUA["pesquisas_semanais"] * TESTEKX_QUA["percentual_pesquisas"]
        )
        TESTEKX_QUA.replace([float("inf"), -float("inf")], 0, inplace=True)
        TESTEKX_QUA["pesquisas_diarias_realizadas"] = TESTEKX_QUA[
            "pesquisas_diarias_realizadas"
        ].fillna(int(0))
        TESTEKX_QUA["pesquisas_diarias_realizadas"] = TESTEKX_QUA[
            "pesquisas_diarias_realizadas"
        ].astype(int)

        TESTEKX_QUA.loc[
            TESTEKX_QUA["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias",
        ] = int(0)
        TESTEKX_QUA.loc[
            TESTEKX_QUA["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias_realizadas",
        ] = int(0)
        TESTEKX_QUA.loc[
            TESTEKX_QUA["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_diarias_realizadas",
        ] = int(0)
        TESTEKX_QUA.loc[
            TESTEKX_QUA["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_semanais",
        ] = int(0)

        SVS_QUA = TESTEKX_QUA["visitas_diarias"].sum()
        SVSR_QUA = TESTEKX_QUA["visitas_diarias_realizadas"].sum()
        SPS_QUA = TESTEKX_QUA["pesquisas_semanais"].sum()
        SPSR_QUA = TESTEKX_QUA["pesquisas_diarias_realizadas"].sum()

        # Condição para evitar divisão por zero
        if HPTT_QUA != pd.to_timedelta("0 days 00:00:00"):
            MdPROD_QUA = PDTT_QUA / HPTT_QUA
        else:
            MdPROD_QUA = int(0)

        if SVS_QUA != pd.to_timedelta("0 days 00:00:00"):
            MdVST_QUA = SVSR_QUA / SVS_QUA
        else:
            MdVST_QUA = int(0)

        if SPS_QUA != pd.to_timedelta("0 days 00:00:00"):
            MdPSQ_QUA = SPSR_QUA / SPS_QUA
        else:
            MdPSQ_QUA = int(0)

        TESTEKX_QUA = TESTEKX_QUA.rename(
            {"pesquisas_semanais": "pesquisas_diarias"}, axis=1
        )

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha_QUA = pd.DataFrame(
            {
                "Colaborador": ["TOTAL"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PDTT_QUA],
                "DESLOCAMENTO": [DSLCT_QUA],
                "ÓCIO": [OCT_QUA],
                "TOTAL": [TT_QUA],
                "horas_programadas": [HPTT_QUA],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "H N REGISTRADAS": [HNRT_QUA],
                "visitas_diarias": [SVS_QUA],
                "visitas_diarias_realizadas": [SVSR_QUA],
                "pesquisas_diarias": [SPS_QUA],
                "pesquisas_diarias_realizadas": [SPSR_QUA],
                "% produtividade": [MdPROD_QUA],
                "percentual_visitas": [MdVST_QUA],
                "percentual_pesquisas": [MdPSQ_QUA],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        TESTEKX_QUA = pd.concat([TESTEKX_QUA, nova_linha_QUA], ignore_index=True)

        TESTEKX_QUA.loc[TESTEKX_QUA["visitas_diarias"] == 0, "percentual_visitas"] = "-"
        TESTEKX_QUA.loc[
            TESTEKX_QUA["pesquisas_diarias"] == 0, "percentual_pesquisas"
        ] = "-"

        # Convertendo as colunas para tipos numéricos
        TESTEKX_QUA["% produtividade"] = pd.to_numeric(
            TESTEKX_QUA["% produtividade"], errors="coerce"
        )
        TESTEKX_QUA["percentual_visitas"] = pd.to_numeric(
            TESTEKX_QUA["percentual_visitas"], errors="coerce"
        )
        TESTEKX_QUA["percentual_pesquisas"] = pd.to_numeric(
            TESTEKX_QUA["percentual_pesquisas"], errors="coerce"
        )

        # Verificando se há valores não numéricos que foram convertidos para NaN
        if (
            TESTEKX_QUA[
                ["% produtividade", "percentual_visitas", "percentual_pesquisas"]
            ]
            .isnull()
            .values.any()
        ):
            print(
                "Existem valores não numéricos nas colunas que foram convertidos para NaN."
            )
            print(TESTEKX_QUA)

        # Commenting out the actual calculation since the DataFrame is not defined here.
        TESTEKX_QUA["percentual_efetividade"] = (
            TESTEKX_QUA["% produtividade"] * pesos["Produtividade"]
            + TESTEKX_QUA["percentual_visitas"] * pesos["Visitas"]
            + TESTEKX_QUA["percentual_pesquisas"] * pesos["Pesquisas"]
        ) / sum(pesos.values())

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha2_QUA = pd.DataFrame(
            {
                "Colaborador": ["-"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PPROD_QUA],
                "DESLOCAMENTO": [PDESLOC_QUA],
                "ÓCIO": [POCIO_QUA],
                "H N REGISTRADAS": [PHNR_QUA],
                "horas_programadas": [PHT_QUA],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "visitas_diarias": ["-"],
                "visitas_diarias_realizadas": ["-"],
                "pesquisas_diarias": ["-"],
                "pesquisas_diarias_realizadas": ["-"],
                "% produtividade": ["-"],
                "percentual_visitas": ["-"],
                "percentual_pesquisas": ["-"],
                "percentual_efetividade": ["-"],
                "justificativas_qua": ["-"],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        PDOHCOMPLETO_QUA = pd.concat([TESTEKX_QUA, nova_linha2_QUA], ignore_index=True)

        PDOHCOMPLETO_QUA["nome_do_dia"] = "Quarta-feira"
        PDOHCOMPLETO_QUA["data"] = teste_x_QUA

        PDOHCOMPLETO_QUA = PDOHCOMPLETO_QUA.rename(
            {
                "Colaborador": "colaborador",
                "SUPERIOR": "superior",
                "UF": "estado",
                "DESLOCAMENTO": "deslocamento",
                "ÓCIO": "ocio",
                "PRODUTIVIDADE": "produtividade",
                "H N REGISTRADAS": "horas_nao_registradas",
                "horas_programadas": "horas_programadas",
                "Média de tempo em loja": "media_tempo_em_loja",
                "% produtividade": "percentual_produtividade",
                "percentual_visitas": "percentual_visitas",
                "percentual_pesquisas": "percentual_pesquisas",
                "percentual_efetividade": "percentual_efetividade",
            },
            axis=1,
        )

        PDOHCOMPLETO_QUA.loc[
            PDOHCOMPLETO_QUA["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_produtividade",
        ] = "-"
        PDOHCOMPLETO_QUA.loc[
            PDOHCOMPLETO_QUA["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_visitas",
        ] = "-"
        PDOHCOMPLETO_QUA.loc[
            PDOHCOMPLETO_QUA["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_pesquisas",
        ] = "-"
        PDOHCOMPLETO_QUA.loc[
            PDOHCOMPLETO_QUA["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_efetividade",
        ] = "-"

        PDOHCOMPLETO_QUA_XXX = PDOHCOMPLETO_QUA[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_qua",
            ]
        ]

        PDOHCOMPLETO_QUA = PDOHCOMPLETO_QUA[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_qua",
            ]
        ]

        concat_analise.append(PDOHCOMPLETO_QUA_XXX)

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------QUARTA-FEIRA-----------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------QUINTA-FEIRA-----------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    if True in dias_da_semana["nome_do_dia"].isin(["Quinta-feira"]).tolist():
        PDOH_QUI = PDOH.loc[PDOH["nome_do_dia"] == "Quinta-feira"]

        def AnalisePDOH(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador", "data"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida  #####################################################

        def AnalisePor(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida

        Ocio_QUI = AnalisePDOH(PDOH_QUI, "Ocio")
        print(Ocio_QUI)

        teste_x_QUI = Ocio_QUI.iloc[0, 1]


        #Criando variável para alocar o corte do almoco dentro do Ocio
        ocio_sem_almoco_QUI = Ocio_QUI

        ocio_sem_almoco_QUI = pd.merge(ocio_sem_almoco_QUI, colaboradores_QUI, on="Colaborador", how="right")

        ocio_sem_almoco_QUI = ocio_sem_almoco_QUI[["Colaborador", "Ocio", "almoco"]]

        ocio_sem_almoco_QUI = ocio_sem_almoco_QUI.rename({"almoco":"almoco"}, axis=1)

        ocio_sem_almoco_QUI["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        ocio_sem_almoco_QUI["Ocio_sem_almoco"] = ocio_sem_almoco_QUI["Ocio"] - ocio_sem_almoco_QUI["almoco"]

        ocio_sem_almoco_QUI.loc[ocio_sem_almoco_QUI["Ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco_QUI["almoco"] - ocio_sem_almoco_QUI["Ocio"]

        ocio_sem_almoco_QUI.loc[ocio_sem_almoco_QUI["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_QUI.loc[ocio_sem_almoco_QUI["Ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco_QUI["Ocio_sem_almoco"]
        ocio_sem_almoco_QUI.loc[ocio_sem_almoco_QUI["Ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_QUI["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Ocio_QUI = ocio_sem_almoco_QUI[["Colaborador", "Ocio_real"]]

        Ocio_QUI.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)


        # --------------------------------------- DESLOCAMENTO
        Deslocamento_QUI = AnalisePor(PDOH_QUI,"Deslocamento")

        desloc_com_cortes_QUI = pd.merge(Deslocamento_QUI, ocio_sem_almoco_QUI, on="Colaborador", how="right")

        desloc_com_cortes_QUI["Deslocamento_pos_corte"] = desloc_com_cortes_QUI["Deslocamento"] - desloc_com_cortes_QUI["Corte_pendente"]

        desloc_com_cortes_QUI["corte_pendente_prod"] = desloc_com_cortes_QUI["Corte_pendente"] - desloc_com_cortes_QUI["Deslocamento"]

        desloc_com_cortes_QUI.loc[desloc_com_cortes_QUI["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")

        desloc_com_cortes_QUI.loc[desloc_com_cortes_QUI["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes_QUI["Deslocamento_pos_corte"]

        desloc_com_cortes_QUI.loc[desloc_com_cortes_QUI["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

        Deslocamento_QUI = desloc_com_cortes_QUI[["Colaborador", "Deslocamento"]]


        Produtividadex_QUI = AnalisePor(PDOH_QUI,"Produtividade")

        Produtividadex_QUI = pd.merge(Produtividadex_QUI, desloc_com_cortes_QUI, how = "right", on = "Colaborador")

        Produtividadex_QUI.loc[Produtividadex_QUI["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex_QUI["Produtividade"] - Produtividadex_QUI["corte_pendente_prod"]

        Produtividadex_QUI.loc[Produtividadex_QUI["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")

        Produtividadex_QUI.loc[Produtividadex_QUI["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")

        Produtividadex_QUI["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Produtividade_QUI = Produtividadex_QUI[["Colaborador", "Produtividade"]]

        htr_QUI = pd.merge(
            Produtividade_QUI, Deslocamento_QUI, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_QUI = pd.merge(
            htr_QUI, Ocio_QUI, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_QUI["Tempo Investido"] = (
            HorasTotaisRegistradas_QUI["Produtividade"]
            + HorasTotaisRegistradas_QUI["Deslocamento"]
            + HorasTotaisRegistradas_QUI["Ocio"]
        )

        Horastotaisxxxx_QUI = pd.merge(
            HorasTotaisRegistradas_QUI, colaboradores_QUI, how="right", on="Colaborador"
        )

        Horastotaisxxxx_QUI["justificativas_qui"].fillna(" ", inplace=True)

        Horastotaisxxxx_QUI.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Horastotaisxxxx_QUI["HT sem almoco"] = Horastotaisxxxx_QUI["Tempo Investido"]

        JornadaDeTrabalho_QUI = Horastotaisxxxx_QUI[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "HT sem almoco",
                "Jornada de Trabalho",
                "almoco",
                "justificativas_qui",
            ]
        ]

        JornadaDeTrabalho_QUI["Hora Excedente"] = (
            JornadaDeTrabalho_QUI["HT sem almoco"]
            - JornadaDeTrabalho_QUI["Jornada de Trabalho"]
        )

        JornadaDeTrabalho_QUI["Hora Pendente"] = (
            JornadaDeTrabalho_QUI["Jornada de Trabalho"]
            - JornadaDeTrabalho_QUI["HT sem almoco"]
        )

        JornadaDeTrabalho_QUI.loc[
            JornadaDeTrabalho_QUI["Hora Excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Hora Excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        JornadaDeTrabalho_QUI.loc[
            JornadaDeTrabalho_QUI["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
            "Hora Pendente",
        ] = pd.to_timedelta("0 days 00:00:00")

        Horasnaoregistradas_QUI = JornadaDeTrabalho_QUI[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Hora Pendente",
                "almoco",
                "justificativas_qui",
            ]
        ]

        corteszx_QUI = pd.merge(
            JornadaDeTrabalho_QUI, Deslocamento_QUI, how="left", on="Colaborador"
        )

        corteszx_QUI["Deslocamento_cortado"] = (
            corteszx_QUI["Deslocamento"] - corteszx_QUI["Hora Excedente"]
        )

        corteszx_QUI["Deslocamento_cortado_excedente"] = (
            corteszx_QUI["Hora Excedente"] - corteszx_QUI["Deslocamento"]
        )

        corteszx_QUI.loc[
            corteszx_QUI["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        corteszx_QUI.loc[
            corteszx_QUI["Deslocamento_cortado_excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        cortesyy_QUI = pd.merge(corteszx_QUI, Ocio_QUI, how="outer", on="Colaborador")

        cortesyy_QUI["Ocio_cortado"] = (
            cortesyy_QUI["Ocio"] - cortesyy_QUI["Deslocamento_cortado_excedente"]
        )

        cortesyy_QUI["Ocio_cortado_excedente"] = (
            cortesyy_QUI["Deslocamento_cortado_excedente"] - cortesyy_QUI["Ocio"]
        )

        cortesyy_QUI.loc[
            cortesyy_QUI["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        cortesyy_QUI.loc[
            cortesyy_QUI["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_QUI = pd.merge(
            cortesyy_QUI, Produtividade_QUI, how="right", on="Colaborador"
        )

        corteszz_QUI["Produtividade_cortada"] = (
            corteszz_QUI["Produtividade"] - corteszz_QUI["Ocio_cortado_excedente"]
        )

        corteszz_QUI.loc[
            corteszz_QUI["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
            "Produtividade_cortada",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_QUI["TOTAL DE HR"] = (
            corteszz_QUI["Produtividade_cortada"]
            + corteszz_QUI["Deslocamento_cortado"]
            + corteszz_QUI["Ocio_cortado"]
        )

        cortes_QUI = corteszz_QUI[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Produtividade_cortada",
                "Deslocamento_cortado",
                "Ocio_cortado",
                "TOTAL DE HR",
                "Jornada de Trabalho",
                "Hora Pendente",
                "Hora Excedente",
                "almoco",
                "justificativas_qui",
            ]
        ]

        cortes_QUI = cortes_QUI.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
        cortes_QUI = cortes_QUI.rename({"almoco": "almoco"}, axis=1)
        cortes_QUI = cortes_QUI.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

        PDOH_P_QUI = cortes_QUI.rename(
            {
                "Produtividade_cortada": "PRODUTIVIDADE",
                "Deslocamento_cortado": "DESLOCAMENTO",
                "Ocio_cortado": "ÓCIO",
                "TOTAL DE HR": "TOTAL",
                "Hora Pendente": "H N REGISTRADAS",
                "Jornada de Trabalho": "horas_programadas",
            },
            axis=1,
        )

        PDOH_P_QUI = pd.merge(PDOH_P_QUI, checkin_QUI, on=["Colaborador"], how="left")

        ########
        visitasgerencial_QUI = gerencial[
            [
                "Colaborador",
                "Data da Visita",
                "Tipo de Check-in",
                "Tempo em Loja Executado",
            ]
        ]

        visitasgerencial_QUI["Data da Visita"] = pd.to_datetime(
            visitasgerencial_QUI["Data da Visita"], format="%d/%m/%Y"
        )

        dias_ptbr = {
            "Sunday": "Domingo",
            "Monday": "Segunda-feira",
            "Tuesday": "Terça-feira",
            "Wednesday": "Quarta-feira",
            "Thursday": "Quinta-feira",
            "Friday": "Sexta-feira",
            "Saturday": "Sábado",
        }

        visitasgerencial_QUI["nome_do_dia"] = (
            visitasgerencial_QUI["Data da Visita"].dt.day_name().replace(dias_ptbr)
        )

        visitasgerencial_QUI = visitasgerencial_QUI.loc[
            visitasgerencial_QUI["nome_do_dia"] == "Quinta-feira"
        ]

        ########

        visitasgerencial_QUI["visitas_diarias"] = "1"
        visitasgerencial_QUI["percentual_visitas"] = "1"
        visitasgerencial_QUI["MANUAL E GPS"] = "0"
        visitasgerencial_QUI["MANUAL"] = "0"
        visitasgerencial_QUI["SEM CHECKIN"] = "0"

        visitasgerencial_QUI.loc[
            visitasgerencial_QUI["Tipo de Check-in"] == "Sem Checkin",
            "percentual_visitas",
        ] = "0"
        visitasgerencial_QUI.loc[
            visitasgerencial_QUI["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
        ] = "1"
        visitasgerencial_QUI.loc[
            visitasgerencial_QUI["Tipo de Check-in"] == "Manual", "MANUAL"
        ] = "1"
        visitasgerencial_QUI.loc[
            visitasgerencial_QUI["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
        ] = "1"

        # transformando em int() os valores das colunas que atribuimos.

        visitasgerencial_QUI["visitas_diarias"] = visitasgerencial_QUI[
            "visitas_diarias"
        ].astype(int)
        visitasgerencial_QUI["percentual_visitas"] = visitasgerencial_QUI[
            "percentual_visitas"
        ].astype(int)
        visitasgerencial_QUI["MANUAL E GPS"] = visitasgerencial_QUI[
            "MANUAL E GPS"
        ].astype(int)
        visitasgerencial_QUI["MANUAL"] = visitasgerencial_QUI["MANUAL"].astype(int)
        visitasgerencial_QUI["SEM CHECKIN"] = visitasgerencial_QUI[
            "SEM CHECKIN"
        ].astype(int)

        porcentagem_visitas_QUI = visitasgerencial_QUI[
            ["Colaborador", "visitas_diarias", "percentual_visitas"]
        ]

        porcentagem_visitas_QUI = porcentagem_visitas_QUI.groupby(["Colaborador"]).agg(
            {"visitas_diarias": "sum", "percentual_visitas": "sum"}
        )

        porcentagem_visitas_QUI["% de Visitas R"] = (
            porcentagem_visitas_QUI["percentual_visitas"]
            / porcentagem_visitas_QUI["visitas_diarias"]
        )

        porcentagem_visitas_QUI = porcentagem_visitas_QUI.rename(
            {
                "visitas_diarias": "visitas_diarias",
                "percentual_visitas": "visitas_diarias_realizadas",
                "% de Visitas R": "percentual_visitas",
            },
            axis=1,
        )

        TESTEKX_QUI = pd.merge(
            PDOH_P_QUI, porcentagem_visitas_QUI, on=["Colaborador"], how="left"
        )

        md_tempoemloja_QUI = visitasgerencial_QUI[
            ["Colaborador", "Tempo em Loja Executado"]
        ]
        md_tempoemloja_QUI["Tempo em Loja Executado"] = pd.to_timedelta(
            md_tempoemloja_QUI["Tempo em Loja Executado"]
        )

        md_tempoemloja_QUI = md_tempoemloja_QUI.groupby(
            ["Colaborador"], as_index=False
        ).agg({"Tempo em Loja Executado": "mean"})

        md_tempoemloja_QUI = md_tempoemloja_QUI.rename(
            {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
        )

        TESTEKX_QUI = pd.merge(
            TESTEKX_QUI, md_tempoemloja_QUI, on=["Colaborador"], how="left"
        )

        ########
        pesquisasporcentagem_QUI = pesquisasporcentagem_x[
            ["Responsável", "Status", "Data de Expiração"]
        ]

        pesquisasporcentagem_QUI["Data de Expiração"] = pd.to_datetime(
            pesquisasporcentagem_QUI["Data de Expiração"]
        )

        pesquisasporcentagem_QUI["nome_do_dia"] = (
            pesquisasporcentagem_QUI["Data de Expiração"]
            .dt.day_name()
            .replace(dias_ptbr)
        )

        pesquisasporcentagem_QUI = pesquisasporcentagem_QUI.loc[
            pesquisasporcentagem_QUI["nome_do_dia"] == "Quinta-feira"
        ]

        ########

        pesquisasporcentagem_QUI["pesquisas_semanais"] = "1"
        pesquisasporcentagem_QUI["% de Pesquisasx"] = "1"

        pesquisasporcentagem_QUI.loc[
            pesquisasporcentagem_QUI["Status"] != "Respondida", "% de Pesquisasx"
        ] = "0"

        # transformando em int() os valores das colunas que atribuimos.

        pesquisasporcentagem_QUI["pesquisas_semanais"] = pesquisasporcentagem_QUI[
            "pesquisas_semanais"
        ].astype(int)
        pesquisasporcentagem_QUI["% de Pesquisasx"] = pesquisasporcentagem_QUI[
            "% de Pesquisasx"
        ].astype(int)

        pesquisasporcentagem_QUI = pesquisasporcentagem_QUI.rename(
            {"Responsável": "Colaborador"}, axis=1
        )

        pesquisasporcentagem_QUI = pesquisasporcentagem_QUI[
            ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
        ]

        pesquisasporcentagem_QUI = pesquisasporcentagem_QUI.groupby(
            ["Colaborador"]
        ).agg({"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"})

        pesquisasporcentagem_QUI["percentual_pesquisas"] = (
            pesquisasporcentagem_QUI["% de Pesquisasx"]
            / pesquisasporcentagem_QUI["pesquisas_semanais"]
        )

        pesquisasporcentagem_QUI = pesquisasporcentagem_QUI[
            ["pesquisas_semanais", "percentual_pesquisas"]
        ]

        TESTEKX_QUI = pd.merge(
            TESTEKX_QUI, pesquisasporcentagem_QUI, on=["Colaborador"], how="left"
        )

        TESTEKX_QUI["pesquisas_diarias_realizadas"] = TESTEKX_QUI[
            "pesquisas_semanais"
        ].astype(float) * TESTEKX_QUI["percentual_pesquisas"].astype(float)

        TESTEKX_QUI["pesquisas_semanais"].fillna(int(0), inplace=True)
        TESTEKX_QUI["percentual_pesquisas"].fillna(int(0), inplace=True)
        TESTEKX_QUI["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

        TESTEKX_QUI["visitas_diarias"].fillna(int(0), inplace=True)
        TESTEKX_QUI["percentual_visitas"].fillna(int(0), inplace=True)

        TESTEKX_QUI["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_QUI["DESLOCAMENTO"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_QUI["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        # TESTEKX_QUI["PROD_SOPRANO"] = TESTEKX_QUI["PRODUTIVIDADE"] + TESTEKX_QUI["ÓCIO"]

        # TESTEKX_QUI["PRODUTIVIDADE"] = TESTEKX_QUI["PROD_SOPRANO"]

        TESTEKX_QUI["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_QUI.loc[
            TESTEKX_QUI["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
            "% produtividade",
        ] = (
            TESTEKX_QUI["PRODUTIVIDADE"] / TESTEKX_QUI["horas_programadas"]
        )

        TESTEKX_QUI.loc[
            (TESTEKX_QUI["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
            & (
                TESTEKX_QUI["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))
            ),
            "% produtividade",
        ] = int(0)

        PDTT_QUI = TESTEKX_QUI["PRODUTIVIDADE"].sum()

        DSLCT_QUI = TESTEKX_QUI["DESLOCAMENTO"].sum()

        OCT_QUI = TESTEKX_QUI["ÓCIO"].sum()

        TT_QUI = TESTEKX_QUI["TOTAL"].sum()

        HPTT_QUI = TESTEKX_QUI["horas_programadas"].sum()

        HNRT_QUI = TESTEKX_QUI["H N REGISTRADAS"].sum()

        # Condição para evitar divisão por zero
        if HPTT_QUI != pd.to_timedelta("0 days 00:00:00"):
            PDESLOC_QUI = DSLCT_QUI / HPTT_QUI
        else:
            PDESLOC_QUI = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUI != pd.to_timedelta("0 days 00:00:00"):
            POCIO_QUI = OCT_QUI / HPTT_QUI
        else:
            POCIO_QUI = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUI != pd.to_timedelta("0 days 00:00:00"):
            PPROD_QUI = PDTT_QUI / HPTT_QUI
        else:
            PPROD_QUI = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUI != pd.to_timedelta("0 days 00:00:00"):
            PHNR_QUI = HNRT_QUI / HPTT_QUI
        else:
            PHNR_QUI = pd.to_timedelta("0 days 00:00:00")

        if HPTT_QUI != pd.to_timedelta("0 days 00:00:00"):
            PHT_QUI = HPTT_QUI / HPTT_QUI
        else:
            PHT_QUI = pd.to_timedelta("0 days 00:00:00")

        TESTEKX_QUI["pesquisas_diarias_realizadas"] = (
            TESTEKX_QUI["pesquisas_semanais"] * TESTEKX_QUI["percentual_pesquisas"]
        )
        TESTEKX_QUI.replace([float("inf"), -float("inf")], 0, inplace=True)
        TESTEKX_QUI["pesquisas_diarias_realizadas"] = TESTEKX_QUI[
            "pesquisas_diarias_realizadas"
        ].fillna(int(0))
        TESTEKX_QUI["pesquisas_diarias_realizadas"] = TESTEKX_QUI[
            "pesquisas_diarias_realizadas"
        ].astype(int)

        TESTEKX_QUI.loc[
            TESTEKX_QUI["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias",
        ] = int(0)
        TESTEKX_QUI.loc[
            TESTEKX_QUI["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias_realizadas",
        ] = int(0)
        TESTEKX_QUI.loc[
            TESTEKX_QUI["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_diarias_realizadas",
        ] = int(0)
        TESTEKX_QUI.loc[
            TESTEKX_QUI["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_semanais",
        ] = int(0)

        SVS_QUI = TESTEKX_QUI["visitas_diarias"].sum()
        SVSR_QUI = TESTEKX_QUI["visitas_diarias_realizadas"].sum()
        SPS_QUI = TESTEKX_QUI["pesquisas_semanais"].sum()
        SPSR_QUI = TESTEKX_QUI["pesquisas_diarias_realizadas"].sum()

        # Condição para evitar divisão por zero
        if HPTT_QUI != pd.to_timedelta("0 days 00:00:00"):
            MdPROD_QUI = PDTT_QUI / HPTT_QUI
        else:
            MdPROD_QUI = int(0)

        if SVS_QUI != pd.to_timedelta("0 days 00:00:00"):
            MdVST_QUI = SVSR_QUI / SVS_QUI
        else:
            MdVST_QUI = int(0)

        if SPS_QUI != pd.to_timedelta("0 days 00:00:00"):
            MdPSQ_QUI = SPSR_QUI / SPS_QUI
        else:
            MdPSQ_QUI = int(0)

        TESTEKX_QUI = TESTEKX_QUI.rename(
            {"pesquisas_semanais": "pesquisas_diarias"}, axis=1
        )

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha_QUI = pd.DataFrame(
            {
                "Colaborador": ["TOTAL"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PDTT_QUI],
                "DESLOCAMENTO": [DSLCT_QUI],
                "ÓCIO": [OCT_QUI],
                "TOTAL": [TT_QUI],
                "horas_programadas": [HPTT_QUI],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "H N REGISTRADAS": [HNRT_QUI],
                "visitas_diarias": [SVS_QUI],
                "visitas_diarias_realizadas": [SVSR_QUI],
                "pesquisas_diarias": [SPS_QUI],
                "pesquisas_diarias_realizadas": [SPSR_QUI],
                "% produtividade": [MdPROD_QUI],
                "percentual_visitas": [MdVST_QUI],
                "percentual_pesquisas": [MdPSQ_QUI],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        TESTEKX_QUI = pd.concat([TESTEKX_QUI, nova_linha_QUI], ignore_index=True)

        TESTEKX_QUI.loc[TESTEKX_QUI["visitas_diarias"] == 0, "percentual_visitas"] = "-"
        TESTEKX_QUI.loc[
            TESTEKX_QUI["pesquisas_diarias"] == 0, "percentual_pesquisas"
        ] = "-"

        # Convertendo as colunas para tipos numéricos
        TESTEKX_QUI["% produtividade"] = pd.to_numeric(
            TESTEKX_QUI["% produtividade"], errors="coerce"
        )
        TESTEKX_QUI["percentual_visitas"] = pd.to_numeric(
            TESTEKX_QUI["percentual_visitas"], errors="coerce"
        )
        TESTEKX_QUI["percentual_pesquisas"] = pd.to_numeric(
            TESTEKX_QUI["percentual_pesquisas"], errors="coerce"
        )

        # Verificando se há valores não numéricos que foram convertidos para NaN
        if (
            TESTEKX_QUI[
                ["% produtividade", "percentual_visitas", "percentual_pesquisas"]
            ]
            .isnull()
            .values.any()
        ):
            print(
                "Existem valores não numéricos nas colunas que foram convertidos para NaN."
            )
            print(TESTEKX_QUI)

        # Commenting out the actual calculation since the DataFrame is not defined here.
        TESTEKX_QUI["percentual_efetividade"] = (
            TESTEKX_QUI["% produtividade"] * pesos["Produtividade"]
            + TESTEKX_QUI["percentual_visitas"] * pesos["Visitas"]
            + TESTEKX_QUI["percentual_pesquisas"] * pesos["Pesquisas"]
        ) / sum(pesos.values())

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha2_QUI = pd.DataFrame(
            {
                "Colaborador": ["-"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PPROD_QUI],
                "DESLOCAMENTO": [PDESLOC_QUI],
                "ÓCIO": [POCIO_QUI],
                "H N REGISTRADAS": [PHNR_QUI],
                "horas_programadas": [PHT_QUI],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "visitas_diarias": ["-"],
                "visitas_diarias_realizadas": ["-"],
                "pesquisas_diarias": ["-"],
                "pesquisas_diarias_realizadas": ["-"],
                "% produtividade": ["-"],
                "percentual_visitas": ["-"],
                "percentual_pesquisas": ["-"],
                "percentual_efetividade": ["-"],
                "justificativas_qui": ["-"],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        PDOHCOMPLETO_QUI = pd.concat([TESTEKX_QUI, nova_linha2_QUI], ignore_index=True)

        PDOHCOMPLETO_QUI["nome_do_dia"] = "Quinta-feira"
        PDOHCOMPLETO_QUI["data"] = teste_x_QUI

        PDOHCOMPLETO_QUI = PDOHCOMPLETO_QUI.rename(
            {
                "Colaborador": "colaborador",
                "SUPERIOR": "superior",
                "UF": "estado",
                "DESLOCAMENTO": "deslocamento",
                "ÓCIO": "ocio",
                "PRODUTIVIDADE": "produtividade",
                "H N REGISTRADAS": "horas_nao_registradas",
                "horas_programadas": "horas_programadas",
                "Média de tempo em loja": "media_tempo_em_loja",
                "% produtividade": "percentual_produtividade",
                "percentual_visitas": "percentual_visitas",
                "percentual_pesquisas": "percentual_pesquisas",
                "percentual_efetividade": "percentual_efetividade",
            },
            axis=1,
        )

        PDOHCOMPLETO_QUI.loc[
            PDOHCOMPLETO_QUI["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_produtividade",
        ] = "-"
        PDOHCOMPLETO_QUI.loc[
            PDOHCOMPLETO_QUI["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_visitas",
        ] = "-"
        PDOHCOMPLETO_QUI.loc[
            PDOHCOMPLETO_QUI["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_pesquisas",
        ] = "-"
        PDOHCOMPLETO_QUI.loc[
            PDOHCOMPLETO_QUI["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_efetividade",
        ] = "-"

        PDOHCOMPLETO_QUI_XXX = PDOHCOMPLETO_QUI[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_qui",
            ]
        ]

        PDOHCOMPLETO_QUI = PDOHCOMPLETO_QUI[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_qui",
            ]
        ]

        concat_analise.append(PDOHCOMPLETO_QUI_XXX)

    # --------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------QUINTA-FEIRA--------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------------
    # -----------------------------------------------SEXTA-FEIRA----------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    if True in dias_da_semana["nome_do_dia"].isin(["Sexta-feira"]).tolist():
        PDOH_SEX = PDOH.loc[PDOH["nome_do_dia"] == "Sexta-feira"]

        def AnalisePDOH(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador", "data"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida  #####################################################

        def AnalisePor(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida

        Ocio_SEX = AnalisePDOH(PDOH_SEX, "Ocio")
        print(Ocio_SEX)

        teste_x_SEX = Ocio_SEX.iloc[0, 1]


        #Criando variável para alocar o corte do almoco dentro do Ocio
        ocio_sem_almoco_SEX = Ocio_SEX

        ocio_sem_almoco_SEX = pd.merge(ocio_sem_almoco_SEX, colaboradores_SEX, on="Colaborador", how="right")

        ocio_sem_almoco_SEX = ocio_sem_almoco_SEX[["Colaborador", "Ocio", "almoco"]]

        ocio_sem_almoco_SEX = ocio_sem_almoco_SEX.rename({"almoco":"almoco"}, axis=1)

        ocio_sem_almoco_SEX["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        ocio_sem_almoco_SEX["Ocio_sem_almoco"] = ocio_sem_almoco_SEX["Ocio"] - ocio_sem_almoco_SEX["almoco"]

        ocio_sem_almoco_SEX.loc[ocio_sem_almoco_SEX["Ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco_SEX["almoco"] - ocio_sem_almoco_SEX["Ocio"]

        ocio_sem_almoco_SEX.loc[ocio_sem_almoco_SEX["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_SEX.loc[ocio_sem_almoco_SEX["Ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco_SEX["Ocio_sem_almoco"]
        ocio_sem_almoco_SEX.loc[ocio_sem_almoco_SEX["Ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_SEX["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Ocio_SEX = ocio_sem_almoco_SEX[["Colaborador", "Ocio_real"]]

        Ocio_SEX.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)


        # --------------------------------------- DESLOCAMENTO
        Deslocamento_SEX = AnalisePor(PDOH_SEX,"Deslocamento")

        desloc_com_cortes_SEX = pd.merge(Deslocamento_SEX, ocio_sem_almoco_SEX, on="Colaborador", how="right")

        desloc_com_cortes_SEX["Deslocamento_pos_corte"] = desloc_com_cortes_SEX["Deslocamento"] - desloc_com_cortes_SEX["Corte_pendente"]

        desloc_com_cortes_SEX["corte_pendente_prod"] = desloc_com_cortes_SEX["Corte_pendente"] - desloc_com_cortes_SEX["Deslocamento"]

        desloc_com_cortes_SEX.loc[desloc_com_cortes_SEX["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")

        desloc_com_cortes_SEX.loc[desloc_com_cortes_SEX["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes_SEX["Deslocamento_pos_corte"]

        desloc_com_cortes_SEX.loc[desloc_com_cortes_SEX["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

        Deslocamento_SEX = desloc_com_cortes_SEX[["Colaborador", "Deslocamento"]]


        Produtividadex_SEX = AnalisePor(PDOH_SEX,"Produtividade")

        Produtividadex_SEX = pd.merge(Produtividadex_SEX, desloc_com_cortes_SEX, how = "right", on = "Colaborador")

        Produtividadex_SEX.loc[Produtividadex_SEX["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex_SEX["Produtividade"] - Produtividadex_SEX["corte_pendente_prod"]

        Produtividadex_SEX.loc[Produtividadex_SEX["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")

        Produtividadex_SEX.loc[Produtividadex_SEX["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")

        Produtividadex_SEX["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Produtividade_SEX = Produtividadex_SEX[["Colaborador", "Produtividade"]]

        htr_SEX = pd.merge(
            Produtividade_SEX, Deslocamento_SEX, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_SEX = pd.merge(
            htr_SEX, Ocio_SEX, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_SEX["Tempo Investido"] = (
            HorasTotaisRegistradas_SEX["Produtividade"]
            + HorasTotaisRegistradas_SEX["Deslocamento"]
            + HorasTotaisRegistradas_SEX["Ocio"]
        )

        Horastotaisxxxx_SEX = pd.merge(
            HorasTotaisRegistradas_SEX, colaboradores_SEX, how="right", on="Colaborador"
        )

        Horastotaisxxxx_SEX["justificativas_sex"].fillna(" ", inplace=True)

        Horastotaisxxxx_SEX.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Horastotaisxxxx_SEX["HT sem almoco"] = Horastotaisxxxx_SEX["Tempo Investido"]

        JornadaDeTrabalho_SEX = Horastotaisxxxx_SEX[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "HT sem almoco",
                "Jornada de Trabalho",
                "almoco",
                "justificativas_sex",
            ]
        ]

        JornadaDeTrabalho_SEX["Hora Excedente"] = (
            JornadaDeTrabalho_SEX["HT sem almoco"]
            - JornadaDeTrabalho_SEX["Jornada de Trabalho"]
        )

        JornadaDeTrabalho_SEX["Hora Pendente"] = (
            JornadaDeTrabalho_SEX["Jornada de Trabalho"]
            - JornadaDeTrabalho_SEX["HT sem almoco"]
        )

        JornadaDeTrabalho_SEX.loc[
            JornadaDeTrabalho_SEX["Hora Excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Hora Excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        JornadaDeTrabalho_SEX.loc[
            JornadaDeTrabalho_SEX["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
            "Hora Pendente",
        ] = pd.to_timedelta("0 days 00:00:00")

        Horasnaoregistradas_SEX = JornadaDeTrabalho_SEX[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Hora Pendente",
                "almoco",
                "justificativas_sex",
            ]
        ]

        corteszx_SEX = pd.merge(
            JornadaDeTrabalho_SEX, Deslocamento_SEX, how="left", on="Colaborador"
        )

        corteszx_SEX["Deslocamento_cortado"] = (
            corteszx_SEX["Deslocamento"] - corteszx_SEX["Hora Excedente"]
        )

        corteszx_SEX["Deslocamento_cortado_excedente"] = (
            corteszx_SEX["Hora Excedente"] - corteszx_SEX["Deslocamento"]
        )

        corteszx_SEX.loc[
            corteszx_SEX["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        corteszx_SEX.loc[
            corteszx_SEX["Deslocamento_cortado_excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        cortesyy_SEX = pd.merge(corteszx_SEX, Ocio_SEX, how="outer", on="Colaborador")

        cortesyy_SEX["Ocio_cortado"] = (
            cortesyy_SEX["Ocio"] - cortesyy_SEX["Deslocamento_cortado_excedente"]
        )

        cortesyy_SEX["Ocio_cortado_excedente"] = (
            cortesyy_SEX["Deslocamento_cortado_excedente"] - cortesyy_SEX["Ocio"]
        )

        cortesyy_SEX.loc[
            cortesyy_SEX["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        cortesyy_SEX.loc[
            cortesyy_SEX["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_SEX = pd.merge(
            cortesyy_SEX, Produtividade_SEX, how="right", on="Colaborador"
        )

        corteszz_SEX["Produtividade_cortada"] = (
            corteszz_SEX["Produtividade"] - corteszz_SEX["Ocio_cortado_excedente"]
        )

        corteszz_SEX.loc[
            corteszz_SEX["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
            "Produtividade_cortada",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_SEX["TOTAL DE HR"] = (
            corteszz_SEX["Produtividade_cortada"]
            + corteszz_SEX["Deslocamento_cortado"]
            + corteszz_SEX["Ocio_cortado"]
        )

        cortes_SEX = corteszz_SEX[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Produtividade_cortada",
                "Deslocamento_cortado",
                "Ocio_cortado",
                "TOTAL DE HR",
                "Jornada de Trabalho",
                "Hora Pendente",
                "Hora Excedente",
                "almoco",
                "justificativas_sex",
            ]
        ]

        cortes_SEX = cortes_SEX.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
        cortes_SEX = cortes_SEX.rename({"almoco": "almoco"}, axis=1)
        cortes_SEX = cortes_SEX.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

        PDOH_P_SEX = cortes_SEX.rename(
            {
                "Produtividade_cortada": "PRODUTIVIDADE",
                "Deslocamento_cortado": "DESLOCAMENTO",
                "Ocio_cortado": "ÓCIO",
                "TOTAL DE HR": "TOTAL",
                "Hora Pendente": "H N REGISTRADAS",
                "Jornada de Trabalho": "horas_programadas",
            },
            axis=1,
        )

        PDOH_P_SEX = pd.merge(PDOH_P_SEX, checkin_SEX, on=["Colaborador"], how="left")

        ########
        visitasgerencial_SEX = gerencial[
            [
                "Colaborador",
                "Data da Visita",
                "Tipo de Check-in",
                "Tempo em Loja Executado",
            ]
        ]

        visitasgerencial_SEX["Data da Visita"] = pd.to_datetime(
            visitasgerencial_SEX["Data da Visita"], format="%d/%m/%Y"
        )

        dias_ptbr = {
            "Sunday": "Domingo",
            "Monday": "Segunda-feira",
            "Tuesday": "Terça-feira",
            "Wednesday": "Quarta-feira",
            "Thursday": "Quinta-feira",
            "Friday": "Sexta-feira",
            "Saturday": "Sábado",
        }

        visitasgerencial_SEX["nome_do_dia"] = (
            visitasgerencial_SEX["Data da Visita"].dt.day_name().replace(dias_ptbr)
        )

        visitasgerencial_SEX = visitasgerencial_SEX.loc[
            visitasgerencial_SEX["nome_do_dia"] == "Sexta-feira"
        ]

        ########

        visitasgerencial_SEX["visitas_diarias"] = "1"
        visitasgerencial_SEX["percentual_visitas"] = "1"
        visitasgerencial_SEX["MANUAL E GPS"] = "0"
        visitasgerencial_SEX["MANUAL"] = "0"
        visitasgerencial_SEX["SEM CHECKIN"] = "0"

        visitasgerencial_SEX.loc[
            visitasgerencial_SEX["Tipo de Check-in"] == "Sem Checkin",
            "percentual_visitas",
        ] = "0"
        visitasgerencial_SEX.loc[
            visitasgerencial_SEX["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
        ] = "1"
        visitasgerencial_SEX.loc[
            visitasgerencial_SEX["Tipo de Check-in"] == "Manual", "MANUAL"
        ] = "1"
        visitasgerencial_SEX.loc[
            visitasgerencial_SEX["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
        ] = "1"

        # transformando em int() os valores das colunas que atribuimos.

        visitasgerencial_SEX["visitas_diarias"] = visitasgerencial_SEX[
            "visitas_diarias"
        ].astype(int)
        visitasgerencial_SEX["percentual_visitas"] = visitasgerencial_SEX[
            "percentual_visitas"
        ].astype(int)
        visitasgerencial_SEX["MANUAL E GPS"] = visitasgerencial_SEX[
            "MANUAL E GPS"
        ].astype(int)
        visitasgerencial_SEX["MANUAL"] = visitasgerencial_SEX["MANUAL"].astype(int)
        visitasgerencial_SEX["SEM CHECKIN"] = visitasgerencial_SEX[
            "SEM CHECKIN"
        ].astype(int)

        porcentagem_visitas_SEX = visitasgerencial_SEX[
            ["Colaborador", "visitas_diarias", "percentual_visitas"]
        ]

        porcentagem_visitas_SEX = porcentagem_visitas_SEX.groupby(["Colaborador"]).agg(
            {"visitas_diarias": "sum", "percentual_visitas": "sum"}
        )

        porcentagem_visitas_SEX["% de Visitas R"] = (
            porcentagem_visitas_SEX["percentual_visitas"]
            / porcentagem_visitas_SEX["visitas_diarias"]
        )

        porcentagem_visitas_SEX = porcentagem_visitas_SEX.rename(
            {
                "visitas_diarias": "visitas_diarias",
                "percentual_visitas": "visitas_diarias_realizadas",
                "% de Visitas R": "percentual_visitas",
            },
            axis=1,
        )

        TESTEKX_SEX = pd.merge(
            PDOH_P_SEX, porcentagem_visitas_SEX, on=["Colaborador"], how="left"
        )

        md_tempoemloja_SEX = visitasgerencial_SEX[
            ["Colaborador", "Tempo em Loja Executado"]
        ]
        md_tempoemloja_SEX["Tempo em Loja Executado"] = pd.to_timedelta(
            md_tempoemloja_SEX["Tempo em Loja Executado"]
        )

        md_tempoemloja_SEX = md_tempoemloja_SEX.groupby(
            ["Colaborador"], as_index=False
        ).agg({"Tempo em Loja Executado": "mean"})

        md_tempoemloja_SEX = md_tempoemloja_SEX.rename(
            {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
        )

        TESTEKX_SEX = pd.merge(
            TESTEKX_SEX, md_tempoemloja_SEX, on=["Colaborador"], how="left"
        )

        ########
        pesquisasporcentagem_SEX = pesquisasporcentagem_x[
            ["Responsável", "Status", "Data de Expiração"]
        ]

        pesquisasporcentagem_SEX["Data de Expiração"] = pd.to_datetime(
            pesquisasporcentagem_SEX["Data de Expiração"]
        )

        pesquisasporcentagem_SEX["nome_do_dia"] = (
            pesquisasporcentagem_SEX["Data de Expiração"]
            .dt.day_name()
            .replace(dias_ptbr)
        )

        pesquisasporcentagem_SEX = pesquisasporcentagem_SEX.loc[
            pesquisasporcentagem_SEX["nome_do_dia"] == "Sexta-feira"
        ]

        ########

        pesquisasporcentagem_SEX["pesquisas_semanais"] = "1"
        pesquisasporcentagem_SEX["% de Pesquisasx"] = "1"

        pesquisasporcentagem_SEX.loc[
            pesquisasporcentagem_SEX["Status"] != "Respondida", "% de Pesquisasx"
        ] = "0"

        # transformando em int() os valores das colunas que atribuimos.

        pesquisasporcentagem_SEX["pesquisas_semanais"] = pesquisasporcentagem_SEX[
            "pesquisas_semanais"
        ].astype(int)
        pesquisasporcentagem_SEX["% de Pesquisasx"] = pesquisasporcentagem_SEX[
            "% de Pesquisasx"
        ].astype(int)

        pesquisasporcentagem_SEX = pesquisasporcentagem_SEX.rename(
            {"Responsável": "Colaborador"}, axis=1
        )

        pesquisasporcentagem_SEX = pesquisasporcentagem_SEX[
            ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
        ]

        pesquisasporcentagem_SEX = pesquisasporcentagem_SEX.groupby(
            ["Colaborador"]
        ).agg({"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"})

        pesquisasporcentagem_SEX["percentual_pesquisas"] = (
            pesquisasporcentagem_SEX["% de Pesquisasx"]
            / pesquisasporcentagem_SEX["pesquisas_semanais"]
        )

        pesquisasporcentagem_SEX = pesquisasporcentagem_SEX[
            ["pesquisas_semanais", "percentual_pesquisas"]
        ]

        TESTEKX_SEX = pd.merge(
            TESTEKX_SEX, pesquisasporcentagem_SEX, on=["Colaborador"], how="left"
        )

        TESTEKX_SEX["pesquisas_diarias_realizadas"] = TESTEKX_SEX[
            "pesquisas_semanais"
        ].astype(float) * TESTEKX_SEX["percentual_pesquisas"].astype(float)

        TESTEKX_SEX["pesquisas_semanais"].fillna(int(0), inplace=True)
        TESTEKX_SEX["percentual_pesquisas"].fillna(int(0), inplace=True)
        TESTEKX_SEX["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

        TESTEKX_SEX["visitas_diarias"].fillna(int(0), inplace=True)
        TESTEKX_SEX["percentual_visitas"].fillna(int(0), inplace=True)

        TESTEKX_SEX["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SEX["DESLOCAMENTO"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SEX["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        # TESTEKX_SEX["PROD_SOPRANO"] = TESTEKX_SEX["PRODUTIVIDADE"] + TESTEKX_SEX["ÓCIO"]

        # TESTEKX_SEX["PRODUTIVIDADE"] = TESTEKX_SEX["PROD_SOPRANO"]

        TESTEKX_SEX["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SEX.loc[
            TESTEKX_SEX["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
            "% produtividade",
        ] = (
            TESTEKX_SEX["PRODUTIVIDADE"] / TESTEKX_SEX["horas_programadas"]
        )

        TESTEKX_SEX.loc[
            (TESTEKX_SEX["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
            & (
                TESTEKX_SEX["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))
            ),
            "% produtividade",
        ] = int(0)

        PDTT_SEX = TESTEKX_SEX["PRODUTIVIDADE"].sum()

        DSLCT_SEX = TESTEKX_SEX["DESLOCAMENTO"].sum()

        OCT_SEX = TESTEKX_SEX["ÓCIO"].sum()

        TT_SEX = TESTEKX_SEX["TOTAL"].sum()

        HPTT_SEX = TESTEKX_SEX["horas_programadas"].sum()

        HNRT_SEX = TESTEKX_SEX["H N REGISTRADAS"].sum()

        # Condição para evitar divisão por zero
        if HPTT_SEX != pd.to_timedelta("0 days 00:00:00"):
            PDESLOC_SEX = DSLCT_SEX / HPTT_SEX
        else:
            PDESLOC_SEX = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEX != pd.to_timedelta("0 days 00:00:00"):
            POCIO_SEX = OCT_SEX / HPTT_SEX
        else:
            POCIO_SEX = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEX != pd.to_timedelta("0 days 00:00:00"):
            PPROD_SEX = PDTT_SEX / HPTT_SEX
        else:
            PPROD_SEX = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEX != pd.to_timedelta("0 days 00:00:00"):
            PHNR_SEX = HNRT_SEX / HPTT_SEX
        else:
            PHNR_SEX = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SEX != pd.to_timedelta("0 days 00:00:00"):
            PHT_SEX = HPTT_SEX / HPTT_SEX
        else:
            PHT_SEX = pd.to_timedelta("0 days 00:00:00")

        TESTEKX_SEX["pesquisas_diarias_realizadas"] = (
            TESTEKX_SEX["pesquisas_semanais"] * TESTEKX_SEX["percentual_pesquisas"]
        )
        TESTEKX_SEX.replace([float("inf"), -float("inf")], 0, inplace=True)
        TESTEKX_SEX["pesquisas_diarias_realizadas"] = TESTEKX_SEX[
            "pesquisas_diarias_realizadas"
        ].fillna(int(0))
        TESTEKX_SEX["pesquisas_diarias_realizadas"] = TESTEKX_SEX[
            "pesquisas_diarias_realizadas"
        ].astype(int)

        TESTEKX_SEX.loc[
            TESTEKX_SEX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias",
        ] = int(0)
        TESTEKX_SEX.loc[
            TESTEKX_SEX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias_realizadas",
        ] = int(0)
        TESTEKX_SEX.loc[
            TESTEKX_SEX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_diarias_realizadas",
        ] = int(0)
        TESTEKX_SEX.loc[
            TESTEKX_SEX["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_semanais",
        ] = int(0)

        SVS_SEX = TESTEKX_SEX["visitas_diarias"].sum()
        SVSR_SEX = TESTEKX_SEX["visitas_diarias_realizadas"].sum()
        SPS_SEX = TESTEKX_SEX["pesquisas_semanais"].sum()
        SPSR_SEX = TESTEKX_SEX["pesquisas_diarias_realizadas"].sum()

        # Condição para evitar divisão por zero
        if HPTT_SEX != pd.to_timedelta("0 days 00:00:00"):
            MdPROD_SEX = PDTT_SEX / HPTT_SEX
        else:
            MdPROD_SEX = int(0)

        if SVS_SEX != pd.to_timedelta("0 days 00:00:00"):
            MdVST_SEX = SVSR_SEX / SVS_SEX
        else:
            MdVST_SEX = int(0)

        if SPS_SEX != pd.to_timedelta("0 days 00:00:00"):
            MdPSQ_SEX = SPSR_SEX / SPS_SEX
        else:
            MdPSQ_SEX = int(0)

        TESTEKX_SEX = TESTEKX_SEX.rename(
            {"pesquisas_semanais": "pesquisas_diarias"}, axis=1
        )

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha_SEX = pd.DataFrame(
            {
                "Colaborador": ["TOTAL"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PDTT_SEX],
                "DESLOCAMENTO": [DSLCT_SEX],
                "ÓCIO": [OCT_SEX],
                "TOTAL": [TT_SEX],
                "horas_programadas": [HPTT_SEX],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "H N REGISTRADAS": [HNRT_SEX],
                "visitas_diarias": [SVS_SEX],
                "visitas_diarias_realizadas": [SVSR_SEX],
                "pesquisas_diarias": [SPS_SEX],
                "pesquisas_diarias_realizadas": [SPSR_SEX],
                "% produtividade": [MdPROD_SEX],
                "percentual_visitas": [MdVST_SEX],
                "percentual_pesquisas": [MdPSQ_SEX],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        TESTEKX_SEX = pd.concat([TESTEKX_SEX, nova_linha_SEX], ignore_index=True)

        TESTEKX_SEX.loc[TESTEKX_SEX["visitas_diarias"] == 0, "percentual_visitas"] = "-"
        TESTEKX_SEX.loc[
            TESTEKX_SEX["pesquisas_diarias"] == 0, "percentual_pesquisas"
        ] = "-"

        # Convertendo as colunas para tipos numéricos
        TESTEKX_SEX["% produtividade"] = pd.to_numeric(
            TESTEKX_SEX["% produtividade"], errors="coerce"
        )
        TESTEKX_SEX["percentual_visitas"] = pd.to_numeric(
            TESTEKX_SEX["percentual_visitas"], errors="coerce"
        )
        TESTEKX_SEX["percentual_pesquisas"] = pd.to_numeric(
            TESTEKX_SEX["percentual_pesquisas"], errors="coerce"
        )

        # Verificando se há valores não numéricos que foram convertidos para NaN
        if (
            TESTEKX_SEX[
                ["% produtividade", "percentual_visitas", "percentual_pesquisas"]
            ]
            .isnull()
            .values.any()
        ):
            print(
                "Existem valores não numéricos nas colunas que foram convertidos para NaN."
            )
            print(TESTEKX_SEX)

        # Commenting out the actual calculation since the DataFrame is not defined here.
        TESTEKX_SEX["percentual_efetividade"] = (
            TESTEKX_SEX["% produtividade"] * pesos["Produtividade"]
            + TESTEKX_SEX["percentual_visitas"] * pesos["Visitas"]
            + TESTEKX_SEX["percentual_pesquisas"] * pesos["Pesquisas"]
        ) / sum(pesos.values())

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha2_SEX = pd.DataFrame(
            {
                "Colaborador": ["-"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PPROD_SEX],
                "DESLOCAMENTO": [PDESLOC_SEX],
                "ÓCIO": [POCIO_SEX],
                "H N REGISTRADAS": [PHNR_SEX],
                "horas_programadas": [PHT_SEX],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "visitas_diarias": ["-"],
                "visitas_diarias_realizadas": ["-"],
                "pesquisas_diarias": ["-"],
                "pesquisas_diarias_realizadas": ["-"],
                "% produtividade": ["-"],
                "percentual_visitas": ["-"],
                "percentual_pesquisas": ["-"],
                "percentual_efetividade": ["-"],
                "justificativas_sex": ["-"],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        PDOHCOMPLETO_SEX = pd.concat([TESTEKX_SEX, nova_linha2_SEX], ignore_index=True)

        PDOHCOMPLETO_SEX["nome_do_dia"] = "Sexta-feira"
        PDOHCOMPLETO_SEX["data"] = teste_x_SEX

        PDOHCOMPLETO_SEX = PDOHCOMPLETO_SEX.rename(
            {
                "Colaborador": "colaborador",
                "SUPERIOR": "superior",
                "UF": "estado",
                "DESLOCAMENTO": "deslocamento",
                "ÓCIO": "ocio",
                "PRODUTIVIDADE": "produtividade",
                "H N REGISTRADAS": "horas_nao_registradas",
                "horas_programadas": "horas_programadas",
                "Média de tempo em loja": "media_tempo_em_loja",
                "% produtividade": "percentual_produtividade",
                "percentual_visitas": "percentual_visitas",
                "percentual_pesquisas": "percentual_pesquisas",
                "percentual_efetividade": "percentual_efetividade",
            },
            axis=1,
        )

        PDOHCOMPLETO_SEX.loc[
            PDOHCOMPLETO_SEX["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_produtividade",
        ] = "-"
        PDOHCOMPLETO_SEX.loc[
            PDOHCOMPLETO_SEX["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_visitas",
        ] = "-"
        PDOHCOMPLETO_SEX.loc[
            PDOHCOMPLETO_SEX["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_pesquisas",
        ] = "-"
        PDOHCOMPLETO_SEX.loc[
            PDOHCOMPLETO_SEX["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_efetividade",
        ] = "-"

        PDOHCOMPLETO_SEX_XXX = PDOHCOMPLETO_SEX[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_sex",
            ]
        ]

        PDOHCOMPLETO_SEX = PDOHCOMPLETO_SEX[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_sex",
            ]
        ]

        concat_analise.append(PDOHCOMPLETO_SEX_XXX)

    # --------------------------------------------------------------------------------------------------------------------
    # -----------------------------------------------SEXTA-FEIRA----------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------SÁBADO-----------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    if True in dias_da_semana["nome_do_dia"].isin(["Sábado"]).tolist():
        PDOH_SAB = PDOH.loc[PDOH["nome_do_dia"] == "Sábado"]

        def AnalisePDOH(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador", "data"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida  #####################################################

        def AnalisePor(df, filtro):
            selecao = df["ID"] == filtro
            filtrado = df[selecao]
            agrupado = filtrado.groupby(["Colaborador"])
            somado = (agrupado["Tempo Investido"].sum()).to_frame()
            IndexReset = somado.reset_index()
            saida = IndexReset.rename(columns={"Tempo Investido": filtro})
            return saida

        Ocio_SAB = AnalisePDOH(PDOH_SAB, "Ocio")
        print(Ocio_SAB)

        teste_x_SAB = Ocio_SAB.iloc[0, 1]


        #Criando variável para alocar o corte do almoco dentro do Ocio
        ocio_sem_almoco_SAB = Ocio_SAB

        ocio_sem_almoco_SAB = pd.merge(ocio_sem_almoco_SAB, colaboradores_SAB, on="Colaborador", how="right")

        ocio_sem_almoco_SAB = ocio_sem_almoco_SAB[["Colaborador", "Ocio", "almoco"]]

        ocio_sem_almoco_SAB = ocio_sem_almoco_SAB.rename({"almoco":"almoco"}, axis=1)

        ocio_sem_almoco_SAB["Ocio"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        ocio_sem_almoco_SAB["Ocio_sem_almoco"] = ocio_sem_almoco_SAB["Ocio"] - ocio_sem_almoco_SAB["almoco"]

        ocio_sem_almoco_SAB.loc[ocio_sem_almoco_SAB["Ocio_sem_almoco"] < pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = ocio_sem_almoco_SAB["almoco"] - ocio_sem_almoco_SAB["Ocio"]

        ocio_sem_almoco_SAB.loc[ocio_sem_almoco_SAB["Corte_pendente"] <= pd.to_timedelta("0 days 00:00:00"), "Corte_pendente"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_SAB.loc[ocio_sem_almoco_SAB["Ocio_sem_almoco"] > pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = ocio_sem_almoco_SAB["Ocio_sem_almoco"]
        ocio_sem_almoco_SAB.loc[ocio_sem_almoco_SAB["Ocio_sem_almoco"] <= pd.to_timedelta("0 days 00:00:00"), "Ocio_real"] = pd.to_timedelta("0 days 00:00:00")

        ocio_sem_almoco_SAB["Corte_pendente"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Ocio_SAB = ocio_sem_almoco_SAB[["Colaborador", "Ocio_real"]]

        Ocio_SAB.rename({"Ocio_real":"Ocio"}, axis=1, inplace=True)


        # --------------------------------------- DESLOCAMENTO
        Deslocamento_SAB = AnalisePor(PDOH_SAB,"Deslocamento")

        desloc_com_cortes_SAB = pd.merge(Deslocamento_SAB, ocio_sem_almoco_SAB, on="Colaborador", how="right")

        desloc_com_cortes_SAB["Deslocamento_pos_corte"] = desloc_com_cortes_SAB["Deslocamento"] - desloc_com_cortes_SAB["Corte_pendente"]

        desloc_com_cortes_SAB["corte_pendente_prod"] = desloc_com_cortes_SAB["Corte_pendente"] - desloc_com_cortes_SAB["Deslocamento"]

        desloc_com_cortes_SAB.loc[desloc_com_cortes_SAB["corte_pendente_prod"] < pd.to_timedelta("0 days 00:00:00"), "corte_pendente_prod"] = pd.to_timedelta("0 days 00:00:00")

        desloc_com_cortes_SAB.loc[desloc_com_cortes_SAB["Deslocamento_pos_corte"] > pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = desloc_com_cortes_SAB["Deslocamento_pos_corte"]

        desloc_com_cortes_SAB.loc[desloc_com_cortes_SAB["Deslocamento_pos_corte"] <= pd.to_timedelta("0 days 00:00:00"), "Deslocamento"] = pd.to_timedelta("0 days 00:00:00")

        Deslocamento_SAB = desloc_com_cortes_SAB[["Colaborador", "Deslocamento"]]


        Produtividadex_SAB = AnalisePor(PDOH_SAB,"Produtividade")

        Produtividadex_SAB = pd.merge(Produtividadex_SAB, desloc_com_cortes_SAB, how = "right", on = "Colaborador")

        Produtividadex_SAB.loc[Produtividadex_SAB["corte_pendente_prod"] > pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = Produtividadex_SAB["Produtividade"] - Produtividadex_SAB["corte_pendente_prod"]

        Produtividadex_SAB.loc[Produtividadex_SAB["Produtividade"] <= pd.to_timedelta("0 days 00:00:00"), "Produtividade"] = pd.to_timedelta("0 days 00:00:00")

        Produtividadex_SAB.loc[Produtividadex_SAB["Produtividade"] > pd.to_timedelta("1 days 20:00:00"), "Produtividade"] = pd.to_timedelta("1 days 20:00:00")

        Produtividadex_SAB["Produtividade"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Produtividade_SAB = Produtividadex_SAB[["Colaborador", "Produtividade"]]

        htr_SAB = pd.merge(
            Produtividade_SAB, Deslocamento_SAB, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_SAB = pd.merge(
            htr_SAB, Ocio_SAB, how="outer", on="Colaborador"
        )
        HorasTotaisRegistradas_SAB["Tempo Investido"] = (
            HorasTotaisRegistradas_SAB["Produtividade"]
            + HorasTotaisRegistradas_SAB["Deslocamento"]
            + HorasTotaisRegistradas_SAB["Ocio"]
        )

        Horastotaisxxxx_SAB = pd.merge(
            HorasTotaisRegistradas_SAB, colaboradores_SAB, how="right", on="Colaborador"
        )

        Horastotaisxxxx_SAB["justificativas_sab"].fillna(" ", inplace=True)

        Horastotaisxxxx_SAB.fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        Horastotaisxxxx_SAB["HT sem almoco"] = Horastotaisxxxx_SAB["Tempo Investido"]

        JornadaDeTrabalho_SAB = Horastotaisxxxx_SAB[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "HT sem almoco",
                "Jornada de Trabalho",
                "almoco",
                "justificativas_sab",
            ]
        ]

        JornadaDeTrabalho_SAB["Hora Excedente"] = (
            JornadaDeTrabalho_SAB["HT sem almoco"]
            - JornadaDeTrabalho_SAB["Jornada de Trabalho"]
        )

        JornadaDeTrabalho_SAB["Hora Pendente"] = (
            JornadaDeTrabalho_SAB["Jornada de Trabalho"]
            - JornadaDeTrabalho_SAB["HT sem almoco"]
        )

        JornadaDeTrabalho_SAB.loc[
            JornadaDeTrabalho_SAB["Hora Excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Hora Excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        JornadaDeTrabalho_SAB.loc[
            JornadaDeTrabalho_SAB["Hora Pendente"] < pd.to_timedelta("0 days 00:00:00"),
            "Hora Pendente",
        ] = pd.to_timedelta("0 days 00:00:00")

        Horasnaoregistradas_SAB = JornadaDeTrabalho_SAB[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Hora Pendente",
                "almoco",
                "justificativas_sab",
            ]
        ]

        corteszx_SAB = pd.merge(
            JornadaDeTrabalho_SAB, Deslocamento_SAB, how="left", on="Colaborador"
        )

        corteszx_SAB["Deslocamento_cortado"] = (
            corteszx_SAB["Deslocamento"] - corteszx_SAB["Hora Excedente"]
        )

        corteszx_SAB["Deslocamento_cortado_excedente"] = (
            corteszx_SAB["Hora Excedente"] - corteszx_SAB["Deslocamento"]
        )

        corteszx_SAB.loc[
            corteszx_SAB["Deslocamento_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        corteszx_SAB.loc[
            corteszx_SAB["Deslocamento_cortado_excedente"]
            < pd.to_timedelta("0 days 00:00:00"),
            "Deslocamento_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        cortesyy_SAB = pd.merge(corteszx_SAB, Ocio_SAB, how="outer", on="Colaborador")

        cortesyy_SAB["Ocio_cortado"] = (
            cortesyy_SAB["Ocio"] - cortesyy_SAB["Deslocamento_cortado_excedente"]
        )

        cortesyy_SAB["Ocio_cortado_excedente"] = (
            cortesyy_SAB["Deslocamento_cortado_excedente"] - cortesyy_SAB["Ocio"]
        )

        cortesyy_SAB.loc[
            cortesyy_SAB["Ocio_cortado"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado",
        ] = pd.to_timedelta("0 days 00:00:00")
        cortesyy_SAB.loc[
            cortesyy_SAB["Ocio_cortado_excedente"] < pd.to_timedelta("0 days 00:00:00"),
            "Ocio_cortado_excedente",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_SAB = pd.merge(
            cortesyy_SAB, Produtividade_SAB, how="right", on="Colaborador"
        )

        corteszz_SAB["Produtividade_cortada"] = (
            corteszz_SAB["Produtividade"] - corteszz_SAB["Ocio_cortado_excedente"]
        )

        corteszz_SAB.loc[
            corteszz_SAB["Produtividade_cortada"] < pd.to_timedelta("0 days 00:00:00"),
            "Produtividade_cortada",
        ] = pd.to_timedelta("0 days 00:00:00")

        corteszz_SAB["TOTAL DE HR"] = (
            corteszz_SAB["Produtividade_cortada"]
            + corteszz_SAB["Deslocamento_cortado"]
            + corteszz_SAB["Ocio_cortado"]
        )

        cortes_SAB = corteszz_SAB[
            [
                "Colaborador",
                "Colaborador Superior",
                "UF",
                "Produtividade_cortada",
                "Deslocamento_cortado",
                "Ocio_cortado",
                "TOTAL DE HR",
                "Jornada de Trabalho",
                "Hora Pendente",
                "Hora Excedente",
                "almoco",
                "justificativas_sab",
            ]
        ]

        cortes_SAB = cortes_SAB.rename({"Hora Excedente": "HORAS EXCEDENTES"}, axis=1)
        cortes_SAB = cortes_SAB.rename({"almoco": "almoco"}, axis=1)
        cortes_SAB = cortes_SAB.rename({"Colaborador Superior": "SUPERIOR"}, axis=1)

        PDOH_P_SAB = cortes_SAB.rename(
            {
                "Produtividade_cortada": "PRODUTIVIDADE",
                "Deslocamento_cortado": "DESLOCAMENTO",
                "Ocio_cortado": "ÓCIO",
                "TOTAL DE HR": "TOTAL",
                "Hora Pendente": "H N REGISTRADAS",
                "Jornada de Trabalho": "horas_programadas",
            },
            axis=1,
        )

        PDOH_P_SAB = pd.merge(PDOH_P_SAB, checkin_SABD, on=["Colaborador"], how="left")

        ########
        visitasgerencial_SAB = gerencial[
            [
                "Colaborador",
                "Data da Visita",
                "Tipo de Check-in",
                "Tempo em Loja Executado",
            ]
        ]

        visitasgerencial_SAB["Data da Visita"] = pd.to_datetime(
            visitasgerencial_SAB["Data da Visita"], format="%d/%m/%Y"
        )

        dias_ptbr = {
            "Sunday": "Domingo",
            "Monday": "Segunda-feira",
            "Tuesday": "Terça-feira",
            "Wednesday": "Quarta-feira",
            "Thursday": "Quinta-feira",
            "Friday": "Sexta-feira",
            "Saturday": "Sábado",
        }

        visitasgerencial_SAB["nome_do_dia"] = (
            visitasgerencial_SAB["Data da Visita"].dt.day_name().replace(dias_ptbr)
        )

        visitasgerencial_SAB = visitasgerencial_SAB.loc[
            visitasgerencial_SAB["nome_do_dia"] == "Sábado"
        ]

        ########

        visitasgerencial_SAB["visitas_diarias"] = "1"
        visitasgerencial_SAB["percentual_visitas"] = "1"
        visitasgerencial_SAB["MANUAL E GPS"] = "0"
        visitasgerencial_SAB["MANUAL"] = "0"
        visitasgerencial_SAB["SEM CHECKIN"] = "0"

        visitasgerencial_SAB.loc[
            visitasgerencial_SAB["Tipo de Check-in"] == "Sem Checkin",
            "percentual_visitas",
        ] = "0"
        visitasgerencial_SAB.loc[
            visitasgerencial_SAB["Tipo de Check-in"] == "Manual e GPS", "MANUAL E GPS"
        ] = "1"
        visitasgerencial_SAB.loc[
            visitasgerencial_SAB["Tipo de Check-in"] == "Manual", "MANUAL"
        ] = "1"
        visitasgerencial_SAB.loc[
            visitasgerencial_SAB["Tipo de Check-in"] == "Sem Checkin", "SEM CHECKIN"
        ] = "1"

        # transformando em int() os valores das colunas que atribuimos.

        visitasgerencial_SAB["visitas_diarias"] = visitasgerencial_SAB[
            "visitas_diarias"
        ].astype(int)
        visitasgerencial_SAB["percentual_visitas"] = visitasgerencial_SAB[
            "percentual_visitas"
        ].astype(int)
        visitasgerencial_SAB["MANUAL E GPS"] = visitasgerencial_SAB[
            "MANUAL E GPS"
        ].astype(int)
        visitasgerencial_SAB["MANUAL"] = visitasgerencial_SAB["MANUAL"].astype(int)
        visitasgerencial_SAB["SEM CHECKIN"] = visitasgerencial_SAB[
            "SEM CHECKIN"
        ].astype(int)

        porcentagem_visitas_SAB = visitasgerencial_SAB[
            ["Colaborador", "visitas_diarias", "percentual_visitas"]
        ]

        porcentagem_visitas_SAB = porcentagem_visitas_SAB.groupby(["Colaborador"]).agg(
            {"visitas_diarias": "sum", "percentual_visitas": "sum"}
        )

        porcentagem_visitas_SAB["% de Visitas R"] = (
            porcentagem_visitas_SAB["percentual_visitas"]
            / porcentagem_visitas_SAB["visitas_diarias"]
        )

        porcentagem_visitas_SAB = porcentagem_visitas_SAB.rename(
            {
                "visitas_diarias": "visitas_diarias",
                "percentual_visitas": "visitas_diarias_realizadas",
                "% de Visitas R": "percentual_visitas",
            },
            axis=1,
        )

        TESTEKX_SAB = pd.merge(
            PDOH_P_SAB, porcentagem_visitas_SAB, on=["Colaborador"], how="left"
        )

        md_tempoemloja_SAB = visitasgerencial_SAB[
            ["Colaborador", "Tempo em Loja Executado"]
        ]
        md_tempoemloja_SAB["Tempo em Loja Executado"] = pd.to_timedelta(
            md_tempoemloja_SAB["Tempo em Loja Executado"]
        )

        md_tempoemloja_SAB = md_tempoemloja_SAB.groupby(
            ["Colaborador"], as_index=False
        ).agg({"Tempo em Loja Executado": "mean"})

        md_tempoemloja_SAB = md_tempoemloja_SAB.rename(
            {"Tempo em Loja Executado": "Média de tempo em loja"}, axis=1
        )

        TESTEKX_SAB = pd.merge(
            TESTEKX_SAB, md_tempoemloja_SAB, on=["Colaborador"], how="left"
        )

        ########
        pesquisasporcentagem_SAB = pesquisasporcentagem_x[
            ["Responsável", "Status", "Data de Expiração"]
        ]

        pesquisasporcentagem_SAB["Data de Expiração"] = pd.to_datetime(
            pesquisasporcentagem_SAB["Data de Expiração"]
        )

        pesquisasporcentagem_SAB["nome_do_dia"] = (
            pesquisasporcentagem_SAB["Data de Expiração"]
            .dt.day_name()
            .replace(dias_ptbr)
        )

        pesquisasporcentagem_SAB = pesquisasporcentagem_SAB.loc[
            pesquisasporcentagem_SAB["nome_do_dia"] == "Sábado"
        ]

        ########

        pesquisasporcentagem_SAB["pesquisas_semanais"] = "1"
        pesquisasporcentagem_SAB["% de Pesquisasx"] = "1"

        pesquisasporcentagem_SAB.loc[
            pesquisasporcentagem_SAB["Status"] != "Respondida", "% de Pesquisasx"
        ] = "0"

        # transformando em int() os valores das colunas que atribuimos.

        pesquisasporcentagem_SAB["pesquisas_semanais"] = pesquisasporcentagem_SAB[
            "pesquisas_semanais"
        ].astype(int)
        pesquisasporcentagem_SAB["% de Pesquisasx"] = pesquisasporcentagem_SAB[
            "% de Pesquisasx"
        ].astype(int)

        pesquisasporcentagem_SAB = pesquisasporcentagem_SAB.rename(
            {"Responsável": "Colaborador"}, axis=1
        )

        pesquisasporcentagem_SAB = pesquisasporcentagem_SAB[
            ["Colaborador", "pesquisas_semanais", "% de Pesquisasx"]
        ]

        pesquisasporcentagem_SAB = pesquisasporcentagem_SAB.groupby(
            ["Colaborador"]
        ).agg({"pesquisas_semanais": "sum", "% de Pesquisasx": "sum"})

        pesquisasporcentagem_SAB["percentual_pesquisas"] = (
            pesquisasporcentagem_SAB["% de Pesquisasx"]
            / pesquisasporcentagem_SAB["pesquisas_semanais"]
        )

        pesquisasporcentagem_SAB = pesquisasporcentagem_SAB[
            ["pesquisas_semanais", "percentual_pesquisas"]
        ]

        TESTEKX_SAB = pd.merge(
            TESTEKX_SAB, pesquisasporcentagem_SAB, on=["Colaborador"], how="left"
        )

        TESTEKX_SAB["pesquisas_diarias_realizadas"] = TESTEKX_SAB[
            "pesquisas_semanais"
        ].astype(float) * TESTEKX_SAB["percentual_pesquisas"].astype(float)

        TESTEKX_SAB["pesquisas_semanais"].fillna(int(0), inplace=True)
        TESTEKX_SAB["percentual_pesquisas"].fillna(int(0), inplace=True)
        TESTEKX_SAB["pesquisas_diarias_realizadas"].fillna(int(0), inplace=True)

        TESTEKX_SAB["visitas_diarias"].fillna(int(0), inplace=True)
        TESTEKX_SAB["percentual_visitas"].fillna(int(0), inplace=True)

        TESTEKX_SAB["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SAB["DESLOCAMENTO"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SAB["ÓCIO"].fillna(pd.to_timedelta("0 days 00:00:00"), inplace=True)

        # TESTEKX_SAB["PROD_SOPRANO"] = TESTEKX_SAB["PRODUTIVIDADE"] + TESTEKX_SAB["ÓCIO"]

        # TESTEKX_SAB["PRODUTIVIDADE"] = TESTEKX_SAB["PROD_SOPRANO"]

        TESTEKX_SAB["PRODUTIVIDADE"].fillna(
            pd.to_timedelta("0 days 00:00:00"), inplace=True
        )

        TESTEKX_SAB.loc[
            TESTEKX_SAB["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00")),
            "% produtividade",
        ] = (
            TESTEKX_SAB["PRODUTIVIDADE"] / TESTEKX_SAB["horas_programadas"]
        )

        TESTEKX_SAB.loc[
            (TESTEKX_SAB["PRODUTIVIDADE"] == (pd.to_timedelta("0 days 00:00:00")))
            & (
                TESTEKX_SAB["horas_programadas"] != (pd.to_timedelta("0 days 00:00:00"))
            ),
            "% produtividade",
        ] = int(0)

        PDTT_SAB = TESTEKX_SAB["PRODUTIVIDADE"].sum()

        DSLCT_SAB = TESTEKX_SAB["DESLOCAMENTO"].sum()

        OCT_SAB = TESTEKX_SAB["ÓCIO"].sum()

        TT_SAB = TESTEKX_SAB["TOTAL"].sum()

        HPTT_SAB = TESTEKX_SAB["horas_programadas"].sum()

        HNRT_SAB = TESTEKX_SAB["H N REGISTRADAS"].sum()

        # Condição para evitar divisão por zero
        if HPTT_SAB != pd.to_timedelta("0 days 00:00:00"):
            PDESLOC_SAB = DSLCT_SAB / HPTT_SAB
        else:
            PDESLOC_SAB = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SAB != pd.to_timedelta("0 days 00:00:00"):
            POCIO_SAB = OCT_SAB / HPTT_SAB
        else:
            POCIO_SAB = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SAB != pd.to_timedelta("0 days 00:00:00"):
            PPROD_SAB = PDTT_SAB / HPTT_SAB
        else:
            PPROD_SAB = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SAB != pd.to_timedelta("0 days 00:00:00"):
            PHNR_SAB = HNRT_SAB / HPTT_SAB
        else:
            PHNR_SAB = pd.to_timedelta("0 days 00:00:00")

        if HPTT_SAB != pd.to_timedelta("0 days 00:00:00"):
            PHT_SAB = HPTT_SAB / HPTT_SAB
        else:
            PHT_SAB = pd.to_timedelta("0 days 00:00:00")

        TESTEKX_SAB["pesquisas_diarias_realizadas"] = (
            TESTEKX_SAB["pesquisas_semanais"] * TESTEKX_SAB["percentual_pesquisas"]
        )
        TESTEKX_SAB.replace([float("inf"), -float("inf")], 0, inplace=True)
        TESTEKX_SAB["pesquisas_diarias_realizadas"] = TESTEKX_SAB[
            "pesquisas_diarias_realizadas"
        ].fillna(int(0))
        TESTEKX_SAB["pesquisas_diarias_realizadas"] = TESTEKX_SAB[
            "pesquisas_diarias_realizadas"
        ].astype(int)

        TESTEKX_SAB.loc[
            TESTEKX_SAB["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias",
        ] = int(0)
        TESTEKX_SAB.loc[
            TESTEKX_SAB["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "visitas_diarias_realizadas",
        ] = int(0)
        TESTEKX_SAB.loc[
            TESTEKX_SAB["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_diarias_realizadas",
        ] = int(0)
        TESTEKX_SAB.loc[
            TESTEKX_SAB["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
            "pesquisas_semanais",
        ] = int(0)

        SVS_SAB = TESTEKX_SAB["visitas_diarias"].sum()
        SVSR_SAB = TESTEKX_SAB["visitas_diarias_realizadas"].sum()
        SPS_SAB = TESTEKX_SAB["pesquisas_semanais"].sum()
        SPSR_SAB = TESTEKX_SAB["pesquisas_diarias_realizadas"].sum()

        # Condição para evitar divisão por zero
        if HPTT_SAB != pd.to_timedelta("0 days 00:00:00"):
            MdPROD_SAB = PDTT_SAB / HPTT_SAB
        else:
            MdPROD_SAB = int(0)

        if SVS_SAB != pd.to_timedelta("0 days 00:00:00"):
            MdVST_SAB = SVSR_SAB / SVS_SAB
        else:
            MdVST_SAB = int(0)

        if SPS_SAB != pd.to_timedelta("0 days 00:00:00"):
            MdPSQ_SAB = SPSR_SAB / SPS_SAB
        else:
            MdPSQ_SAB = int(0)

        TESTEKX_SAB = TESTEKX_SAB.rename(
            {"pesquisas_semanais": "pesquisas_diarias"}, axis=1
        )

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha_SAB = pd.DataFrame(
            {
                "Colaborador": ["TOTAL"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PDTT_SAB],
                "DESLOCAMENTO": [DSLCT_SAB],
                "ÓCIO": [OCT_SAB],
                "TOTAL": [TT_SAB],
                "horas_programadas": [HPTT_SAB],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "H N REGISTRADAS": [HNRT_SAB],
                "visitas_diarias": [SVS_SAB],
                "visitas_diarias_realizadas": [SVSR_SAB],
                "pesquisas_diarias": [SPS_SAB],
                "pesquisas_diarias_realizadas": [SPSR_SAB],
                "% produtividade": [MdPROD_SAB],
                "percentual_visitas": [MdVST_SAB],
                "percentual_pesquisas": [MdPSQ_SAB],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        TESTEKX_SAB = pd.concat([TESTEKX_SAB, nova_linha_SAB], ignore_index=True)

        TESTEKX_SAB.loc[TESTEKX_SAB["visitas_diarias"] == 0, "percentual_visitas"] = "-"
        TESTEKX_SAB.loc[
            TESTEKX_SAB["pesquisas_diarias"] == 0, "percentual_pesquisas"
        ] = "-"

        # Convertendo as colunas para tipos numéricos
        TESTEKX_SAB["% produtividade"] = pd.to_numeric(
            TESTEKX_SAB["% produtividade"], errors="coerce"
        )
        TESTEKX_SAB["percentual_visitas"] = pd.to_numeric(
            TESTEKX_SAB["percentual_visitas"], errors="coerce"
        )
        TESTEKX_SAB["percentual_pesquisas"] = pd.to_numeric(
            TESTEKX_SAB["percentual_pesquisas"], errors="coerce"
        )

        # Verificando se há valores não numéricos que foram convertidos para NaN
        if (
            TESTEKX_SAB[
                ["% produtividade", "percentual_visitas", "percentual_pesquisas"]
            ]
            .isnull()
            .values.any()
        ):
            print(
                "Existem valores não numéricos nas colunas que foram convertidos para NaN."
            )
            print(TESTEKX_SAB)

        # Commenting out the actual calculation since the DataFrame is not defined here.
        TESTEKX_SAB["percentual_efetividade"] = (
            TESTEKX_SAB["% produtividade"] * pesos["Produtividade"]
            + TESTEKX_SAB["percentual_visitas"] * pesos["Visitas"]
            + TESTEKX_SAB["percentual_pesquisas"] * pesos["Pesquisas"]
        ) / sum(pesos.values())

        # Criar um novo DataFrame com a linha que deseja adicionar
        nova_linha2_SAB = pd.DataFrame(
            {
                "Colaborador": ["-"],
                "SUPERIOR": ["-"],
                "UF": ["-"],
                "PRODUTIVIDADE": [PPROD_SAB],
                "DESLOCAMENTO": [PDESLOC_SAB],
                "ÓCIO": [POCIO_SAB],
                "H N REGISTRADAS": [PHNR_SAB],
                "horas_programadas": [PHT_SAB],
                "almoco": "-",
                "Média de tempo em loja": ["-"],
                "primeiro_checkin": "-",
                "ultimo_checkout": "-",
                "visitas_diarias": ["-"],
                "visitas_diarias_realizadas": ["-"],
                "pesquisas_diarias": ["-"],
                "pesquisas_diarias_realizadas": ["-"],
                "% produtividade": ["-"],
                "percentual_visitas": ["-"],
                "percentual_pesquisas": ["-"],
                "percentual_efetividade": ["-"],
                "justificativas_sab": ["-"],
            }
        )

        # Usar a função concat para adicionar a nova linha ao DataFrame
        PDOHCOMPLETO_SAB = pd.concat([TESTEKX_SAB, nova_linha2_SAB], ignore_index=True)

        PDOHCOMPLETO_SAB["nome_do_dia"] = "Sábado"
        PDOHCOMPLETO_SAB["data"] = teste_x_SAB

        PDOHCOMPLETO_SAB = PDOHCOMPLETO_SAB.rename(
            {
                "Colaborador": "colaborador",
                "SUPERIOR": "superior",
                "UF": "estado",
                "DESLOCAMENTO": "deslocamento",
                "ÓCIO": "ocio",
                "PRODUTIVIDADE": "produtividade",
                "H N REGISTRADAS": "horas_nao_registradas",
                "horas_programadas": "horas_programadas",
                "Média de tempo em loja": "media_tempo_em_loja",
                "% produtividade": "percentual_produtividade",
                "percentual_visitas": "percentual_visitas",
                "percentual_pesquisas": "percentual_pesquisas",
                "percentual_efetividade": "percentual_efetividade",
            },
            axis=1,
        )

        PDOHCOMPLETO_SAB.loc[
            PDOHCOMPLETO_SAB["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_produtividade",
        ] = "-"
        PDOHCOMPLETO_SAB.loc[
            PDOHCOMPLETO_SAB["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_visitas",
        ] = "-"
        PDOHCOMPLETO_SAB.loc[
            PDOHCOMPLETO_SAB["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_pesquisas",
        ] = "-"
        PDOHCOMPLETO_SAB.loc[
            PDOHCOMPLETO_SAB["horas_programadas"]
            == (pd.to_timedelta("0 days 00:00:00")),
            "percentual_efetividade",
        ] = "-"

        PDOHCOMPLETO_SAB_XXX = PDOHCOMPLETO_SAB[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_sab",
            ]
        ]

        PDOHCOMPLETO_SAB = PDOHCOMPLETO_SAB[
            [
                "colaborador",
                "superior",
                "estado",
                "data",
                "nome_do_dia",
                "deslocamento",
                "ocio",
                "produtividade",
                "horas_nao_registradas",
                "horas_programadas",
                "almoco",
                "media_tempo_em_loja",
                "primeiro_checkin",
                "ultimo_checkout",
                "visitas_diarias",
                "visitas_diarias_realizadas",
                "pesquisas_diarias",
                "pesquisas_diarias_realizadas",
                "percentual_produtividade",
                "percentual_visitas",
                "percentual_pesquisas",
                "percentual_efetividade",
                "justificativas_sab",
            ]
        ]

        concat_analise.append(PDOHCOMPLETO_SAB_XXX)

    # --------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------SÁBADO-----------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------------

    PDOH_ANALISE = pd.concat(concat_analise)

    PDOH_ANALISE = PDOH_ANALISE.sort_values(["colaborador"], ascending=True)

    # Criar uma máscara booleana identificando as linhas que contêm "exemplo"
    mask = PDOH_ANALISE["colaborador"].str.contains("TOTAL", "-")

    # Filtrar o DataFrame para manter apenas as linhas que não contêm "exemplo"
    PDOH_ANALISE = PDOH_ANALISE.loc[~mask]

    PDOH_ANALISE = PDOH_ANALISE.loc[PDOH_ANALISE["colaborador"] != "-"]

    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "produtividade",
    ] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "deslocamento",
    ] = ""
    # PDOH_ANALISE.loc[PDOH_ANALISE["horas_programadas"]==(pd.to_timedelta("0 days 00:00:00")), "ocio"] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "horas_nao_registradas",
    ] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "almoco",
    ] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "visitas_diarias",
    ] = int(0)
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "visitas_diarias_realizadas",
    ] = int(0)
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "pesquisas_diarias_realizadas",
    ] = int(0)
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "pesquisas_diarias",
    ] = int(0)
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_produtividade",
    ] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_visitas",
    ] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_pesquisas",
    ] = ""
    PDOH_ANALISE.loc[
        PDOH_ANALISE["horas_programadas"] == (pd.to_timedelta("0 days 00:00:00")),
        "percentual_efetividade",
    ] = ""
    # PDOH_ANALISE.loc[PDOH_ANALISE["horas_programadas"]==(pd.to_timedelta("0 days 00:00:00")), "justificativas"] = ""

    # PDOH_ANALISE.to_excel(f"outputs/PDOH_ANALISE {nome_do_cliente} {primeira_data_formatada} à {ultima_data_formatada}.xlsx", sheet_name="PDOH_ANALISE", index=False)

    # Usar strftime para formatar as datas
    PDOH_ANALISE["data"] = PDOH_ANALISE["data"].dt.strftime("%d/%m/%Y")

    # Salvando o arquivo PDOH_ANALISE
    pdoh_analise_path = f"outputs/PDOH_ANALISE {nome_do_cliente} {primeira_data_formatada} à {ultima_data_formatada}.xlsx"
    PDOH_ANALISE.to_excel(pdoh_analise_path, sheet_name="PDOH_ANALISE", index=False)

    # Salvando o arquivo PDOH corretamente
    output_path = f"outputs/PDOH {nome_do_cliente} {primeira_data_formatada} à {ultima_data_formatada}.xlsx"
    writer = pd.ExcelWriter(
        output_path, engine="openpyxl"
    )  # Correção para usar a variável corretamente

    PDOHCOMPLETO.to_excel(writer, sheet_name="PDOH SEMANAL", index=False)
    if True in dias_da_semana["nome_do_dia"].isin(["Segunda-feira"]).tolist():
        PDOHCOMPLETO_SEG.to_excel(writer, sheet_name="PDOH SEG", index=False)
    if True in dias_da_semana["nome_do_dia"].isin(["Terça-feira"]).tolist():
        PDOHCOMPLETO_TER.to_excel(writer, sheet_name="PDOH TER", index=False)
    if True in dias_da_semana["nome_do_dia"].isin(["Quarta-feira"]).tolist():
        PDOHCOMPLETO_QUA.to_excel(writer, sheet_name="PDOH QUA", index=False)
    if True in dias_da_semana["nome_do_dia"].isin(["Quinta-feira"]).tolist():
        PDOHCOMPLETO_QUI.to_excel(writer, sheet_name="PDOH QUI", index=False)
    if True in dias_da_semana["nome_do_dia"].isin(["Sexta-feira"]).tolist():
        PDOHCOMPLETO_SEX.to_excel(writer, sheet_name="PDOH SEX", index=False)
    if True in dias_da_semana["nome_do_dia"].isin(["Sábado"]).tolist():
        PDOHCOMPLETO_SAB.to_excel(writer, sheet_name="PDOH SAB", index=False)

    # Salvando o arquivo Excel PDOH corretamente
    writer.close()

    # Função para formatar as abas do Excel
    def format_sheet(sheet):
        # Colunas para formatar conforme o tipo
        time_format_columns = [
            "deslocamento",
            "ocio",
            "produtividade",
            "horas_nao_registradas",
            "horas_programadas",
            "media_tempo_em_loja",
            "almoco",
        ]
        hour_format_columns = [
            "media_primeiro_checkin",
            "media_ultimo_checkout",
            "primeiro_checkin_sab",
            "ultimo_checkout_sab",
            "primeiro_checkin",
            "ultimo_checkout",
        ]
        # percentage_format_columns = [
        #     "percentual_produtividade",
        #     "percentual_visitas",
        #     "percentual_pesquisas",
        #     "percentual_efetividade",
        # ]
        date_format_columns = ["data"]  # Nova coluna para formato de data

        header_fill = PatternFill(
            start_color="FF000080", end_color="FF000080", fill_type="solid"
        )
        header_font = Font(color="FFFFFFFF", bold=True)

        max_length = {}
        headers = {}
        for cell in sheet[1]:
            headers[cell.value] = cell.column_letter
            max_length[cell.column_letter] = len(cell.value)
            cell.fill = header_fill
            cell.font = header_font

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                column_letter = cell.column_letter
                if len(str(cell.value)) > max_length.get(column_letter, 0):
                    max_length[column_letter] = len(str(cell.value))

                if column_letter in [headers.get(col) for col in time_format_columns]:
                    cell.number_format = "[h]:mm:ss"
                elif column_letter in [headers.get(col) for col in hour_format_columns]:
                    cell.number_format = "hh:mm:ss"
                # elif column_letter in [
                #     headers.get(col) for col in percentage_format_columns
                # ]:
                #     cell.number_format = "0%"
                elif column_letter in [headers.get(col) for col in date_format_columns]:
                    cell.number_format = "dd/mm/yyyy"  # Formato de data
                cell.alignment = Alignment(horizontal="center")

        for col, length in max_length.items():
            sheet.column_dimensions[col].width = length + 2

    # Função para formatar e salvar múltiplos arquivos
    def format_and_save_workbooks(paths):
        for path in paths:
            workbook = openpyxl.load_workbook(path)
            for sheet in workbook.worksheets:
                format_sheet(sheet)
            workbook.save(path)

    # Lista de caminhos dos arquivos a serem formatados
    output_paths = [pdoh_analise_path, output_path]

    # Formatando e salvando ambos os arquivos
    format_and_save_workbooks(output_paths)

    # # Certificando-se de fechar corretamente os arquivos antes de deletar
    # for path in output_paths:
    #     if os.path.exists(path):
    #         try:
    #             os.remove(path)
    #         except PermissionError:
    #             print(f"O arquivo {path} ainda está em uso e não pode ser deletado.")


except Exception as e:
    # Captura a exceção e registra o erro em um arquivo de texto
    error_message = traceback.format_exc()
    with open("Erro na automação.txt", "w") as file:
        file.write(error_message)
